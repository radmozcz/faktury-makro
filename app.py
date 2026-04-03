"""
Aplikace pro správu přijatých faktur
Spuštění: python app.py
"""

import os
import json
import sqlite3
try:
    import psycopg2
    import psycopg2.extras
except ImportError:
    psycopg2 = None
import csv
import io
import re
import base64
import anthropic

# Google Cloud Storage
try:
    from google.cloud import storage as gcs_storage
    from google.oauth2 import service_account
    GCS_SUPPORT = True
except ImportError:
    GCS_SUPPORT = False
    print("⚠  google-cloud-storage není nainstalován – GCS nebude fungovat")
from datetime import datetime, date, timedelta
from flask import Flask, render_template, request, jsonify, send_file, redirect, url_for, send_from_directory, session
from werkzeug.utils import secure_filename
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

try:
    import pdfplumber
    PDF_SUPPORT = True
except ImportError:
    PDF_SUPPORT = False
    print("⚠  pdfplumber není nainstalován – PDF parsing nebude fungovat")

try:
    import pytesseract
    from PIL import Image
    import os as _os
    _tess_path = r"C:\Program Files\Tesseract-OCR\	esseract.exe"
    if _os.path.exists(_tess_path):
        pytesseract.pytesseract.tesseract_cmd = _tess_path
    OCR_SUPPORT = True
except ImportError:
    OCR_SUPPORT = False
    print("⚠  pytesseract/Pillow není nainstalován – OCR obrázků nebude fungovat")

BASE_DIR    = os.path.dirname(os.path.abspath(__file__))
DB_PATH     = os.path.join(BASE_DIR, "faktury.db")
UPLOAD_DIR  = os.path.join(BASE_DIR, "uploads")
CONFIG_PATH = os.path.join(BASE_DIR, "config.json")
ALLOWED_EXT = {"pdf", "png", "jpg", "jpeg", "tiff", "bmp"}

os.makedirs(UPLOAD_DIR, exist_ok=True)

def get_gcs_client():
    """Vrátí GCS bucket nebo None pokud není nakonfigurováno."""
    if not GCS_SUPPORT:
        return None
    creds_json = os.environ.get("GCS_CREDENTIALS_JSON", "")
    bucket_name = os.environ.get("GCS_BUCKET_NAME", "")
    if not creds_json or not bucket_name:
        return None
    try:
        creds_info = json.loads(creds_json)
        creds = service_account.Credentials.from_service_account_info(creds_info)
        client = gcs_storage.Client(credentials=creds, project=creds_info.get("project_id"))
        return client.bucket(bucket_name)
    except Exception as e:
        print(f"⚠  GCS init error: {e}")
        return None

def upload_to_gcs(local_path, filename):
    """Nahraje soubor do GCS a vrátí signed URL (platné 7 dní) nebo None."""
    bucket = get_gcs_client()
    if not bucket:
        return None
    try:
        blob = bucket.blob(f"faktury/{filename}")
        blob.upload_from_filename(local_path)
        # Signed URL platné 7 dní
        url = blob.generate_signed_url(
            expiration=timedelta(days=7),
            method="GET",
            version="v4"
        )
        return url
    except Exception as e:
        print(f"⚠  GCS upload error: {e}")
        return None

def get_gcs_url(filename):
    """Vrátí čerstvé signed URL pro existující soubor v GCS."""
    bucket = get_gcs_client()
    if not bucket:
        return None
    try:
        blob = bucket.blob(f"faktury/{filename}")
        if not blob.exists():
            return None
        return blob.generate_signed_url(
            expiration=timedelta(days=7),
            method="GET",
            version="v4"
        )
    except Exception as e:
        print(f"⚠  GCS url error: {e}")
        return None

app = Flask(__name__)
app.config["MAX_CONTENT_LENGTH"] = 50 * 1024 * 1024
app.secret_key = os.environ.get("SECRET_KEY", "bistro-tajny-klic-2024-zmen-me")

# ── Přihlašování ────────────────────────────────────────────────────────────────
# Role: admin, verunka, ucetni
ROLE_NAMES = {
    "admin":   "ADMIN",
    "verunka": "VERUNKA",
    "ucetni":  "UCETNI",
}

# Výchozí oprávnění (co smí kdo vidět/dělat)
# Klíče odpovídají sekcím v aplikaci
DEFAULT_PRAVA = {
    "verunka": {
        "faktury_zobrazit":  True,
        "faktury_upravit":   True,
        "faktury_smazat":    False,
        "faktury_export":    True,
        "reporty_zobrazit":  True,
        "reporty_upravit":   True,
        "vyplaty_zobrazit":  True,
        "vyplaty_upravit":   False,
        "zbozi_zobrazit":    True,
        "vydaje_zobrazit":              True,
        "vydaje_upravit":               True,
        "vydaje_smazat":                False,
        "soukrome_vydaje_zobrazit":     False,
        "soukrome_vydaje_upravit":      False,
        "soukrome_vydaje_smazat":       False,
        "naklady_zobrazit":             False,
        "bankovni_vypisy":              False,
        "banky_soukrome":               False,
        "statistiky":                   False,
        "nastaveni":                    False,
        "vystavene_zobrazit":           False,
        "vystavene_upravit":            False,
        "kalkulace":                    False,
        "upozorneni":                   False,
        "nastenka":                     False,
        "radek_sekce":                  False,
    },
    "ucetni": {
        "faktury_zobrazit":  True,
        "faktury_upravit":   False,
        "faktury_smazat":    False,
        "faktury_export":    True,
        "reporty_zobrazit":  False,
        "reporty_upravit":   False,
        "vyplaty_zobrazit":  False,
        "vyplaty_upravit":   False,
        "zbozi_zobrazit":    False,
        "vydaje_zobrazit":              True,
        "vydaje_upravit":               False,
        "vydaje_smazat":                False,
        "soukrome_vydaje_zobrazit":     False,
        "soukrome_vydaje_upravit":      False,
        "soukrome_vydaje_smazat":       False,
        "naklady_zobrazit":             True,
        "bankovni_vypisy":              True,
        "banky_soukrome":               False,
        "statistiky":                   False,
        "nastaveni":                    False,
        "vystavene_zobrazit":           True,
        "vystavene_upravit":            False,
        "kalkulace":                    False,
        "upozorneni":                   False,
        "nastenka":                     False,
        "radek_sekce":                  False,
    },
}

def get_prava_z_db():
    """Načte matici oprávnění z databáze, nebo vrátí výchozí."""
    try:
        with get_db() as conn:
            cur = conn.execute("SELECT role, sekce, povoleno FROM prava")
            rows = cur.fetchall()
        if not rows:
            return DEFAULT_PRAVA.copy()
        prava = {"verunka": {}, "ucetni": {}}
        for r in rows:
            role = r["role"] if isinstance(r, dict) else r[0]
            sekce = r["sekce"] if isinstance(r, dict) else r[1]
            povoleno = r["povoleno"] if isinstance(r, dict) else r[2]
            if role in prava:
                prava[role][sekce] = bool(povoleno)
        # Doplnit chybějící klíče výchozími hodnotami
        for role in ["verunka", "ucetni"]:
            for sekce, val in DEFAULT_PRAVA[role].items():
                if sekce not in prava[role]:
                    prava[role][sekce] = val
        return prava
    except Exception:
        return DEFAULT_PRAVA.copy()

def ma_pravo(sekce):
    """Zkontroluje zda přihlášený uživatel má právo na danou sekci."""
    role = session.get("role", "")
    if role == "admin":
        return True
    prava = get_prava_z_db()
    return prava.get(role, {}).get(sekce, False)

def vyzaduj_prihlaseni(f):
    """Dekorátor – vrátí 401 pokud uživatel není přihlášen."""
    from functools import wraps
    @wraps(f)
    def wrapper(*args, **kwargs):
        if not session.get("role"):
            return jsonify({"error": "Nejsi přihlášen", "login_required": True}), 401
        return f(*args, **kwargs)
    return wrapper

DEFAULT_CONFIG = {
    "firmy": ["FP", "MR", "CFF"],
    "app_nazev": "Správa faktur",
    "ico_map": {},
    "terminal_limit": 100000,
    "dph_limit": 2000000,
    "terminal_aktivni": {},
    "terminal_od": {}
}

def load_config():
    if os.path.exists(CONFIG_PATH):
        with open(CONFIG_PATH, "r", encoding="utf-8") as f:
            return json.load(f)
    return DEFAULT_CONFIG.copy()

def save_config(cfg):
    with open(CONFIG_PATH, "w", encoding="utf-8") as f:
        json.dump(cfg, f, ensure_ascii=False, indent=2)

DATABASE_URL = os.environ.get("DATABASE_URL", "")
_USE_PG = bool(DATABASE_URL)


def _first_val(row):
    """Vrátí první hodnotu z řádku – funguje pro dict (PG) i tuple/Row (SQLite)."""
    if row is None:
        return 0
    if isinstance(row, dict):
        v = list(row.values())[0]
    else:
        v = row[0]
    return v if v is not None else 0


def _row_to_dict(row):
    """Převede DB řádek na JSON-serializovatelný dict (ošetří Decimal, None atd.)."""
    from decimal import Decimal
    if isinstance(row, dict):
        d = row
    else:
        try:
            d = dict(row)
        except Exception:
            return {}
    result = {}
    for k, v in d.items():
        if isinstance(v, Decimal):
            result[k] = float(v)
        elif v is None:
            result[k] = None
        else:
            result[k] = v
    return result


def _safe_json(rows):
    """Serializuje seznam DB řádků do JSON bezpečně."""
    import json as _json
    from decimal import Decimal
    class _Enc(_json.JSONEncoder):
        def default(self, o):
            if isinstance(o, Decimal): return float(o)
            return super().default(o)
    return _json.dumps([_row_to_dict(r) for r in rows], ensure_ascii=False, indent=2, cls=_Enc)


class _PgCursor:
    def __init__(self, cur, is_insert=False):
        self._cur = cur
        self._lastrowid = None
        if is_insert:
            try:
                r = self._cur.fetchone()
                if r is not None:
                    try:
                        self._lastrowid = r["id"]
                    except (KeyError, TypeError):
                        try:
                            self._lastrowid = r[0]
                        except Exception:
                            self._lastrowid = None
            except Exception:
                self._lastrowid = None

    def __iter__(self): return iter(self._cur)
    def fetchall(self): return [dict(r) for r in self._cur.fetchall()]
    def fetchone(self):
        r = self._cur.fetchone()
        return dict(r) if r else None

    @property
    def lastrowid(self):
        return self._lastrowid

    @property
    def rowcount(self): return self._cur.rowcount

class _PgConn:
    def __init__(self, conn): self._conn = conn
    def __enter__(self): return self
    def __exit__(self, exc_type, *_):
        if exc_type: self._conn.rollback()
        else: self._conn.commit()
        self._conn.close()
    def commit(self): self._conn.commit()
    def rollback(self): self._conn.rollback()
    def close(self): self._conn.close()
    @staticmethod
    def _adapt(sql):
        import re as _re
        sql = sql.replace("INTEGER PRIMARY KEY AUTOINCREMENT", "SERIAL PRIMARY KEY")
        sql = sql.replace("DEFAULT (datetime('now','localtime'))", "DEFAULT NOW()")
        sql = sql.replace("datetime('now','localtime')", "NOW()")
        sql = sql.replace("date('now','-12 months')", "(CURRENT_DATE - INTERVAL '12 months')")
        sql = sql.replace("date('now')", "CURRENT_DATE::text")
        # datum_vystaveni a datum jsou TEXT sloupce – při porovnání s datem je nutný cast
        # Ale nenahrazujeme porovnání s prázdným stringem (datum > '')
        ssql = _re.sub(r"\bdatum_vystaveni\b(\s*)(>=|<=|>|<)(?!\s*'')", r"NULLIF(datum_vystaveni,'')::date\1\2", sql)
        sql = _re.sub(r"\bdatum\b(\s*)(>=|<=|>|<)(?!\s*'')", r"NULLIF(datum,'')::date\1\2", sql)
        sql = _re.sub(r"strftime\('%Y',\s*([^,)]+)\)", r"TO_CHAR(NULLIF(\1,'')::date, 'YYYY')", sql)
        sql = _re.sub(r"strftime\('%m',\s*([^,)]+)\)", r"TO_CHAR(NULLIF(\1,'')::date, 'MM')", sql)
        sql = _re.sub(r"strftime\('%Y-%m',\s*([^,)]+)\)", r"TO_CHAR(NULLIF(\1,'')::date, 'YYYY-MM')", sql)
        return sql
    def execute(self, sql, params=()):
        if sql.strip().upper().startswith("PRAGMA"):
            class _D:
                lastrowid=None; rowcount=0
                def fetchone(self): return None
                def fetchall(self): return []
                def __iter__(self): return iter([])
            return _D()
        sql = self._adapt(sql)
        sql_pg = sql.replace("?", "%s")
        cur = self._conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
        is_insert = sql_pg.strip().upper().startswith("INSERT")
        if is_insert and "RETURNING" not in sql_pg.upper():
            sql_pg = sql_pg.rstrip().rstrip(";") + " RETURNING id"
        cur.execute(sql_pg, params if params else None)
        return _PgCursor(cur, is_insert=is_insert)
    def executescript(self, sql):
        sql = self._adapt(sql)
        sql = sql.replace("PRAGMA journal_mode=WAL;", "").replace("PRAGMA foreign_keys=ON;", "")
        cur = self._conn.cursor()
        cur.execute(sql)
        self._conn.commit()

def get_db():
    if _USE_PG:
        conn = psycopg2.connect(DATABASE_URL)
        conn.autocommit = False
        return _PgConn(conn)
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA journal_mode=WAL")
    conn.execute("PRAGMA foreign_keys=ON")
    return conn


def init_db():
    TABLES = [
        ("zbozi", """CREATE TABLE IF NOT EXISTS zbozi (
            id               SERIAL PRIMARY KEY,
            nazev_canonical  TEXT    NOT NULL UNIQUE,
            poznamka         TEXT
        )"""),
        ("zbozi_aliasy", """CREATE TABLE IF NOT EXISTS zbozi_aliasy (
            id        SERIAL PRIMARY KEY,
            zbozi_id  INTEGER NOT NULL REFERENCES zbozi(id) ON DELETE CASCADE,
            alias     TEXT    NOT NULL UNIQUE
        )"""),
        ("faktury", """CREATE TABLE IF NOT EXISTS faktury (
            id              SERIAL PRIMARY KEY,
            firma_zkratka   TEXT    NOT NULL,
            dodavatel       TEXT    NOT NULL,
            cislo_faktury   TEXT,
            datum_vystaveni TEXT,
            datum_splatnosti TEXT,
            zpusob_uhrady   TEXT,
            stav            TEXT    DEFAULT 'ceka',
            celkem_s_dph    REAL    DEFAULT 0,
            soubor_cesta    TEXT,
            soubor_url      TEXT,
            zdroj           TEXT    DEFAULT 'rucni',
            created_at      TEXT    DEFAULT NOW()
        )"""),
        ("polozky", """CREATE TABLE IF NOT EXISTS polozky (
            id                    SERIAL PRIMARY KEY,
            faktura_id            INTEGER NOT NULL REFERENCES faktury(id) ON DELETE CASCADE,
            nazev                 TEXT    NOT NULL,
            mnozstvi              REAL    DEFAULT 1,
            jednotka              TEXT    DEFAULT 'ks',
            cena_za_jednotku_s_dph REAL   DEFAULT 0,
            celkem_s_dph          REAL    DEFAULT 0,
            zbozi_id              INTEGER REFERENCES zbozi(id) ON DELETE SET NULL
        )"""),
        ("vyplaty", """CREATE TABLE IF NOT EXISTS vyplaty (
            id          SERIAL PRIMARY KEY,
            jmeno       TEXT    NOT NULL,
            datum       TEXT    NOT NULL,
            castka      REAL    NOT NULL DEFAULT 0,
            poznamka    TEXT,
            firma_zkratka TEXT  DEFAULT '',
            created_at  TEXT    DEFAULT NOW()
        )"""),
        ("reporty", """CREATE TABLE IF NOT EXISTS reporty (
            id            SERIAL PRIMARY KEY,
            datum         TEXT    NOT NULL UNIQUE,
            den           TEXT,
            smena         TEXT,
            karty         REAL    DEFAULT 0,
            kov           REAL    DEFAULT 0,
            papir         REAL    DEFAULT 0,
            hotovost      REAL    DEFAULT 0,
            vydaje        REAL    DEFAULT 0,
            trzba         REAL    DEFAULT 0,
            trzba_vcpk    REAL    DEFAULT 0,
            pk50_ks       INTEGER DEFAULT 0,
            pk100_ks      INTEGER DEFAULT 0,
            pk_celkem     REAL    DEFAULT 0,
            pizza_cela    INTEGER DEFAULT 0,
            pizza_ctvrt   INTEGER DEFAULT 0,
            burger        INTEGER DEFAULT 0,
            talire        INTEGER DEFAULT 0,
            burtgulas     INTEGER DEFAULT 0,
            hotdog        INTEGER DEFAULT 0,
            snidane       INTEGER DEFAULT 0,
            nakupy        INTEGER DEFAULT 0,
            foto_cesta    TEXT,
            firma_zkratka TEXT    DEFAULT '',
            poznamka      TEXT,
            created_at    TEXT    DEFAULT NOW()
        )"""),
        ("prava", """CREATE TABLE IF NOT EXISTS prava (
            id      SERIAL PRIMARY KEY,
            role    TEXT NOT NULL,
            sekce   TEXT NOT NULL,
            povoleno INTEGER DEFAULT 0,
            UNIQUE(role, sekce)
        )"""),
        ("pausalni_odvody", """CREATE TABLE IF NOT EXISTS pausalni_odvody (
            id          SERIAL PRIMARY KEY,
            jmeno       TEXT NOT NULL,
            nazev       TEXT NOT NULL,
            castka      REAL NOT NULL DEFAULT 0,
            poradi      INTEGER DEFAULT 0,
            platnost_od TEXT DEFAULT '2020-01-01'
        )"""),
        ("bankovni_pohyby", """CREATE TABLE IF NOT EXISTS bankovni_pohyby (
            id              SERIAL PRIMARY KEY,
            banka           TEXT NOT NULL,
            datum           TEXT NOT NULL,
            castka          REAL NOT NULL DEFAULT 0,
            protiucet       TEXT DEFAULT '',
            nazev_protiucet TEXT DEFAULT '',
            typ_transakce   TEXT DEFAULT '',
            zprava          TEXT DEFAULT '',
            id_transakce    TEXT UNIQUE,
            firma_zkratka   TEXT DEFAULT '',
            created_at      TEXT DEFAULT NOW()
        )"""),
        ("vydaje", """CREATE TABLE IF NOT EXISTS vydaje (
            id              SERIAL PRIMARY KEY,
            firma_zkratka   TEXT NOT NULL,
            dodavatel       TEXT DEFAULT '',
            datum           TEXT DEFAULT '',
            datum_splatnosti TEXT DEFAULT '',
            castka          REAL NOT NULL DEFAULT 0,
            zpusob_uhrady   TEXT DEFAULT 'hotovost',
            stav            TEXT DEFAULT 'nezaplaceno',
            popis           TEXT DEFAULT '',
            poznamka        TEXT DEFAULT '',
            soubor_cesta    TEXT DEFAULT '',
            soubor_url      TEXT DEFAULT '',
            zdroj           TEXT DEFAULT 'rucni',
            typ             TEXT DEFAULT 'provozni',
            created_at      TEXT DEFAULT NOW()
        )"""),
        ("vydaje_polozky", """CREATE TABLE IF NOT EXISTS vydaje_polozky (
            id          SERIAL PRIMARY KEY,
            vydaj_id    INTEGER NOT NULL,
            nazev       TEXT NOT NULL,
            castka      REAL NOT NULL DEFAULT 0
        )"""),
        ("vystavene_faktury", """CREATE TABLE IF NOT EXISTS vystavene_faktury (
            id                SERIAL PRIMARY KEY,
            firma_zkratka     TEXT    NOT NULL,
            cislo_faktury     TEXT    DEFAULT '',
            datum             TEXT    DEFAULT '',
            datum_splatnosti  TEXT    DEFAULT '',
            odberatel         TEXT    DEFAULT '',
            popis             TEXT    DEFAULT '',
            castka            REAL    NOT NULL DEFAULT 0,
            stav              TEXT    DEFAULT 'nezaplaceno',
            soubor_url        TEXT    DEFAULT '',
            duplicita_id      INTEGER DEFAULT NULL,
            created_at        TEXT    DEFAULT NOW()
        )"""),
    ]
    with get_db() as conn:
        for name, sql in TABLES:
            if not _USE_PG:
                sql = sql.replace('SERIAL PRIMARY KEY', 'INTEGER PRIMARY KEY AUTOINCREMENT')
                sql = sql.replace('DEFAULT NOW()', "DEFAULT (datetime('now','localtime'))")
            conn.execute(sql)
    print("init_db OK")


def migrate_db():
    # Migrace: obdobi_od, obdobi_do ve vyplatach
    with get_db() as conn:
        if _USE_PG:
            cur = conn.execute("SELECT column_name FROM information_schema.columns WHERE table_name='vyplaty'")
            vypl_cols = [r["column_name"] for r in cur.fetchall()]
        else:
            vypl_cols = [row[1] for row in conn.execute("PRAGMA table_info(vyplaty)").fetchall()]
        for col, typ in [("obdobi_od","TEXT"), ("obdobi_do","TEXT")]:
            if col not in vypl_cols:
                try: conn.execute(f"ALTER TABLE vyplaty ADD COLUMN {col} {typ}")
                except Exception: pass

    # Odebrat UNIQUE constraint na datum v reporty – samostatná transakce
    if _USE_PG:
        try:
            with get_db() as conn2:
                row = conn2.execute("""
                    SELECT constraint_name FROM information_schema.table_constraints
                    WHERE table_name='reporty' AND constraint_type='UNIQUE'
                    AND constraint_name LIKE '%datum%'
                """).fetchone()
                if row:
                    cname = row["constraint_name"] if isinstance(row, dict) else row[0]
                    conn2.execute(f"ALTER TABLE reporty DROP CONSTRAINT {cname}")
        except Exception: pass

    with get_db() as conn:
        if _USE_PG:
            cur = conn.execute("SELECT column_name FROM information_schema.columns WHERE table_name='reporty'")
            existing = [r["column_name"] for r in cur.fetchall()]
        else:
            existing = [row[1] for row in conn.execute("PRAGMA table_info(reporty)").fetchall()]
        for col, typ in [
            ("burtgulas","INTEGER DEFAULT 0"),("hotdog","INTEGER DEFAULT 0"),
            ("snidane","INTEGER DEFAULT 0"),("nakupy","INTEGER DEFAULT 0"),
            ("foto_cesta","TEXT"),("firma_zkratka","TEXT DEFAULT ''"),
            ("soubor_url","TEXT"),("duplicita_id","INTEGER"),
        ]:
            if col not in existing:
                try: conn.execute(f"ALTER TABLE reporty ADD COLUMN {col} {typ}")
                except Exception: pass
        if _USE_PG:
            cur2 = conn.execute("SELECT column_name FROM information_schema.columns WHERE table_name='faktury'")
            fakt_cols = [r["column_name"] for r in cur2.fetchall()]
        else:
            fakt_cols = [row[1] for row in conn.execute("PRAGMA table_info(faktury)").fetchall()]
        if "duplicita_id" not in fakt_cols:
            try: conn.execute("ALTER TABLE faktury ADD COLUMN duplicita_id INTEGER")
            except Exception: pass
        if "soubor_url" not in fakt_cols:
            try: conn.execute("ALTER TABLE faktury ADD COLUMN soubor_url TEXT")
            except Exception: pass
        # Migrace vydaje
        if _USE_PG:
            cur3 = conn.execute("SELECT column_name FROM information_schema.columns WHERE table_name='vydaje'")
            vydaj_cols = [r["column_name"] for r in cur3.fetchall()]
        else:
            vydaj_cols = [row[1] for row in conn.execute("PRAGMA table_info(vydaje)").fetchall()]
        if "popis" not in vydaj_cols:
            try: conn.execute("ALTER TABLE vydaje ADD COLUMN popis TEXT DEFAULT ''")
            except Exception: pass
        if "stav" not in vydaj_cols:
            try: conn.execute("ALTER TABLE vydaje ADD COLUMN stav TEXT DEFAULT 'nezaplaceno'")
            except Exception: pass
        if "datum_splatnosti" not in vydaj_cols:
            try: conn.execute("ALTER TABLE vydaje ADD COLUMN datum_splatnosti TEXT DEFAULT ''")
            except Exception: pass
        if "datum_uhrady" not in vydaj_cols:
            try: conn.execute("ALTER TABLE vydaje ADD COLUMN datum_uhrady TEXT DEFAULT ''")
            except Exception: pass
        if "banka_uhrady" not in vydaj_cols:
            try: conn.execute("ALTER TABLE vydaje ADD COLUMN banka_uhrady TEXT DEFAULT ''")
            except Exception: pass
        if "typ" not in vydaj_cols:
            try: conn.execute("ALTER TABLE vydaje ADD COLUMN typ TEXT DEFAULT 'provozni'")
            except Exception: pass
            try: conn.execute("ALTER TABLE vydaje ADD COLUMN stitky TEXT DEFAULT ''")
            except Exception: pass
            try: conn.execute("ALTER TABLE vydaje ADD COLUMN duplicita_id INTEGER DEFAULT NULL")
            except Exception: pass
    # Migrace vystavene_faktury
    with get_db() as conn:
        if _USE_PG:
            cur4 = conn.execute("SELECT column_name FROM information_schema.columns WHERE table_name='vystavene_faktury'")
            vyst_cols = [r["column_name"] for r in cur4.fetchall()]
        else:
            vyst_cols = [row[1] for row in conn.execute("PRAGMA table_info(vystavene_faktury)").fetchall()]
        if "datum_splatnosti" not in vyst_cols:
            try: conn.execute("ALTER TABLE vystavene_faktury ADD COLUMN datum_splatnosti TEXT DEFAULT ''")
            except Exception: pass
        if "duplicita_id" not in vyst_cols:
            try: conn.execute("ALTER TABLE vystavene_faktury ADD COLUMN duplicita_id INTEGER DEFAULT NULL")
            except Exception: pass
    # Drive tabulky
    with get_db() as conn:
        conn.execute("""CREATE TABLE IF NOT EXISTS drive_zpracovane (
            id SERIAL PRIMARY KEY, file_id TEXT UNIQUE, zpracovano_at TEXT)""")
        conn.execute("""CREATE TABLE IF NOT EXISTS drive_channels (
            id SERIAL PRIMARY KEY, channel_id TEXT, resource_id TEXT, expiration TEXT)""")
    # Kalkulace
    with get_db() as conn:
        conn.execute("""CREATE TABLE IF NOT EXISTS kalkulace (
            id              SERIAL PRIMARY KEY,
            nazev           TEXT NOT NULL,
            popis           TEXT DEFAULT '',
            prodejni_cena   REAL DEFAULT 0,
            cil_marze_pct   REAL DEFAULT 200,
            created_at      TEXT DEFAULT NOW(),
            updated_at      TEXT DEFAULT NOW()
        )""")
        conn.execute("""CREATE TABLE IF NOT EXISTS kalkulace_pausalni (
            id              SERIAL PRIMARY KEY,
            kalkulace_id    INTEGER NOT NULL,
            nazev           TEXT NOT NULL,
            castka          REAL DEFAULT 0
        )""")
        conn.execute("""CREATE TABLE IF NOT EXISTS kalkulace_polozky (
            id              SERIAL PRIMARY KEY,
            kalkulace_id    INTEGER NOT NULL,
            nazev           TEXT NOT NULL,
            mnozstvi        REAL DEFAULT 1,
            jednotka        TEXT DEFAULT 'ks',
            cena_za_jednotku REAL DEFAULT 0,
            je_baleni       INTEGER DEFAULT 0,
            baleni_ks       REAL DEFAULT 1,
            zdroj_ceny      TEXT DEFAULT 'rucni'
        )""")
    # Ruční statistická data (průměry za roky bez dat v DB)
    with get_db() as conn:
        conn.execute("""CREATE TABLE IF NOT EXISTS stat_rucni_data (
            id      SERIAL PRIMARY KEY,
            rok     TEXT NOT NULL,
            mesic   TEXT NOT NULL,
            hodnota REAL NOT NULL DEFAULT 0,
            typ     TEXT NOT NULL DEFAULT 'trzba_vcpk_prumer',
            UNIQUE(rok, mesic, typ)
        )""")
    # Migrace pausalni_odvody — platnost_od
    with get_db() as conn:
        if _USE_PG:
            cur = conn.execute("SELECT column_name FROM information_schema.columns WHERE table_name='pausalni_odvody'")
            po_cols = [r["column_name"] for r in cur.fetchall()]
        else:
            po_cols = [row[1] for row in conn.execute("PRAGMA table_info(pausalni_odvody)").fetchall()]
        if "platnost_od" not in po_cols:
            try: conn.execute("ALTER TABLE pausalni_odvody ADD COLUMN platnost_od TEXT DEFAULT '2020-01-01'")
            except Exception: pass
# Reset sekvencí pro SERIAL sloupce (oprava null id)
    if _USE_PG:
        for tbl in ["vystavene_faktury", "faktury", "reporty", "vyplaty", "vydaje", "bankovni_pohyby", "zbozi", "polozky", "prava"]:
            try:
                with get_db() as conn:
                    conn.execute(f"SELECT setval(pg_get_serial_sequence('{tbl}', 'id'), COALESCE((SELECT MAX(id) FROM {tbl}), 0) + 1, false)")
            except Exception as e:
                print(f"⚠ Sekvence {tbl}: {e}")
    # paska_url ve vyplaty
    with get_db() as conn:
        if _USE_PG:
            cur = conn.execute("SELECT column_name FROM information_schema.columns WHERE table_name='vyplaty'")
            vypl_cols = [r["column_name"] for r in cur.fetchall()]
        else:
            vypl_cols = [row[1] for row in conn.execute("PRAGMA table_info(vyplaty)").fetchall()]
        if "paska_url" not in vypl_cols:
            try: conn.execute("ALTER TABLE vyplaty ADD COLUMN paska_url TEXT")
            except Exception: pass
        try: conn.execute("UPDATE reporty SET firma_zkratka='FP' WHERE firma_zkratka IS NULL OR firma_zkratka=''")
        except Exception: pass
    print("migrate_db OK")
    # Zrušit UNIQUE constraint na alias v zbozi_aliasy (více zboží může mít stejný alias)
    if _USE_PG:
        try:
            with get_db() as conn:
                conn.execute("ALTER TABLE zbozi_aliasy DROP CONSTRAINT IF EXISTS zbozi_aliasy_alias_key")
        except Exception:
            pass
            # Přidat UNIQUE constraint na prava tabulku pokud chybí
    if _USE_PG:
        try:
            with get_db() as conn:
                conn.execute("""
                    ALTER TABLE prava ADD CONSTRAINT prava_role_sekce_unique UNIQUE (role, sekce)
                """)
        except Exception:
            pass

    # Peněženka — hotovostní záznamy
    with get_db() as conn:
        conn.execute("""CREATE TABLE IF NOT EXISTS penezenka (
            id        SERIAL PRIMARY KEY,
            datum     TEXT NOT NULL,
            hotovost  REAL NOT NULL DEFAULT 0,
            rb_fp     REAL NOT NULL DEFAULT 0,
            rb_mr     REAL NOT NULL DEFAULT 0,
            rb_cff    REAL NOT NULL DEFAULT 0,
            rb_radek  REAL NOT NULL DEFAULT 0,
            air_fp    REAL NOT NULL DEFAULT 0,
            air_mr    REAL NOT NULL DEFAULT 0,
            air_cff   REAL NOT NULL DEFAULT 0,
            air_radek REAL NOT NULL DEFAULT 0,
            kb_radek  REAL NOT NULL DEFAULT 0,
            xtb_czk   REAL NOT NULL DEFAULT 0,
            xtb_eur   REAL NOT NULL DEFAULT 0,
            t212      REAL NOT NULL DEFAULT 0,
            etoro     REAL NOT NULL DEFAULT 0,
            sporeni   REAL NOT NULL DEFAULT 0,
            extras    TEXT DEFAULT '[]',
            poznamka  TEXT DEFAULT '',
            created_at TEXT DEFAULT NOW()
        )""")
        # Migrace — přidat sloupce pokud tabulka existuje bez nich
        if _USE_PG:
            cur = conn.execute("SELECT column_name FROM information_schema.columns WHERE table_name='penezenka'")
            pen_cols = [r["column_name"] for r in cur.fetchall()]
        else:
            pen_cols = [row[1] for row in conn.execute("PRAGMA table_info(penezenka)").fetchall()]
        for col in ["hotovost","rb_fp","rb_mr","rb_cff","rb_radek","air_fp","air_mr","air_cff","air_radek","kb_radek","xtb_czk","xtb_eur","t212","etoro","sporeni"]:
            if col not in pen_cols:
                try: conn.execute(f"ALTER TABLE penezenka ADD COLUMN {col} REAL DEFAULT 0")
                except Exception: pass
        if "extras" not in pen_cols:
            try: conn.execute("ALTER TABLE penezenka ADD COLUMN extras TEXT DEFAULT '[]'")
            except Exception: pass
        # Přejmenovat stary sloupec stav_skutecny → hotovost pokud existuje
        if "stav_skutecny" in pen_cols and "hotovost" not in pen_cols:
            try: conn.execute("ALTER TABLE penezenka RENAME COLUMN stav_skutecny TO hotovost")
            except Exception: pass
        # Přejmenovat akcie → xtb_czk pokud existuje (starý sloupec)
        if "akcie" in pen_cols and "xtb_czk" not in pen_cols:
            try: conn.execute("ALTER TABLE penezenka RENAME COLUMN akcie TO xtb_czk")
            except Exception: pass
            # Párování výpisů — nové sloupce
            try: conn.execute("ALTER TABLE vydaje ADD COLUMN var_sym TEXT DEFAULT ''")
            except Exception: pass
            try: conn.execute("ALTER TABLE vydaje ADD COLUMN datum_zaplaceno TEXT DEFAULT ''")
            except Exception: pass
            try: conn.execute("ALTER TABLE faktury ADD COLUMN var_sym TEXT DEFAULT ''")
            except Exception: pass
            try: conn.execute("ALTER TABLE faktury ADD COLUMN datum_zaplaceno TEXT DEFAULT ''")
            except Exception: pass
            try: conn.execute("ALTER TABLE vystavene_faktury ADD COLUMN var_sym TEXT DEFAULT ''")
            except Exception: pass
            try: conn.execute("ALTER TABLE vystavene_faktury ADD COLUMN datum_zaplaceno TEXT DEFAULT ''")
            except Exception: pass
            try: conn.execute("ALTER TABLE bankovni_pohyby ADD COLUMN var_sym TEXT DEFAULT ''")
            except Exception: pass
            try: conn.execute("ALTER TABLE bankovni_pohyby ADD COLUMN sparovano INTEGER DEFAULT 0")
            except Exception: pass
            try: conn.execute("ALTER TABLE bankovni_pohyby ADD COLUMN sparovano_typ TEXT DEFAULT ''")
            except Exception: pass
            try: conn.execute("ALTER TABLE bankovni_pohyby ADD COLUMN sparovano_id INTEGER DEFAULT NULL")
            except Exception: pass

    # Trvalé příkazy (soukromé výdaje)
    with get_db() as conn:
        conn.execute("""CREATE TABLE IF NOT EXISTS trvale_prikazy (
            id            SERIAL PRIMARY KEY,
            lokace        TEXT NOT NULL DEFAULT '',
            dodavatel     TEXT NOT NULL DEFAULT '',
            popis         TEXT NOT NULL DEFAULT '',
            zpusob_uhrady TEXT NOT NULL DEFAULT 'převodem',
            castka        REAL NOT NULL DEFAULT 0,
            den_v_mesici  INTEGER NOT NULL DEFAULT 1,
            aktivni       INTEGER NOT NULL DEFAULT 1,
            poznamka      TEXT DEFAULT '',
            created_at    TEXT DEFAULT NOW()
        )""")

    # Dluhy — půjčky kamarádům
    with get_db() as conn:
        conn.execute("""CREATE TABLE IF NOT EXISTS dluhy_osoby (
            id         SERIAL PRIMARY KEY,
            jmeno      TEXT NOT NULL UNIQUE,
            poznamka   TEXT DEFAULT '',
            created_at TEXT DEFAULT NOW()
        )""")
        conn.execute("""CREATE TABLE IF NOT EXISTS dluhy_transakce (
            id         SERIAL PRIMARY KEY,
            osoba_id   INTEGER NOT NULL REFERENCES dluhy_osoby(id) ON DELETE CASCADE,
            datum      TEXT NOT NULL,
            castka     REAL NOT NULL,
            poznamka   TEXT DEFAULT '',
            created_at TEXT DEFAULT NOW()
        )""")


def allowed_file(filename):
    return "." in filename and filename.rsplit(".", 1)[1].lower() in ALLOWED_EXT
    today = date.today().isoformat()
    with get_db() as conn:
        conn.execute("""
            UPDATE faktury SET stav = 'po_splatnosti'
            WHERE stav = 'ceka'
              AND datum_splatnosti IS NOT NULL
              AND datum_splatnosti < ?
        """, (today,))

def recalc_faktura_total(conn, faktura_id):
    row = conn.execute("SELECT COALESCE(SUM(celkem_s_dph),0) as total FROM polozky WHERE faktura_id=?", (faktura_id,)).fetchone()
    total = _first_val(row)
    conn.execute("UPDATE faktury SET celkem_s_dph=? WHERE id=?", (total, faktura_id))


# ── MAKRO parser ───────────────────────────────────────────────────────────────
def parse_makro_pdf(filepath):
    if not PDF_SUPPORT:
        return None, "pdfplumber není nainstalován"

    result = {
        "cislo_faktury":   "",
        "datum_vystaveni": "",
        "datum_splatnosti":"",
        "zpusob_uhrady":   "",
        "dodavatel":       "MAKRO Cash & Carry ČR s.r.o.",
        "celkem_s_dph":    0,
        "polozky":         []
    }

    try:
        from collections import defaultdict
        import re as _re

        all_items      = []
        full_text_lines = []

        with pdfplumber.open(filepath) as pdf:
            first_text = pdf.pages[0].extract_text() or ""
            first_despaced = _re.sub(r"(?<=\S) (?=\S)", "", first_text)
            if "Súpistovaru" in first_despaced and "FAKTURA" not in first_despaced:
                return None, "Tento soubor je 'Súpis tovaru' (interní doklad MAKRO) – není to daňová faktura. Soubor nebyl nahrán."

            # Naskenované PDF — žádný text → předat Claude API s MAKRO kontextem
            if not first_text.strip():
                api_key = os.environ.get("ANTHROPIC_API_KEY", "")
                if api_key:
                    return parse_faktura_claude(filepath)
                return None, "Naskenované PDF — OCR není k dispozici"

            # Pokud PDF neobsahuje MAKRO text, předej rovnou Claude
            makro_keywords = ["MAKRO", "makro", "Cash & Carry"]
            if not any(kw in first_text for kw in makro_keywords):
                return None, "Není MAKRO faktura"

            for page in pdf.pages:
                full_text_lines += (page.extract_text() or "").splitlines()
                words = page.extract_words(x_tolerance=1, y_tolerance=2)

                rows = defaultdict(list)
                for w in words:
                    y = round(w["top"] / 2) * 2
                    rows[y].append(w)

                for y, ws in sorted(rows.items()):
                    ws = sorted(ws, key=lambda w: w["x0"])

                    left_digits = "".join(
                        w["text"] for w in ws
                        if w["x0"] < 95 and len(w["text"]) == 1 and w["text"].isdigit()
                    )
                    if len(left_digits) < 6:
                        left_tokens = "".join(
                            w["text"] for w in ws if w["x0"] < 95
                        ).replace("*", "").strip()
                        if re.match(r"^\d{6,}", left_tokens):
                            left_digits = left_tokens[:14]

                    unit_chars = "".join(
                        w["text"] for w in ws
                        if 230 <= w["x0"] <= 275 and len(w["text"]) == 1
                        and w["text"].upper() in "PCGKBSLXAW"
                    ).upper()
                    if   unit_chars.startswith("PC"): jed = "PC"
                    elif unit_chars.startswith("KG"): jed = "KG"
                    elif unit_chars.startswith("BG"): jed = "BG"
                    elif unit_chars.startswith("BX"): jed = "BX"
                    elif unit_chars.startswith("KS"): jed = "KS"
                    elif unit_chars.startswith("CA"): jed = "CA"
                    elif unit_chars.startswith("SW"): jed = "SW"
                    elif unit_chars.startswith("WA"): jed = "SW"
                    elif unit_chars.startswith("L"):  jed = "L"
                    else:                              jed = ""

                    all_text_j = "".join(w["text"] for w in ws).lower()
                    is_sleva = "urcenopro" in all_text_j or "kupvice" in all_text_j or "kupvíce" in all_text_j
                    if is_sleva and all_items:
                        right_text = "".join(w["text"] for w in sorted(
                            [w for w in ws if w["x0"] > 265], key=lambda w: w["x0"]
                        ))
                        neg = re.findall(r"-?(\d+,\d{2})", right_text)
                        if neg:
                            sleva = _parse_money(neg[-1])
                            all_items[-1]["celkem_s_dph"] = round(max(0, all_items[-1]["celkem_s_dph"] - sleva), 2)
                            mn = all_items[-1]["mnozstvi"]
                            if mn:
                                all_items[-1]["cena_za_jednotku_s_dph"] = round(all_items[-1]["celkem_s_dph"] / mn, 4)
                        continue

                    if len(left_digits) < 6 or not jed:
                        continue

                    nazev_ws = [w for w in ws if 90 <= w["x0"] <= 237]
                    nazev = _rekonstruuj_nazev(nazev_ws)

                    right_ws = sorted([w for w in ws if w["x0"] > 265], key=lambda w: w["x0"])
                    cf = _makro_reconstruct_numbers(right_ws)

                    if len(cf) < 2:
                        continue

                    idx_dph      = len(cf)
                    idx_celkem_s = idx_dph - 1
                    idx_pocet    = idx_dph - 3

                    celkem_s_dph = cf[idx_celkem_s] if 0 <= idx_celkem_s < len(cf) else 0
                    pocet        = cf[idx_pocet]    if 0 <= idx_pocet    < len(cf) else 1.0
                    if pocet <= 0 or pocet > 10000:
                        pocet = 1.0
                    cena_j = round(celkem_s_dph / pocet, 4) if pocet else celkem_s_dph

                    if not nazev or celkem_s_dph <= 0:
                        continue

                    all_items.append({
                        "nazev":                  nazev,
                        "mnozstvi":               pocet,
                        "jednotka":               _map_unit(jed),
                        "cena_za_jednotku_s_dph": cena_j,
                        "celkem_s_dph":           round(celkem_s_dph, 2)
                    })

        result["polozky"] = all_items

        def despace(s):
            return re.sub(r"(?<=\S) (?=\S)", "", s)

        ico_odberatele = ""
        for line in full_text_lines:
            dl = despace(line)
            if not result["cislo_faktury"]:
                m = re.search(r"Faktura.*?VS.*?:?\s*(\d{7,12})", dl, re.IGNORECASE)
                if m: result["cislo_faktury"] = m.group(1)
            if not result["cislo_faktury"]:
                m = re.search(r"Súpistovaru\s*(\d{7,12})", dl, re.IGNORECASE)
                if m: result["cislo_faktury"] = m.group(1)
            if not result["cislo_faktury"]:
                m = re.search(r"TechnickéID.*?/(\d{7,12})\)", dl, re.IGNORECASE)
                if m: result["cislo_faktury"] = m.group(1)
            if not ico_odberatele:
                m = re.search(r"IČ\s*:\s*(\d{8})", dl)
                if m: ico_odberatele = m.group(1)
            if not result["datum_vystaveni"]:
                m = re.search(r"vystavení.*?(\d{2}-\d{2}-\d{4})", dl, re.IGNORECASE)
                if m: result["datum_vystaveni"] = _makro_date(m.group(1))
            if not result["datum_splatnosti"]:
                m = re.search(r"splatnosti.*?(\d{2}-\d{2}-\d{4})", dl, re.IGNORECASE)
                if m: result["datum_splatnosti"] = _makro_date(m.group(1))
            if not result["zpusob_uhrady"]:
                m = re.search(r"Způsobúhrady:?\s*([A-Za-záéíóúýžšČřďťňÁÉÍÓÚÝŽŠČŘĎŤŇ]+(?:\s+[A-Za-záéíóúýžšČřďťňÁÉÍÓÚÝŽŠČŘĎŤŇ]+)?)", dl, re.IGNORECASE)
                if m:
                    u = m.group(1).strip()
                    if u and u.lower() not in ("praha", "pruhonice", "chudenicka", ""):
                        result["zpusob_uhrady"] = u

            dl_line = despace(line)
            if "Celkov" in dl_line and "stka" in dl_line:
                nums = re.findall(r"(\d{1,3}(?:\s\d{3})*[,\.]\d{2})", line)
                if nums:
                    result["celkem_s_dph"] = _parse_money(nums[-1])

        if result["celkem_s_dph"] == 0 and all_items:
            result["celkem_s_dph"] = round(sum(p["celkem_s_dph"] for p in all_items), 2)

        result["zpusob_uhrady"] = "Hotovost"
        result["stav"] = "zaplaceno"
        result["ico_odberatele"] = ico_odberatele
        result["firma_zkratka"] = _ico_na_firmu(ico_odberatele) or "UNI"

        for p in result["polozky"]:
            p["nazev"] = _format_nazev(p["nazev"])

    except Exception as e:
        return None, str(e)

    return result, None


def _makro_reconstruct_numbers(ws_sorted):
    if not ws_sorted:
        return []
    groups = []
    current = [ws_sorted[0]]
    for prev, curr in zip(ws_sorted, ws_sorted[1:]):
        gap = curr["x0"] - prev["x0"] - 3.5
        if gap > 8:
            groups.append(current)
            current = [curr]
        else:
            current.append(curr)
    groups.append(current)

    DPH_SAZBY = {6.0, 10.0, 15.0, 21.0, 23.0}

    parsed = []
    for g in groups:
        token = "".join(w["text"] for w in g).replace(",", ".")
        x0 = g[0]["x0"]
        if re.match(r"^\d{5,}$", token):
            continue
        if re.match(r"^[A-Za-z]+$", token):
            continue
        try:
            val = float(token)
            parsed.append((x0, val))
        except Exception:
            pass

    dph_idx = None
    for i, (x0, val) in enumerate(parsed):
        if val in DPH_SAZBY and val == int(val) and x0 > 480:
            dph_idx = i
            break

    if dph_idx is not None:
        parsed = parsed[:dph_idx]

    return [val for _, val in parsed]


def _rekonstruuj_nazev(nazev_ws):
    if not nazev_ws:
        return ""
    result = ""
    for i, w in enumerate(nazev_ws):
        if i > 0:
            gap = w["x0"] - nazev_ws[i-1]["x1"]
            if gap > 3.5:
                result += " "
        result += w["text"]
    result = re.sub(r" {2,}", " ", result)
    return result.lstrip("*").strip()


def _format_nazev(nazev):
    result = re.sub(r"  +", " ", nazev).strip()
    return result


def _ico_na_firmu(ico):
    try:
        import json, os
        cfg_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "config.json")
        with open(cfg_path) as f:
            cfg = json.load(f)
        ico_map = cfg.get("ico_map", {})
        return ico_map.get(ico, "")
    except Exception:
        return ""


def _makro_date(s):
    s = s.replace("-", ".")
    try:
        return datetime.strptime(s, "%d.%m.%Y").strftime("%Y-%m-%d")
    except Exception:
        return s


def _parse_makro_items(lines):
    return []


def _ocr_best_orientation(img):
    best_text = ""
    best_score = 0
    for angle in [0, 90, 180, 270]:
        rotated = img.rotate(angle, expand=True) if angle else img
        for lang in ["ces+eng", "ces", "eng"]:
            try:
                text = pytesseract.image_to_string(rotated, lang=lang, config="--psm 6 --oem 3")
                score = sum(text.count(kw) for kw in [
                    "MAKRO", "Faktura", "Datum", "DPH", "Kč", "PC", "KG", "BG",
                    "splatnosti", "vystavení", "Food Plus", "Odběratel"
                ])
                if score > best_score:
                    best_score = score
                    best_text = text
                break
            except Exception:
                continue
    return best_text


def _precti_celkovou_castku_z_pdf(filepath):
    """Přečte celkovou částku přímo z textu PDF přes pdfplumber.
    Pro naskenované PDF použije OCR (Tesseract).
    Vrací float nebo None."""
    if not PDF_SUPPORT:
        return None
    try:
        import re as _re
        klicova_slova = [
            r"celkov.{0,3}\s*.{0,3}stka",
            r"k\s*[uú]hrad[eě]",
            r"celkem\s*s\s*dph",
        ]
        with pdfplumber.open(filepath) as _pdf:
            # Nejdřív zkus textové PDF (stránky od poslední)
            for _page in reversed(_pdf.pages):
                _text = _page.extract_text() or ""
                if _text.strip():
                    for _kw in klicova_slova:
                        for _line in _text.splitlines():
                            if _re.search(_kw, _line, _re.IGNORECASE):
                                _nums = _re.findall(r"(\d{1,3}(?:[\s]\d{3})*[,.]\d{2})", _line)
                                if _nums:
                                    return _parse_money(_nums[-1])

            # Naskenované PDF — použij OCR přes Tesseract jen na poslední stránce
            if OCR_SUPPORT:
                import io as _io
                _page = _pdf.pages[-1]
                _pil = _page.to_image(resolution=150).original
                _text = pytesseract.image_to_string(_pil, lang="ces+eng")
                for _line in _text.splitlines():
                    if _re.search(r"celkov.{0,3}\s*.{0,3}stka", _line, _re.IGNORECASE) \
                            and "strana" not in _line.lower() \
                            and "stran" not in _line.lower():
                        _nums = _re.findall(r"(\d{1,3}(?:[\s]\d{3})*[,.]\d{2})", _line)
                        if _nums:
                            app.logger.info(f"[OCR CASTKA] Nalezeno: '{_line}' -> {_parse_money(_nums[-1])}")
                            return _parse_money(_nums[-1])
                        else:
                            app.logger.info(f"[OCR CASTKA] Řádek nalezen ale bez čísel: '{_line}'")
    except Exception as _ex:
        app.logger.warning(f"[OCR CASTKA] Chyba: {_ex}")
    app.logger.info("[OCR CASTKA] Nic nenalezeno, vrací None")
    return None


def parse_faktura_claude(filepath):
    """Univerzální parser faktur a účtenek přes Claude API – funguje pro PDF i obrázky."""
    api_key = os.environ.get("ANTHROPIC_API_KEY", "")
    if not api_key:
        return None, "ANTHROPIC_API_KEY není nastaven"

    try:
        ext = filepath.rsplit(".", 1)[-1].lower()

        # PDF — poslat přímo Claude jako dokument
        if ext == "pdf":
            castka_z_textu = None  # OCR vypnuto - příliš pomalé na Cloud Run
            with open(filepath, "rb") as f:
                raw = f.read()
            b64 = base64.standard_b64encode(raw).decode("utf-8")
            content_block = {"type": "document", "source": {"type": "base64", "media_type": "application/pdf", "data": b64}}
            content_blocks = None
        else:
            # Obrázek (JPG, PNG...)
            media_map = {"jpg": "image/jpeg", "jpeg": "image/jpeg", "png": "image/png",
                         "bmp": "image/bmp", "tiff": "image/tiff", "webp": "image/webp"}
            media_type = media_map.get(ext, "image/jpeg")
            if OCR_SUPPORT and ext in ("jpg", "jpeg", "png", "bmp", "tiff"):
                try:
                    import io as _io
                    _img = Image.open(filepath)
                    # Zmenšit pokud přesahuje limit Claude API (8000px)
                    _max = 3500
                    if _img.width > _max or _img.height > _max:
                        _img.thumbnail((_max, _max), Image.LANCZOS)
                    _buf = _io.BytesIO()
                    _img.save(_buf, format="JPEG", quality=85)
                    b64 = base64.standard_b64encode(_buf.getvalue()).decode("utf-8")
                    media_type = "image/jpeg"
                except Exception:
                    with open(filepath, "rb") as f:
                        b64 = base64.standard_b64encode(f.read()).decode("utf-8")
            else:
                with open(filepath, "rb") as f:
                    b64 = base64.standard_b64encode(f.read()).decode("utf-8")
            content_block = {
                "type": "image",
                "source": {"type": "base64", "media_type": media_type, "data": b64}
            }
            content_blocks = None

        prompt = """Jsi expert na čtení faktur a účtenek. Přečti tento doklad VELMI PEČLIVĚ.
DŮLEŽITÉ: Dokument může být otočený o 90, 180 nebo 270 stupňů — přečti ho správně bez ohledu na orientaci.
Odpověz POUZE platným JSON objektem, žádný jiný text, žádné backticky, žádné komentáře.

Formát odpovědi:
{
  "dodavatel": "název dodavatele nebo obchodu",
  "cislo_faktury": "pro MAKRO faktury: číslo POUZE z pole Faktura c. / VS (10 číslic, např. 0415000291) — IGNORUJ číslo vpravo nahoře (formát 0015/0135) a IGNORUJ c. zákazníka. Pro ostatní faktury: číslo faktury nebo VS nebo null",
  "datum_vystaveni": "YYYY-MM-DD nebo null",
  "datum_splatnosti": "YYYY-MM-DD nebo null",
  "zpusob_uhrady": "hotově/kartou/převodem nebo null",
  "celkem_s_dph": číslo (celková částka včetně DPH),
  "polozky": [
    {
      "nazev": "název položky",
      "mnozstvi": číslo,
      "jednotka": "ks/kg/l/...",
      "cena_za_jednotku_s_dph": číslo,
      "celkem_s_dph": číslo
    }
  ]
}

PRAVIDLA:
- Všechny částky jsou v Kč, piš jen číslo bez symbolu Kč
- Desetinná čárka nebo tečka = desetinné místo (475,55 = 475.55)
- Pokud není datum splatnosti, vrať null
- Pokud není číslo faktury/VS, vrať null
- Způsob úhrady: pokud vidíš "karta", "card", "kartou" → "kartou"; "cash", "hotov" → "hotově"
- Položky: zahrň všechny položky které vidíš na dokladu
- celkem_s_dph u položky = množství × cena za jednotku
- CELKOVÁ ČÁSTKA (celkem_s_dph): Hledej pole "Celková částka" — je to JEDINÝ řádek s tímto textem na celém dokumentu. U MAKRO faktur je na poslední straně pod tabulkou DPH, těsně nad "Platba kartou/hotově". IGNORUJ: "Strana celkem bez DPH", "Poslední strana celkem bez DPH", čísla v tabulce DPH (hodnota zboží, částka daně, Celkem v DPH tabulce). Správná celková částka je VŽDY nižší než součet položek kvůli slevám CLAP.
"""

        client = anthropic.Anthropic(api_key=api_key)
        # Sestavit obsah zprávy - buď více stránek (content_blocks) nebo jeden blok
        if content_blocks:
            msg_content = content_blocks + [{"type": "text", "text": prompt}]
        else:
            msg_content = [content_block, {"type": "text", "text": prompt}]
        message = client.messages.create(
            model="claude-sonnet-4-20250514",
            max_tokens=4000,
            messages=[{"role": "user", "content": msg_content}]
        )

        text = message.content[0].text.strip()
        text = re.sub(r"^```json\s*", "", text)
        text = re.sub(r"```$", "", text).strip()
        parsed = json.loads(text)

        # Normalizace výstupu
        app.logger.info(f"[PARSE] Finální castka_z_textu={castka_z_textu}, claude_castka={parsed.get('celkem_s_dph')}")
        result = {
            "dodavatel":        parsed.get("dodavatel", ""),
            "cislo_faktury":    parsed.get("cislo_faktury") or "",
            "datum_vystaveni":  parsed.get("datum_vystaveni") or "",
            "datum_splatnosti": parsed.get("datum_splatnosti") or "",
            "zpusob_uhrady":    parsed.get("zpusob_uhrady") or "",
            "celkem_s_dph":     castka_z_textu if castka_z_textu else float(parsed.get("celkem_s_dph") or 0),
            "polozky": [
                {
                    "nazev":                   p.get("nazev", ""),
                    "mnozstvi":                float(p.get("mnozstvi", 1) or 1),
                    "jednotka":                p.get("jednotka", "ks"),
                    "cena_za_jednotku_s_dph":  float(p.get("cena_za_jednotku_s_dph", 0) or 0),
                    "celkem_s_dph":            float(p.get("celkem_s_dph", 0) or 0),
                }
                for p in parsed.get("polozky", [])
                if p.get("nazev", "").strip()
            ]
        }
        return result, None

    except Exception as e:
        return None, str(e)


def parse_makro_image(filepath):
    if not OCR_SUPPORT:
        return None, "pytesseract/Pillow není nainstalován"
    try:
        img = Image.open(filepath)
        img = img.convert("L")
        w, h = img.size
        needs_rotation_check = (w > h * 1.2) or (h > w * 1.2)
        if w < 1200:
            scale = 1200 / w
            img = img.resize((int(w*scale), int(h*scale)), Image.LANCZOS)

        if needs_rotation_check:
            text = _ocr_best_orientation(img)
        else:
            for lang in ["ces+eng", "ces", "eng"]:
                try:
                    text = pytesseract.image_to_string(img, lang=lang, config="--psm 6 --oem 3")
                    break
                except Exception:
                    continue

        lines = text.splitlines()
        result = {
            "cislo_faktury":   "",
            "datum_vystaveni": "",
            "datum_splatnosti":"",
            "zpusob_uhrady":   "Hotovost",
            "stav":            "zaplaceno",
            "dodavatel":       "MAKRO Cash & Carry ČR s.r.o.",
            "celkem_s_dph":    0,
            "firma_zkratka":   "",
            "polozky":         []
        }

        for line in lines:
            ls = line.strip()
            if not result["cislo_faktury"]:
                m = re.search(r"Faktura.*?[Vv][Ss]\s*[;:,.]?\s*([\d\s]{7,15})", ls, re.IGNORECASE)
                if m:
                    vs = re.sub(r"\s+", "", m.group(1))[:12]
                    if vs.isdigit() and len(vs) >= 7: result["cislo_faktury"] = vs
            m = re.search(r"(\d{2})[.\-](\d{2})[.\-](\d{4})", ls)
            if m:
                den, mes, rok = m.group(1), m.group(2), m.group(3)
                if int(mes) > 12:
                    mes = mes.replace("8", "0")
                try:
                    from datetime import datetime
                    datetime(int(rok), int(mes), int(den))
                    d = f"{rok}-{mes}-{den}"
                    if not result["datum_vystaveni"]: result["datum_vystaveni"] = d
                    elif not result["datum_splatnosti"]: result["datum_splatnosti"] = d
                except Exception:
                    pass
            if not result["zpusob_uhrady"] or result["zpusob_uhrady"] == "Hotovost":
                if "Platba kartou" in ls or "platba kartou" in ls:
                    result["zpusob_uhrady"] = "Platba kartou"
            if not result["firma_zkratka"]:
                m = re.search(r"IČ\s*:\s*(\d{8})", ls)
                if m: result["firma_zkratka"] = _ico_na_firmu(m.group(1))
            m = re.search(r"Celkov[aá]\s+[čc][aá]stka\s+([\d\s]{1,10}[,.]\d{2})", ls, re.IGNORECASE)
            if m: result["celkem_s_dph"] = _parse_money(m.group(1))
            m2 = re.search(r"[Ss]trana.{0,10}celkem.{0,10}bez.{0,5}DPH.{0,5}([\d\s]+[,.]\d{2})", ls, re.IGNORECASE)
            if m2 and not result.get("ocr_strana_celkem_bez_dph"):
                result["ocr_strana_celkem_bez_dph"] = _parse_money(m2.group(1))
            m3 = re.search(r"celkem\s+bez\s+DPH\s+([\d\s]+[,.]\d{2})", ls, re.IGNORECASE)
            if m3 and not result.get("ocr_strana_celkem_bez_dph"):
                result["ocr_strana_celkem_bez_dph"] = _parse_money(m3.group(1))

        result["polozky"] = _parse_ocr_items(lines)
        suma_polozek = round(sum(p["celkem_s_dph"] for p in result["polozky"]), 2)
        if result["celkem_s_dph"] == 0:
            result["celkem_s_dph"] = suma_polozek

        if not result["firma_zkratka"]:
            result["firma_zkratka"] = "UNI"

        ocr_bez = result.get("ocr_strana_celkem_bez_dph", 0)
        podezrele = [i for i, p in enumerate(result["polozky"])
                     if p["celkem_s_dph"] == 0 or p["mnozstvi"] > 500]
        result["ocr_kontrola"] = {
            "suma_polozek": suma_polozek,
            "ocr_bez_dph": ocr_bez,
            "ma_celkem": ocr_bez > 0,
            "podezrele_indexy": podezrele,
        }

        return result, None
    except Exception as e:
        return None, str(e)


def _parse_ocr_items(lines):
    items = []
    sleva_kw = ["urceno pro konecnou", "určeno pro konečnou", "kup vice", "kup více"]
    jednotky = {"PC", "KG", "BG", "KS", "BX", "CA", "SW", "BT",
                "B6", "86", "PG", "6G", "BQ", "BC", "2B", "CA"}
    jednotka_map = {"B6": "BG", "86": "BG", "PG": "PC", "6G": "BG",
                    "BQ": "BG", "BC": "BX", "2B": "BG"}

    for line in lines:
        ls = line.strip()
        if not ls: continue
        ll = ls.lower()

        is_sleva = any(kw in ll for kw in sleva_kw)
        if is_sleva and items:
            nums = re.findall(r"-\s*(\d[\d\s]*[,.]\d{2})", ls)
            if nums:
                sleva = _parse_money(nums[-1])
                items[-1]["celkem_s_dph"] = round(max(0, items[-1]["celkem_s_dph"] - sleva), 2)
                mn = items[-1]["mnozstvi"]
                if mn: items[-1]["cena_za_jednotku_s_dph"] = round(items[-1]["celkem_s_dph"] / mn, 4)
            continue

        ls_clean = re.sub(r"^[Ss|lIG]+(?=\d)", "", ls)
        ls_clean = re.sub(r"^[|l]\s+", "", ls_clean)
        m = re.match(r"^(\d{6,14})\s+[\*\-—–|]*\s*(.+)", ls_clean)
        if not m: continue

        rest_after_mm = m.group(2).strip().lstrip("*").strip()

        jednotka = ""
        nazev = rest_after_mm
        cisla_str = ""

        for jed in jednotky:
            pat = r"^(.+?)\s+" + jed + r"\s+(.+)$"
            mj = re.match(pat, rest_after_mm, re.IGNORECASE)
            if mj:
                nazev    = mj.group(1).strip().rstrip("*").strip()
                jednotka = jednotka_map.get(jed, jed)
                cisla_str = mj.group(2)
                break

        if not jednotka:
            mj = re.search(r"\s(PC|KG|BG|KS|BX|CA|SW|BT)\s", rest_after_mm, re.IGNORECASE)
            if mj:
                jednotka = mj.group(1).upper()
                nazev    = rest_after_mm[:mj.start()].strip().rstrip("*")
                cisla_str = rest_after_mm[mj.end():]

        if not cisla_str:
            cisla_str = rest_after_mm

        cisla_str = re.sub(r"(\d+)[,\.](\s+)(\d+)", r"\1.\3", cisla_str)
        cisla_raw = re.findall(r"\d+[,.]\d+|\d+", cisla_str)
        cf = []
        for c in cisla_raw:
            try:
                val = float(c.replace(",", "."))
                if val == int(val) and val >= 10000:
                    continue
                cf.append(val)
            except:
                pass

        if len(cf) < 2: continue

        idx_dph = None
        for i in range(len(cf)-1, -1, -1):
            if cf[i] in (6.0, 10.0, 15.0, 23.0):
                idx_dph = i
                break
        if idx_dph is None: idx_dph = len(cf)

        if (idx_dph >= 2 and
                cf[idx_dph-1] >= 100 and
                cf[idx_dph-2] == int(cf[idx_dph-2]) and
                1 <= cf[idx_dph-2] <= 9):
            celkem = round(cf[idx_dph-2] * 1000 + cf[idx_dph-1], 2)
            pocet  = cf[idx_dph-5] if idx_dph >= 5 else 1.0
        else:
            celkem = cf[idx_dph-1] if idx_dph >= 1 else 0
            pocet  = cf[idx_dph-3] if idx_dph >= 3 else 1.0

        if pocet <= 0 or pocet > 10000: pocet = 1.0
        cena_j = round(celkem / pocet, 4) if pocet else celkem

        nazev = re.sub(r"^[|\-—–\s]+", "", nazev).strip()
        if not nazev or celkem <= 0: continue

        items.append({
            "nazev":                  _format_nazev(nazev),
            "mnozstvi":               pocet,
            "jednotka":               _map_unit(jednotka) if jednotka else "ks",
            "cena_za_jednotku_s_dph": cena_j,
            "celkem_s_dph":           round(celkem, 2)
        })
    return items


def _cz_date(s):
    try:
        return datetime.strptime(s, "%d.%m.%Y").strftime("%Y-%m-%d")
    except Exception:
        return s

def _parse_money(s):
    s = str(s).replace(" ", "").replace(".", "").replace(",", ".")
    try:
        return float(s)
    except Exception:
        return 0.0

def _map_unit(u):
    mapping = {"PC": "ks", "KS": "ks", "KG": "kg", "BG": "bal", "BX": "bal", "CA": "bal", "SW": "bal", "L": "l"}
    return mapping.get(u.upper(), u.lower())


JMENA_MAP = {
    "rada": "Ráďa", "radek": "Ráďa", "ráďa": "Ráďa", "radi": "Ráďa",
    "verka": "Věrka", "vera": "Věrka", "věra": "Věrka", "věrka": "Věrka",
    "renča": "Renča", "renata": "Renča", "renca": "Renča",
    "vendy": "Vendy", "wendy": "Vendy",
    "vali": "Vali",
}

def normalize_jmena(text):
    if not text:
        return ""
    parts = re.split(r"[,/\s]+", text.strip())
    result = []
    for p in parts:
        p = p.strip().lower().rstrip(".,")
        if not p:
            continue
        canonical = JMENA_MAP.get(p, p.capitalize())
        result.append(canonical)
    return ", ".join(result)


def parse_report_image_claude(filepath):
    api_key = os.environ.get("ANTHROPIC_API_KEY", "")
    if not api_key:
        return None, "ANTHROPIC_API_KEY není nastaven"

    try:
        with open(filepath, "rb") as f:
            img_data = base64.standard_b64encode(f.read()).decode("utf-8")

        ext = filepath.rsplit(".", 1)[-1].lower()
        media_map = {"jpg": "image/jpeg", "jpeg": "image/jpeg", "png": "image/png",
                     "bmp": "image/bmp", "tiff": "image/tiff"}
        media_type = media_map.get(ext, "image/jpeg")

        client = anthropic.Anthropic(api_key=api_key)

        _t = date.today()
        today = f"{_t.day}.{_t.month}"
        prompt = f"""Jsi expert na čtení ručně psaných restauračních reportů z bistra.
Odpověz POUZE platným JSON objektem, žádný jiný text, žádné backticky.

Formát odpovědi:
{{
  "datum": "D.M" nebo null,
  "den": "název dne česky" nebo null,
  "smena": "jména oddělená čárkou" nebo null,
  "karty": číslo nebo 0,
  "kov": číslo nebo 0,
  "papir": číslo nebo 0,
  "vydaje": číslo nebo 0,
  "pk50_ks": celé číslo nebo 0,
  "pk100_ks": celé číslo nebo 0,
  "pizza_cela": celé číslo nebo 0,
  "pizza_ctvrt": celé číslo nebo 0,
  "burger": celé číslo nebo 0,
  "talire": celé číslo nebo 0,
  "burtgulas": celé číslo nebo 0
}}

=== DATUM ===
- Hledej nahoře na lístku formát "D.M" nebo "D/M" — jen den a měsíc, bez roku
- Příklad: "19.3" → "19.3", "5/2" → "5.2"
- Neplést s jinými čísly na lístku (karty, tržba...)
- Pokud datum není, vrať dnešní: "{today}"
- POZOR na záměnu číslic v měsíci: "3" a "5" jsou si podobné — měsíc 3 = březen, měsíc 5 = květen
- Pole "den" NEVYPLŇUJ — vrať vždy null, den spočítáme sami z data

=== ČÍSLA — ZÁMĚNY ČÍSLIC ===
- Tečka nebo čárka uvnitř čísla = oddělovač tisíců: 9.582 = 9582, 4.900 = 4900, 8.527 = 8527
- Pomlčka nebo lomítko za číslem (9.582,-) = ignoruj
- Výsledek rovnice: "14.521 + 5 = 14.636" → ber číslo ZA "=" = 14636
- KRITICKÉ — při ručním psaní jsou si podobné: 3↔5, 1↔7, 0↔6, 4↔9, 1↔4
- Číslo vždy ověř pomocí kontrolního součtu: KARTY + KOV + PAPÍR + VÝDAJE = TRŽBA CELKEM
- Pokud součet nesedí, zkus alternativní čtení záměnných číslic (3↔5, 1↔4) dokud součet nesedí

=== KARTY, KOV, PAPÍR, VÝDAJE ===
- KARTY = platby kartou (větší číslo, typicky 4000–15000 Kč)
- KOV = drobné mince (malé číslo, typicky 20–200 Kč, NIKDY tisíce)
- PAPÍR = papírové bankovky (stovky až tisíce Kč)
- VÝDAJE = hotovost vydaná ven z kasy — typicky 0–500 Kč, NIKDY tisíce
- TRŽBA CELKEM (bez PK) = KARTY + KOV + PAPÍR + VÝDAJE — použij tento součet pro ověření!
- TRŽBA CELKEM vč. výdajů na lístku = KARTY + KOV + PAPÍR + VÝDAJE (to je správná kontrola)
- KRITICKÉ pro VÝDAJE: jsou to typicky malé částky (100, 200 Kč). Pokud vidíš 400 ale součet nesedí, zkus 100 — záměna 4↔1 je velmi častá!
- KRITICKÉ pro KARTY: číslo okolo 8000 Kč. Pokud vidíš 8327 ale součet nesedí, zkus 8527 — záměna 3↔5 je velmi častá!

=== POUKAZKY (PK) — VELMI DŮLEŽITÉ ===
- Hledej na lístku "PK" nebo "POUKAZ" nebo "POUKAZKA"
- Formát "6x 100 = 600" nebo "6x/100" nebo "6x100" → pk100_ks = 6
- Formát "3x 50 = 150" nebo "3x/50" nebo "3x50" → pk50_ks = 3
- Číslo před "x" = počet kusů, číslo za "x" = hodnota (50 nebo 100 Kč)
- NIKDY nezapisuj 0 pokud PK na lístku je!

=== PIZZA — VELMI DŮLEŽITÉ ===
- Hledej sekci PIZZA na lístku
- CELÁ / CELÉ / C: → pizza_cela (číslo hned za tím, "2x" = 2)
- ČTVRT / ČTVRŤ / 1/4 / Č: → pizza_ctvrt (číslo hned za tím, "8x" = 8)
- NIKDY nezapisuj 0 pokud pizza na lístku je!

=== BURGER, BUŘTGULÁŠ, TALÍŘE — VELMI DŮLEŽITÉ ===
- BURGER / BURGR → burger (číslo za nebo před slovem)
- KRITICKÉ: "1" u burgeru bývá čtena jako "0" nebo přeskočena — pokud vidíš jakékoliv číslo u BURGER, zapiš ho!
- Zápis "1" nebo "1x" nebo ": 1" u burgeru = burger 1, NIKDY 0
- BURTGULÁŠ / BURTGULAS / BURGULÁŠ / BUŘTGULÁŠ / BURTGULÁS → burtgulas
- KRITICKÉ: "7" bývá čtena jako "2" — pokud vidíš "2x" u buřtguláše, zkontroluj znovu
- TALÍŘ / TALIRE / POČET TALÍŘŮ / TAL: → talire
- NIKDY nezapisuj 0 pokud číslo na lístku je — i "1" je číslo!

=== JMÉNA (SMĚNA) ===
Na směně pracují POUZE tyto osoby — žádná jiná jména neexistují:

  "Ráďa"  → variace: Ráďa, Rádá, Rada, Radi, Nada, Náda, Nade, Nadi
             (Ř bývá čteno jako N, Á jako A, Ď jako D)
  "Vendy" → variace: Vendy, Wendy, Vendi, Vends, Vend
             (začíná VEN nebo WEN)
  "Vali"  → variace: Vali, Valy, Voli
  "Věrka" → variace: Věrka, Věra, Verka, Vera
             (začíná VĚ nebo VER — NIKDY VEN!)
  "Renča" → variace: Renča, Renata, Renca, Renata

POSTUP: Přečti každé jméno na lístku → najdi nejpodobnější ze seznamu výše → zapiš správný tvar.
Pokud jméno vůbec neznáš → přiřaď nejbližší ze 5 možností, nikdy nevymýšlej nové.
NIKDY nepiš: Nada, Náda (→ je to Ráďa), ani žádné jiné jméno mimo seznam.
"""

        message = client.messages.create(
            model="claude-sonnet-4-20250514",
            max_tokens=1000,
            messages=[{
                "role": "user",
                "content": [
                    {
                        "type": "image",
                        "source": {
                            "type": "base64",
                            "media_type": media_type,
                            "data": img_data
                        }
                    },
                    {"type": "text", "text": prompt}
                ]
            }]
        )

        text = message.content[0].text.strip()
        text = re.sub(r"^```json\s*", "", text)
        text = re.sub(r"```$", "", text).strip()

        parsed = json.loads(text)
        return parsed, None

    except Exception as e:
        return None, str(e)


def parse_report_text(text):
    api_key = os.environ.get("ANTHROPIC_API_KEY", "")
    if not api_key:
        return None, "ANTHROPIC_API_KEY není nastaven"

    try:
        client = anthropic.Anthropic(api_key=api_key)

        prompt = f"""Přečti tento text denního reportu z restaurace a extrahuj údaje.
Odpověz POUZE platným JSON objektem, žádný jiný text.

Text reportu:
{text}

Formát odpovědi:
{{
  "datum": "DD.M" nebo null,
  "den": "název dne česky" nebo null,
  "smena": "jména oddělená čárkou" nebo null,
  "karty": číslo nebo 0,
  "kov": číslo nebo 0,
  "papir": číslo nebo 0,
  "vydaje": číslo nebo 0,
  "trzba": číslo nebo 0,
  "pk50_ks": počet kusů PK50 nebo 0,
  "pk100_ks": počet kusů PK100 nebo 0,
  "pizza_cela": číslo nebo 0,
  "pizza_ctvrt": číslo nebo 0,
  "burger": číslo nebo 0,
  "talire": číslo nebo 0,
  "burtgulas": číslo nebo 0
}}
"""

        message = client.messages.create(
            model="claude-sonnet-4-20250514",
            max_tokens=1000,
            messages=[{"role": "user", "content": prompt}]
        )

        text_resp = message.content[0].text.strip()
        text_resp = re.sub(r"^```json\s*", "", text_resp)
        text_resp = re.sub(r"```$", "", text_resp).strip()
        parsed = json.loads(text_resp)
        return parsed, None

    except Exception as e:
        return None, str(e)


def datum_to_iso(datum_str, year=None):
    if not datum_str:
        return None
    datum_str = str(datum_str).strip()
    for sep in ["/", ".", "-"]:
        parts = datum_str.split(sep)
        if len(parts) == 2:
            try:
                d, m = int(parts[0]), int(parts[1])
                if year is None:
                    year = date.today().year
                return date(year, m, d).isoformat()
            except Exception:
                pass
    return None


def build_report_from_parsed(parsed, year=None):
    datum_iso = datum_to_iso(parsed.get("datum"), year)

    karty   = float(parsed.get("karty", 0) or 0)
    kov     = float(parsed.get("kov", 0) or 0)
    papir   = float(parsed.get("papir", 0) or 0)
    vydaje  = float(parsed.get("vydaje", 0) or 0)
    hotovost = kov + papir
    trzba    = karty + hotovost

    pk50_ks  = int(parsed.get("pk50_ks", 0) or 0)
    pk100_ks = int(parsed.get("pk100_ks", 0) or 0)
    pk_celkem = pk50_ks * 50 + pk100_ks * 100
    trzba_vcpk = trzba + pk_celkem

    if parsed.get("trzba") and float(parsed.get("trzba")) > 0:
        trzba = float(parsed["trzba"])
        trzba_vcpk = trzba + pk_celkem

    smena = normalize_jmena(parsed.get("smena", ""))

    _DNY = ["pondělí","úterý","středa","čtvrtek","pátek","sobota","neděle"]
    if datum_iso:
        try:
            from datetime import date as _date
            den_auto = _DNY[_date.fromisoformat(datum_iso).weekday()]
        except Exception:
            den_auto = ""
    else:
        den_auto = ""

    return {
        "datum":       datum_iso,
        "den":         den_auto,
        "smena":       smena,
        "karty":       karty,
        "kov":         kov,
        "papir":       papir,
        "hotovost":    hotovost,
        "vydaje":      vydaje,
        "trzba":       trzba,
        "trzba_vcpk":  trzba_vcpk,
        "pk50_ks":     pk50_ks,
        "pk100_ks":    pk100_ks,
        "pk_celkem":   pk_celkem,
        "pizza_cela":  int(parsed.get("pizza_cela", 0) or 0),
        "pizza_ctvrt": int(parsed.get("pizza_ctvrt", 0) or 0),
        "burger":      int(parsed.get("burger", 0) or 0),
        "talire":      int(parsed.get("talire", 0) or 0),
        "burtgulas":   int(parsed.get("burtgulas", 0) or 0),
    }


# ═══════════════════════════════════════════════════════════════════════════════
#  ROUTES
# ═══════════════════════════════════════════════════════════════════════════════


# ── Login / Logout ──────────────────────────────────────────────────────────────
@app.route("/api/login", methods=["POST"])
def api_login():
    heslo = (request.json or {}).get("heslo", "")
    admin_pwd   = os.environ.get("PASSWORD_ADMIN", "")
    verunka_pwd = os.environ.get("PASSWORD_VERUNKA", "")
    ucetni_pwd  = os.environ.get("PASSWORD_UCETNI", "")

    if heslo and heslo == admin_pwd:
        session["role"] = "admin"
    elif heslo and heslo == verunka_pwd:
        session["role"] = "verunka"
    elif heslo and heslo == ucetni_pwd:
        session["role"] = "ucetni"
    else:
        return jsonify({"ok": False, "chyba": "Špatné heslo"}), 401

    role = session["role"]
    return jsonify({
        "ok": True,
        "role": role,
        "jmeno": ROLE_NAMES[role],
        "prava": get_prava_z_db().get(role, {}) if role != "admin" else "vse",
    })

@app.route("/api/logout", methods=["POST"])
def api_logout():
    session.clear()
    return jsonify({"ok": True})

@app.route("/api/me")
def api_me():
    """Vrátí info o přihlášeném uživateli."""
    role = session.get("role")
    if not role:
        return jsonify({"prihlasen": False})
    return jsonify({
        "prihlasen": True,
        "role": role,
        "jmeno": ROLE_NAMES.get(role, role),
        "prava": get_prava_z_db().get(role, {}) if role != "admin" else "vse",
    })

@app.route("/api/prava", methods=["GET"])
@vyzaduj_prihlaseni
def api_prava_get():
    if session.get("role") != "admin":
        return jsonify({"error": "Pouze admin"}), 403
    return jsonify(get_prava_z_db())

@app.route("/api/prava", methods=["POST"])
@vyzaduj_prihlaseni
def api_prava_set():
    if session.get("role") != "admin":
        return jsonify({"error": "Pouze admin"}), 403
    data = request.json or {}
    try:
        with get_db() as conn:
            for role, sekce_dict in data.items():
                if role not in ("verunka", "ucetni"):
                    continue
                for sekce, povoleno in sekce_dict.items():
                    conn.execute("""
    INSERT INTO prava (role, sekce, povoleno)
    VALUES (%s, %s, %s)
    ON CONFLICT (role, sekce) DO UPDATE SET povoleno = excluded.povoleno
""", (role, sekce, 1 if povoleno else 0))
        return jsonify({"ok": True})
    except Exception as e:
        return jsonify({"ok": False, "chyba": str(e)}), 500

@app.route("/")
def index():
    return render_template("index.html", config=load_config())



@app.route("/api/config", methods=["GET", "POST"])
@vyzaduj_prihlaseni
def api_config():
    if request.method == "GET":
        return jsonify(load_config())
    data = request.json
    cfg = load_config()
    if "firmy" in data:
        cfg["firmy"] = [f.strip().upper() for f in data["firmy"] if f.strip()]
    if "ico_map" in data:
        cfg["ico_map"] = data["ico_map"]
    if "app_nazev" in data:
        cfg["app_nazev"] = data["app_nazev"]
    if "terminal_limit" in data:
        cfg["terminal_limit"] = int(data["terminal_limit"] or 100000)
    if "dph_limit" in data:
        cfg["dph_limit"] = int(data["dph_limit"] or 2000000)
    if "terminal_prepnout" in data:
        firma = data["terminal_prepnout"]
        if not cfg.get("terminal_od"):
            cfg["terminal_od"] = {}
        from datetime import date as _date
        cfg["terminal_od"][firma] = _date.today().isoformat()
        cfg["terminal_aktivni"] = {f: (f == firma) for f in cfg.get("firmy", [])}
    save_config(cfg)
    return jsonify({"ok": True})

@app.route("/api/reporty/karty-stats")
@vyzaduj_prihlaseni
def api_karty_stats():
    import datetime as _dt
    cfg = load_config()
    firmy = cfg.get("firmy", [])
    terminal_od = cfg.get("terminal_od", {})
    terminal_limit = cfg.get("terminal_limit", 100000)
    dph_limit = cfg.get("dph_limit", 2000000)
    rok = str(_dt.date.today().year)
    result = {}
    with get_db() as conn:
        for firma in firmy:
            row = conn.execute("""
                SELECT COALESCE(SUM(karty),0) as total
                FROM reporty
                WHERE firma_zkratka=? AND datum>=?
            """, (firma, f"{rok}-01-01")).fetchone()
            rocni = float((row or {}).get("total", 0))

            od = terminal_od.get(firma, f"{rok}-01-01")
            mesic_str = _dt.date.today().strftime("%Y-%m")
            mesic_prvni = mesic_str + "-01"
            row2 = conn.execute("""
                SELECT COALESCE(SUM(karty),0) as total
                FROM reporty
                WHERE firma_zkratka=? AND datum>=?
            """, (firma, mesic_prvni)).fetchone()
            mesicni = float((row2 or {}).get("total", 0))

            row3 = conn.execute("""
                SELECT COALESCE(SUM(hotovost+karty),0) as total
                FROM reporty
                WHERE firma_zkratka=? AND datum>=?
            """, (firma, mesic_prvni)).fetchone()
            trzba_od = float((row3 or {}).get("total", 0))
            row4 = conn.execute("""
                SELECT COALESCE(SUM(karty),0) as k,
                       COALESCE(SUM(hotovost),0) as h,
                       COALESCE(SUM(hotovost+karty),0) as t
                FROM reporty
                WHERE firma_zkratka=? AND datum LIKE ?
            """, (firma, mesic_str + "%")).fetchone()
            karty_mesic = float((row4 or {}).get("k", 0))
            hot_mesic   = float((row4 or {}).get("h", 0))
            trzba_mesic = float((row4 or {}).get("t", 0))

            row5 = conn.execute("""
                SELECT COALESCE(SUM(karty),0) as k,
                       COALESCE(SUM(hotovost),0) as h,
                       COALESCE(SUM(hotovost+karty),0) as t
                FROM reporty
                WHERE firma_zkratka=? AND datum>=?
            """, (firma, f"{rok}-01-01")).fetchone()
            hot_rok   = float((row5 or {}).get("h", 0))
            trzba_rok = float((row5 or {}).get("t", 0))

            aktivni = cfg.get("terminal_aktivni", {}).get(firma, False)
            result[firma] = {
                "rocni": rocni,
                "mesicni": mesicni,
                "trzba_od": trzba_od,
                "karty_mesic": karty_mesic,
                "hot_mesic": hot_mesic,
                "trzba_mesic": trzba_mesic,
                "hot_rok": hot_rok,
                "trzba_rok": trzba_rok,
                "terminal_od": od,
                "terminal_limit": terminal_limit,
                "dph_limit": dph_limit,
                "aktivni": aktivni,
            }
    return jsonify(result)

@app.route("/api/dashboard")
@vyzaduj_prihlaseni
def api_dashboard():
    firma = request.args.get("firma", "")
    with get_db() as conn:
        mesic = date.today().strftime("%Y-%m")
        where_firma = "AND firma_zkratka=?" if firma else ""
        params_base = (firma,) if firma else ()

        like_cond = "AND datum_vystaveni::text LIKE ?" if _USE_PG else "AND datum_vystaveni LIKE ?"
        row = conn.execute(f"""
            SELECT COUNT(*) as pocet, COALESCE(SUM(celkem_s_dph),0) as vydaje
            FROM faktury
            WHERE 1=1 {like_cond} {where_firma}
        """, (mesic + "%",) + params_base).fetchone()
        pocet_mesic  = row["pocet"]  if isinstance(row, dict) else row[0]
        vydaje_mesic = row["vydaje"] if isinstance(row, dict) else row[1]

        row2 = conn.execute(f"""
            SELECT COUNT(*) as pocet, COALESCE(SUM(celkem_s_dph),0) as castka
            FROM faktury WHERE stav='po_splatnosti' {where_firma}
        """, params_base).fetchone()
        pocet_po_spl  = row2["pocet"]  if isinstance(row2, dict) else row2[0]
        castka_po_spl = row2["castka"] if isinstance(row2, dict) else row2[1]

        datum_filter = "AND datum_vystaveni::date >= CURRENT_DATE - INTERVAL '12 months'" if _USE_PG else "AND datum_vystaveni >= date('now','-12 months')"
        graf_sql = "TO_CHAR(NULLIF(datum_vystaveni,'')::date,'YYYY-MM')" if _USE_PG else "strftime('%Y-%m', datum_vystaveni)"
        graf = conn.execute(f"""
            SELECT {graf_sql} as m, COALESCE(SUM(celkem_s_dph),0) as castka
            FROM faktury
            WHERE datum_vystaveni IS NOT NULL AND datum_vystaveni != '' {datum_filter} {where_firma}
            GROUP BY m ORDER BY m
        """, params_base).fetchall()

        posledni = conn.execute(f"""
            SELECT id, dodavatel, cislo_faktury, firma_zkratka, datum_vystaveni,
                   datum_splatnosti, celkem_s_dph, stav
            FROM faktury {('WHERE firma_zkratka=?' if firma else '')}
            ORDER BY created_at DESC LIMIT 5
        """, params_base).fetchall()

        karty_row = conn.execute("""
            SELECT COALESCE(SUM(karty),0) as karty
            FROM reporty
            WHERE datum >= CURRENT_DATE - INTERVAL '12 months'
        """).fetchone()
        karty_12m = karty_row["karty"] if isinstance(karty_row, dict) else karty_row[0]

    def graf_row(r):
        if isinstance(r, dict):
            return {"mesic": r["m"], "castka": round(r["castka"], 2)}
        return {"mesic": r[0], "castka": round(r[1], 2)}

    return jsonify({
        "vydaje_mesic": round(vydaje_mesic, 2),
        "pocet_mesic": pocet_mesic,
        "pocet_po_splatnosti": pocet_po_spl,
        "castka_po_splatnosti": round(castka_po_spl, 2),
        "graf": [graf_row(r) for r in graf],
        "posledni_faktury": [dict(r) for r in posledni],
        "karty_12m": round(karty_12m, 2),
        "karty_limit": 1500000,
    })

@app.route("/api/nastenka-rocni-prehled")
@vyzaduj_prihlaseni
def api_nastenka_rocni_prehled():
    """Roční přehled tržeb po měsících pro tabulku na nástěnce."""
    firma = request.args.get("firma", "")
    try:
        import psycopg2 as _pg2
        db_url = os.environ.get("DATABASE_URL", "")
        conn = _pg2.connect(db_url)
        cur = conn.cursor()
        fw = "AND firma_zkratka=%s" if firma else ""
        params = [firma] if firma else []
        cur.execute(f"""
            SELECT
                EXTRACT(YEAR FROM NULLIF(datum,'')::date)::int as rok,
                EXTRACT(MONTH FROM NULLIF(datum,'')::date)::int as mesic,
                ROUND(SUM(trzba_vcpk)::numeric, 0) as trzba,
                COUNT(*) as dni
            FROM reporty
            WHERE datum IS NOT NULL AND datum != '' {fw}
            GROUP BY rok, mesic
            ORDER BY rok, mesic
        """, params)
        rows = cur.fetchall()
        conn.close()

        # Sestav strukturu {rok: {mesic: {trzba, dni, prumer}}}
        data = {}
        for rok, mesic, trzba, dni in rows:
            if rok not in data:
                data[rok] = {}
            trzba_f = float(trzba or 0)
            dni_i = int(dni or 1)
            data[rok][mesic] = {
                "trzba": trzba_f,
                "dni": dni_i,
                "prumer": round(trzba_f / dni_i, 0) if dni_i else 0
            }

        return jsonify({"ok": True, "data": data})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)})

@app.route("/api/nastenka-backup-check")
@vyzaduj_prihlaseni
def api_nastenka_backup_check():
    """Zkontroluje datum poslední GCS zálohy."""
    try:
        from google.cloud import storage as gcs
        bucket_name = os.environ.get("GCS_BUCKET", "faktury-makro-docs")
        client = gcs.Client()
        bucket = client.bucket(bucket_name)
        blobs = sorted(
            [b for b in bucket.list_blobs(prefix="zalohy/")],
            key=lambda b: b.updated, reverse=True
        )
        if not blobs:
            return jsonify({"ok": False, "zprava": "Žádné zálohy nenalezeny", "dnu_od": 999})
        posledni = blobs[0]
        import datetime as _dt
        dnu = (_dt.datetime.now(_dt.timezone.utc) - posledni.updated).days
        return jsonify({
            "ok": dnu <= 7,
            "posledni": posledni.name,
            "dnu_od": dnu,
            "zprava": f"Záloha stará {dnu} dní" if dnu > 7 else f"OK — záloha před {dnu} dny"
        })
    except Exception as e:
        return jsonify({"ok": False, "zprava": str(e), "dnu_od": -1})

@app.route("/api/nastenka-check")
@vyzaduj_prihlaseni
def api_nastenka_check():
    import datetime as _dt
    dnes = _dt.date.today().isoformat()
    cfg = load_config()
    terminal_limit = cfg.get("terminal_limit", 100000)
    dph_limit = cfg.get("dph_limit", 1800000)
    rok = str(_dt.date.today().year)
    mesic_str = _dt.date.today().strftime("%Y-%m")
    mesic_prvni = mesic_str + "-01"

    result = {}

    with get_db() as conn:
        # 1. Přijaté faktury po splatnosti
        r = conn.execute("""
            SELECT COUNT(*) as pocet, COALESCE(SUM(celkem_s_dph),0) as castka
            FROM faktury WHERE stav='po_splatnosti'
        """).fetchone()
        pocet_po_spl = int(_first_val(r) if not isinstance(r, dict) else r["pocet"])
        castka_po_spl = float(r["castka"] if isinstance(r, dict) else r[1])
        rows_po_spl = conn.execute("""
            SELECT id, dodavatel, cislo_faktury, datum_splatnosti, celkem_s_dph, firma_zkratka
            FROM faktury WHERE stav='po_splatnosti'
            ORDER BY datum_splatnosti ASC LIMIT 5
        """).fetchall()
        result["faktury_po_splatnosti"] = {
            "pocet": pocet_po_spl,
            "castka": round(castka_po_spl, 2),
            "items": [dict(r) for r in rows_po_spl],
            "stav": "ok" if pocet_po_spl == 0 else "error",
        }

        # 2. Přijaté faktury čekající na zaplacení (stav ceka, datum splatnosti budoucí)
        # Firemní — přijaté faktury čekající
        r2a = conn.execute("""
            SELECT COUNT(*) as pocet, COALESCE(SUM(celkem_s_dph),0) as castka
            FROM faktury WHERE stav='ceka'
        """).fetchone()
        pocet_fa = int(r2a["pocet"] if isinstance(r2a, dict) else r2a[0])
        castka_fa = float(r2a["castka"] if isinstance(r2a, dict) else r2a[1])

        # Firemní — provozní výdaje nezaplacené
        r2b = conn.execute("""
            SELECT COUNT(*) as pocet, COALESCE(SUM(castka),0) as castka
            FROM vydaje WHERE stav='nezaplaceno' AND COALESCE(typ,'provozni')='provozni'
        """).fetchone()
        pocet_vyd = int(r2b["pocet"] if isinstance(r2b, dict) else r2b[0])
        castka_vyd = float(r2b["castka"] if isinstance(r2b, dict) else r2b[1])

        pocet_firmy = pocet_fa + pocet_vyd
        castka_firmy = castka_fa + castka_vyd
        result["cekajici_firemni"] = {
            "pocet": pocet_firmy,
            "pocet_faktur": pocet_fa,
            "pocet_vydaju": pocet_vyd,
            "castka": round(castka_firmy, 2),
            "castka_faktur": round(castka_fa, 2),
            "castka_vydaju": round(castka_vyd, 2),
            "stav": "ok" if pocet_firmy == 0 else "warning",
        }

        # Soukromé výdaje nezaplacené
        r2c = conn.execute("""
            SELECT COUNT(*) as pocet, COALESCE(SUM(castka),0) as castka
            FROM vydaje WHERE stav='nezaplaceno' AND typ='soukrome'
        """).fetchone()
        pocet_soukr = int(r2c["pocet"] if isinstance(r2c, dict) else r2c[0])
        castka_soukr = float(r2c["castka"] if isinstance(r2c, dict) else r2c[1])
        result["cekajici_soukrome"] = {
            "pocet": pocet_soukr,
            "castka": round(castka_soukr, 2),
            "stav": "ok" if pocet_soukr == 0 else "warning",
        }

        # 3. Duplicitní faktury nevyřešené
        r3 = conn.execute("""
            SELECT COUNT(*) as pocet, COALESCE(SUM(celkem_s_dph),0) as castka
            FROM faktury WHERE stav='duplikat'
        """).fetchone()
        pocet_dup = int(r3["pocet"] if isinstance(r3, dict) else r3[0])
        castka_dup = float(r3["castka"] if isinstance(r3, dict) else r3[1])
        result["duplicitni_faktury"] = {
            "pocet": pocet_dup,
            "castka": round(castka_dup, 2),
            "stav": "ok" if pocet_dup == 0 else "error",
        }

        # Faktury blížící se splatnosti (do 7 dní)
        blizi_cond = "AND datum_splatnosti::date BETWEEN CURRENT_DATE AND CURRENT_DATE + INTERVAL '7 days'" if _USE_PG else "AND datum_splatnosti BETWEEN date('now') AND date('now','+7 days')"
        r_blizi = conn.execute(f"""
            SELECT COUNT(*) as pocet, COALESCE(SUM(celkem_s_dph),0) as castka
            FROM faktury WHERE stav='ceka'
            AND datum_splatnosti IS NOT NULL AND datum_splatnosti != ''
            {blizi_cond}
        """).fetchone()
        pocet_blizi = int(r_blizi["pocet"] if isinstance(r_blizi, dict) else r_blizi[0])
        castka_blizi = float(r_blizi["castka"] if isinstance(r_blizi, dict) else r_blizi[1])
        rows_blizi = conn.execute(f"""
            SELECT id, dodavatel, cislo_faktury, datum_splatnosti, celkem_s_dph, firma_zkratka
            FROM faktury WHERE stav='ceka'
            AND datum_splatnosti IS NOT NULL AND datum_splatnosti != ''
            {blizi_cond}
            ORDER BY datum_splatnosti ASC LIMIT 5
        """).fetchall()
        result["faktury_blizi_splatnost"] = {
            "pocet": pocet_blizi,
            "castka": round(castka_blizi, 2),
            "items": [dict(r) for r in rows_blizi],
            "stav": "ok" if pocet_blizi == 0 else "warning",
        }

        # 4. Vystavené faktury po splatnosti (Bauhaus nezaplatil)
        datum_cond = "AND datum_splatnosti::date < CURRENT_DATE" if _USE_PG else "AND datum_splatnosti < date('now')"
        datum_not_empty = "AND datum_splatnosti IS NOT NULL AND datum_splatnosti != ''" 
        r4 = conn.execute(f"""
            SELECT COUNT(*) as pocet, COALESCE(SUM(castka),0) as castka
            FROM vystavene_faktury
            WHERE stav='nezaplaceno'
              {datum_not_empty}
              {datum_cond}
        """).fetchone()
        pocet_vyst = int(r4["pocet"] if isinstance(r4, dict) else r4[0])
        castka_vyst = float(r4["castka"] if isinstance(r4, dict) else r4[1])
        rows_vyst = conn.execute(f"""
            SELECT id, odberatel, cislo_faktury, datum_splatnosti, castka, firma_zkratka
            FROM vystavene_faktury
            WHERE stav='nezaplaceno'
              {datum_not_empty}
              {datum_cond}
            ORDER BY datum_splatnosti ASC LIMIT 5
        """).fetchall()
        result["vystavene_po_splatnosti"] = {
            "pocet": pocet_vyst,
            "castka": round(castka_vyst, 2),
            "items": [dict(r) for r in rows_vyst],
            "stav": "ok" if pocet_vyst == 0 else "error",
        }

        # 5. Terminál — karty aktuální měsíc
        r5 = conn.execute("""
            SELECT COALESCE(SUM(karty),0) as total FROM reporty
            WHERE datum >= ?
        """, (mesic_prvni,)).fetchone()
        karty_mesic = float(_first_val(r5))
        terminal_pct = round(karty_mesic / terminal_limit * 100, 1) if terminal_limit else 0
        result["terminal_box"] = {
            "castka": round(karty_mesic, 2),
            "limit": terminal_limit,
            "procent": terminal_pct,
            "stav": "ok" if terminal_pct < 80 else ("error" if terminal_pct >= 100 else "warning"),
        }

        # 6. DPH rok — karty od 1.1.
        r6 = conn.execute("""
            SELECT COALESCE(SUM(karty),0) as total FROM reporty
            WHERE datum >= ?
        """, (f"{rok}-01-01",)).fetchone()
        karty_rok = float(_first_val(r6))
        dph_pct = round(karty_rok / dph_limit * 100, 1) if dph_limit else 0
        result["dph_limit"] = {
            "castka": round(karty_rok, 2),
            "limit": dph_limit,
            "procent": dph_pct,
            "stav": "ok" if dph_pct < 75 else ("error" if dph_pct >= 100 else "warning"),
        }

        # 7. Duplicitní reporty
        r7 = conn.execute("""
            SELECT COUNT(*) as pocet FROM reporty
            WHERE duplicita_id IS NOT NULL
        """).fetchone()
        pocet_dup_rep = int(r7["pocet"] if isinstance(r7, dict) else r7[0])
        result["duplicitni_reporty"] = {
            "pocet": pocet_dup_rep,
            "stav": "ok" if pocet_dup_rep == 0 else "warning",
        }

    # 8. Záloha starší 7 dní
    try:
        bucket = get_gcs_client()
        zaloha_stav = "ok"
        zaloha_info = ""
        if bucket:
            blobs = sorted(
                [b for b in bucket.list_blobs(prefix="zalohy/") if b.name.endswith(".json") or b.name.endswith(".sql")],
                key=lambda b: b.updated, reverse=True
            )
            if blobs:
                last = blobs[0]
                stari = (_dt.datetime.now(_dt.timezone.utc) - last.updated).days
                zaloha_info = last.name.replace("zalohy/", "")
                zaloha_stav = "ok" if stari <= 7 else "warning"
                result["zaloha"] = {
                    "stav": zaloha_stav,
                    "soubor": zaloha_info,
                    "dni_stari": stari,
                }
            else:
                result["zaloha"] = {"stav": "warning", "soubor": "", "dni_stari": 999}
        else:
            result["zaloha"] = {"stav": "warning", "soubor": "", "dni_stari": -1}
    except Exception:
        result["zaloha"] = {"stav": "warning", "soubor": "", "dni_stari": -1}

    # 9. Terminál — duplikát z Reportů
    cfg = load_config()
    terminal_limit = cfg.get("terminal_limit", 100000)
    firmy = cfg.get("firmy", [])
    terminal_od = cfg.get("terminal_od", {})
    with get_db() as conn:
        terminal_firmy = {}
        for firma in firmy:
            r_t = conn.execute("""
                SELECT COALESCE(SUM(karty),0) as total FROM reporty
                WHERE firma_zkratka=? AND datum>=?
            """, (firma, mesic_prvni)).fetchone()
            karty_f = float(_first_val(r_t))
            pct_f = round(karty_f / terminal_limit * 100, 1) if terminal_limit else 0
            aktivni = cfg.get("terminal_aktivni", {}).get(firma, False)
            terminal_firmy[firma] = {
                "castka": round(karty_f, 2),
                "procent": pct_f,
                "aktivni": aktivni,
                "stav": "ok" if pct_f < 80 else ("error" if pct_f >= 100 else "warning"),
            }
        result["terminal_firmy"] = terminal_firmy
        result["terminal_limit"] = terminal_limit

        # 10. P&L — aktuální rok
        rok_od = f"{rok}-01-01"
        rok_do = f"{rok}-12-31"
        r_trzba = conn.execute("""
            SELECT COALESCE(SUM(trzba_vcpk),0) as total FROM reporty
            WHERE datum>=? AND datum<=?
        """, (rok_od, rok_do)).fetchone()
        trzba_rok = float(_first_val(r_trzba))

        r_fakt = conn.execute("""
            SELECT COALESCE(SUM(celkem_s_dph),0) as total FROM faktury
            WHERE datum_vystaveni>=? AND datum_vystaveni<=?
        """, (rok_od, rok_do)).fetchone()
        naklady_faktury = float(_first_val(r_fakt))

        r_vyd = conn.execute("""
            SELECT COALESCE(SUM(castka),0) as total FROM vydaje
            WHERE datum>=? AND datum<=? AND COALESCE(typ,'provozni')='provozni'
        """, (rok_od, rok_do)).fetchone()
        naklady_vydaje = float(_first_val(r_vyd))

        r_vypl = conn.execute("""
            SELECT COALESCE(SUM(castka),0) as total FROM vyplaty
            WHERE datum>=? AND datum<=?
        """, (rok_od, rok_do)).fetchone()
        naklady_vyplaty = float(_first_val(r_vypl))

        # Odvody za rok
        odvody_rok = sum(
            _spocitej_pausaly_mesic(conn, f"{rok}-{mi:02d}-01")
            for mi in range(1, _dt.date.today().month + 1)
        )

        naklady_celkem = naklady_faktury + naklady_vydaje + naklady_vyplaty + odvody_rok
        pl_rok = trzba_rok - naklady_celkem

        result["pl"] = {
            "trzba_rok": round(trzba_rok, 0),
            "naklady_faktury": round(naklady_faktury, 0),
            "naklady_vydaje": round(naklady_vydaje, 0),
            "naklady_vyplaty": round(naklady_vyplaty, 0),
            "naklady_odvody": round(odvody_rok, 0),
            "naklady_celkem": round(naklady_celkem, 0),
            "pl_rok": round(pl_rok, 0),
            "stav": "ok" if pl_rok >= 0 else "error",
        }

        # 11. Náklady — faktury + výdaje po měsících (aktuální rok)
        if _USE_PG:
            mesic_sql = "TO_CHAR(NULLIF(datum_vystaveni,'')::date,'YYYY-MM')"
            mesic_sql_vyd = "TO_CHAR(NULLIF(datum,'')::date,'YYYY-MM')"
        else:
            mesic_sql = "strftime('%Y-%m', datum_vystaveni)"
            mesic_sql_vyd = "strftime('%Y-%m', datum)"

        r_fakt_m = conn.execute(f"""
            SELECT {mesic_sql} as m, COALESCE(SUM(celkem_s_dph),0) as castka
            FROM faktury WHERE datum_vystaveni>=? AND datum_vystaveni<=?
            GROUP BY m ORDER BY m
        """, (rok_od, rok_do)).fetchall()

        r_vyd_m = conn.execute(f"""
            SELECT {mesic_sql_vyd} as m, COALESCE(SUM(castka),0) as castka
            FROM vydaje WHERE datum>=? AND datum<=? AND COALESCE(typ,'provozni')='provozni'
            GROUP BY m ORDER BY m
        """, (rok_od, rok_do)).fetchall()

        naklady_mesice = {}
        for r in r_fakt_m:
            m = r["m"] if isinstance(r, dict) else r[0]
            v = float(r["castka"] if isinstance(r, dict) else r[1])
            if m: naklady_mesice[m] = naklady_mesice.get(m, {"faktury": 0, "vydaje": 0})
            if m: naklady_mesice[m]["faktury"] = round(v, 0)
        for r in r_vyd_m:
            m = r["m"] if isinstance(r, dict) else r[0]
            v = float(r["castka"] if isinstance(r, dict) else r[1])
            if m:
                if m not in naklady_mesice: naklady_mesice[m] = {"faktury": 0, "vydaje": 0}
                naklady_mesice[m]["vydaje"] = round(v, 0)

        result["naklady_mesice"] = [
            {"mesic": m, "faktury": d["faktury"], "vydaje": d["vydaje"],
             "celkem": round(d["faktury"] + d["vydaje"], 0)}
            for m, d in sorted(naklady_mesice.items())
        ]

    return jsonify(result)


@app.route("/api/faktury")
@vyzaduj_prihlaseni
def api_faktury():
    firma   = request.args.get("firma", "")
    stav    = request.args.get("stav", "")
    od      = request.args.get("od", "")
    do_     = request.args.get("do", "")
    hledat  = request.args.get("q", "")

    clauses = []
    params  = []
    if firma:
        clauses.append("firma_zkratka=?"); params.append(firma)
    if stav:
        clauses.append("stav=?"); params.append(stav)
    if od:
        clauses.append("datum_vystaveni>=?"); params.append(od)
    if do_:
        clauses.append("datum_vystaveni<=?"); params.append(do_)
    if hledat:
        clauses.append("(dodavatel LIKE ? OR cislo_faktury LIKE ?)")
        params += [f"%{hledat}%", f"%{hledat}%"]

    where = ("WHERE " + " AND ".join(clauses)) if clauses else ""

    with get_db() as conn:
        rows = conn.execute(f"""
            SELECT id, firma_zkratka, dodavatel, cislo_faktury,
                   datum_vystaveni, datum_splatnosti, celkem_s_dph, stav, zdroj, duplicita_id,
                   soubor_url, soubor_cesta
            FROM faktury {where}
            ORDER BY datum_vystaveni DESC, created_at DESC
        """, params).fetchall()
        total_row = conn.execute(f"SELECT COALESCE(SUM(celkem_s_dph),0) as total FROM faktury {where}", params).fetchone()
        total = _first_val(total_row)

    # Fallback: pokud soubor_url chybí ale je soubor_cesta, zkus GCS
    # Omezit na max 20 volání GCS aby soupis nebyl pomalý
    gcs_calls = 0
    result = []
    for r in rows:
        d = dict(r)
        if not d.get("soubor_url") and d.get("soubor_cesta") and gcs_calls < 20:
            gcs_url = get_gcs_url(d["soubor_cesta"])
            if gcs_url:
                d["soubor_url"] = gcs_url
            gcs_calls += 1
        result.append(d)

    return jsonify({
        "faktury": result,
        "celkem": round(total, 2)
    })

@app.route("/api/faktury/<int:fid>")
@vyzaduj_prihlaseni
def api_faktura_detail(fid):
    with get_db() as conn:
        f = conn.execute("SELECT * FROM faktury WHERE id=?", (fid,)).fetchone()
        if not f:
            return jsonify({"error": "Nenalezeno"}), 404
        polozky = conn.execute("""
            SELECT p.*, z.nazev_canonical as zbozi_nazev
            FROM polozky p
            LEFT JOIN zbozi z ON z.id = p.zbozi_id
            WHERE p.faktura_id=?
        """, (fid,)).fetchall()
    faktura_dict = dict(f)
    if not faktura_dict.get("soubor_url") and faktura_dict.get("soubor_cesta"):
        gcs_url = get_gcs_url(faktura_dict["soubor_cesta"])
        if gcs_url:
            faktura_dict["soubor_url"] = gcs_url
    return jsonify({"faktura": faktura_dict, "polozky": [dict(p) for p in polozky]})

@app.route("/api/faktury/<int:fid>/stav", methods=["POST"])
@vyzaduj_prihlaseni
def api_faktura_stav(fid):
    stav = request.json.get("stav")
    if stav not in ("ceka", "zaplaceno", "po_splatnosti"):
        return jsonify({"error": "Neplatný stav"}), 400
    with get_db() as conn:
        conn.execute("UPDATE faktury SET stav=? WHERE id=?", (stav, fid))
    return jsonify({"ok": True})

@app.route("/api/faktury/<int:fid>", methods=["DELETE"])
def api_faktura_delete(fid):
    reset_drive = request.args.get("reset_drive", "0") == "1"
    with get_db() as conn:
        row = conn.execute("SELECT soubor_cesta, zdroj FROM faktury WHERE id=?", (fid,)).fetchone()
        conn.execute("DELETE FROM faktury WHERE id=?", (fid,))
        if reset_drive and row:
            zdroj = row["zdroj"] if isinstance(row, dict) else row[1]
            soubor_cesta = row["soubor_cesta"] if isinstance(row, dict) else row[0]
            if zdroj == "drive_auto" and soubor_cesta:
                nazev = soubor_cesta.split("/")[-1]
                try:
                    drive_svc, _ = get_drive_service()
                    if drive_svc:
                        results = drive_svc.files().list(
                            q=f"name='{nazev}' and trashed=false",
                            fields="files(id,name)"
                        ).execute()
                        files = results.get("files", [])
                        for df in files:
                            conn.execute("DELETE FROM drive_zpracovane WHERE file_id=?", (df["id"],))
                            app.logger.info(f"Drive reset: smazán záznam {df['id']} ({nazev})")
                except Exception as e:
                    app.logger.warning(f"Drive reset chyba: {e}")
    if row:
        soubor_cesta = row["soubor_cesta"] if isinstance(row, dict) else row[0]
        if soubor_cesta:
            path = os.path.join(UPLOAD_DIR, soubor_cesta)
            if os.path.exists(path):
                os.remove(path)
    return jsonify({"ok": True})

@app.route("/api/faktury/<int:fid>", methods=["PUT"])
@vyzaduj_prihlaseni
def api_faktura_update(fid):
    data = request.json
    fields = ["firma_zkratka","dodavatel","cislo_faktury","datum_vystaveni",
              "datum_splatnosti","zpusob_uhrady","stav","celkem_s_dph","duplicita_id"]
    set_parts = [f"{f}=?" for f in fields if f in data]
    vals = [data[f] for f in fields if f in data]
    if set_parts:
        vals.append(fid)
        with get_db() as conn:
            conn.execute(f"UPDATE faktury SET {','.join(set_parts)} WHERE id=?", vals)

    polozky = data.get("polozky")
    if polozky is not None:
        with get_db() as conn:
            conn.execute("DELETE FROM polozky WHERE faktura_id=?", (fid,))
            for p in polozky:
                nazev = (p.get("nazev") or "").strip()
                if not nazev: continue
                mnozstvi = float(p.get("mnozstvi") or 1)
                celkem   = float(p.get("celkem_s_dph") or 0)
                cena_j   = float(p.get("cena_za_jednotku_s_dph") or 0)
                if cena_j == 0 and mnozstvi:
                    cena_j = celkem / mnozstvi
                jed = (p.get("jednotka") or "").strip()
                zbozi_id = _get_or_create_zbozi(conn, nazev)
                conn.execute("""
                    INSERT INTO polozky (faktura_id, nazev, mnozstvi, jednotka,
                        cena_za_jednotku_s_dph, celkem_s_dph, zbozi_id)
                    VALUES (?,?,?,?,?,?,?)
                """, (fid, nazev, mnozstvi, jed, round(cena_j,4), round(celkem,2), zbozi_id))
            recalc_faktura_total(conn, fid)
    return jsonify({"ok": True})

# ── API: výplaty ──────────────────────────────────────────────────────────────
def _spocitej_pausaly_mesic_jmeno(conn, jmeno, mesic_od):
    """Paušály pro konkrétního zaměstnance k danému měsíci."""
    rows = conn.execute("""
        SELECT nazev, castka, platnost_od
        FROM pausalni_odvody
        WHERE jmeno=? AND platnost_od <= ?
        ORDER BY nazev, platnost_od DESC
    """, (jmeno, mesic_od)).fetchall()
    seen = set()
    total = 0.0
    for r in rows:
        nazev = r["nazev"] if isinstance(r, dict) else r[0]
        castka = float(r["castka"] if isinstance(r, dict) else r[1])
        if nazev not in seen:
            seen.add(nazev)
            total += castka
    return total


def _spocitej_pausaly_mesic(conn, mesic_od):
    """Vrátí součet paušálů platných k danému měsíci (bere nejnovější platnost_od <= mesic_od pro každý jmeno+nazev)."""
    rows = conn.execute("""
        SELECT jmeno, nazev, castka, platnost_od
        FROM pausalni_odvody
        WHERE platnost_od <= ?
        ORDER BY jmeno, nazev, platnost_od DESC
    """, (mesic_od,)).fetchall()
    # Pro každý (jmeno, nazev) vezmi jen nejnovější
    seen = set()
    total = 0.0
    for r in rows:
        jmeno = r["jmeno"] if isinstance(r, dict) else r[0]
        nazev = r["nazev"] if isinstance(r, dict) else r[1]
        castka = float(r["castka"] if isinstance(r, dict) else r[2])
        key = (jmeno, nazev)
        if key not in seen:
            seen.add(key)
            total += castka
    return total


@app.route("/api/vyplaty/prehled")
@vyzaduj_prihlaseni
def api_vyplaty_prehled():
    """Přehled zaměstnanců — souhrn za měsíc a rok + poslední 2 výplaty."""
    import datetime as _dt
    dnes = _dt.date.today()
    mesic_od = f"{dnes.year}-{dnes.month:02d}-01"
    rok_od   = f"{dnes.year}-01-01"
    rok_do   = f"{dnes.year}-12-31"
    with get_db() as conn:
        # Seznam zaměstnanců
        jmena = [r["jmeno"] if isinstance(r, dict) else r[0]
                 for r in conn.execute("SELECT DISTINCT jmeno FROM vyplaty ORDER BY jmeno").fetchall()]

        result = []
        for jmeno in jmena:
            # Částka za aktuální měsíc
            r_mes = conn.execute(
                "SELECT COALESCE(SUM(castka),0) as total FROM vyplaty WHERE jmeno=? AND datum>=?",
                (jmeno, mesic_od)
            ).fetchone()
            castka_mesic = float(_first_val(r_mes))

            # Částka za rok
            r_rok = conn.execute(
                "SELECT COALESCE(SUM(castka),0) as total FROM vyplaty WHERE jmeno=? AND datum>=? AND datum<=?",
                (jmeno, rok_od, rok_do)
            ).fetchone()
            castka_rok = float(_first_val(r_rok))

            # Poslední 2 výplaty
            posledni = conn.execute(
                "SELECT datum, castka FROM vyplaty WHERE jmeno=? ORDER BY datum DESC, id DESC LIMIT 2",
                (jmeno,)
            ).fetchall()
            posledni_list = [{"datum": r["datum"] if isinstance(r, dict) else r[0],
                              "castka": float(r["castka"] if isinstance(r, dict) else r[1])}
                             for r in posledni]

            # Paušály pro tohoto zaměstnance
            odvody_zam_mesic = _spocitej_pausaly_mesic_jmeno(conn, jmeno, mesic_od)
            odvody_zam_rok = sum(
                _spocitej_pausaly_mesic_jmeno(conn, jmeno, f"{dnes.year}-{mi:02d}-01")
                for mi in range(1, dnes.month + 1)
            )

            result.append({
                "jmeno": jmeno,
                "castka_mesic": round(castka_mesic, 2),
                "castka_rok": round(castka_rok, 2),
                "odvody_mesic": round(odvody_zam_mesic, 2),
                "castka_rok_s_odvody": round(castka_rok + odvody_zam_rok, 2),
                "ma_odvody": odvody_zam_mesic > 0,
                "posledni": posledni_list,
            })

        # Celkem za měsíc a rok (všichni zaměstnanci)
        r_total_mes = conn.execute(
            "SELECT COALESCE(SUM(castka),0) as total FROM vyplaty WHERE datum>=?", (mesic_od,)
        ).fetchone()
        r_total_rok = conn.execute(
            "SELECT COALESCE(SUM(castka),0) as total FROM vyplaty WHERE datum>=? AND datum<=?", (rok_od, rok_do)
        ).fetchone()
        total_mesic = float(_first_val(r_total_mes))
        total_rok   = float(_first_val(r_total_rok))

        # Paušály pro aktuální měsíc — vezmi platnou částku k 1. dni měsíce
        odvody_mesic = _spocitej_pausaly_mesic(conn, mesic_od)

        # Paušály za rok — pro každý měsíc zvlášť
        odvody_rok = 0.0
        for mi in range(1, dnes.month + 1):
            m_od = f"{dnes.year}-{mi:02d}-01"
            odvody_rok += _spocitej_pausaly_mesic(conn, m_od)

    return jsonify({
        "zamestnanci": result,
        "souhrn": {
            "mesic_bez_odvodu": round(total_mesic, 2),
            "mesic_s_odvody":   round(total_mesic + odvody_mesic, 2),
            "rok_bez_odvodu":   round(total_rok, 2),
            "rok_s_odvody":     round(total_rok + odvody_rok, 2),
            "odvody_mesic":     round(odvody_mesic, 2),
        }
    })

@app.route("/api/vyplaty/mesice/<jmeno>")
@vyzaduj_prihlaseni
def api_vyplaty_mesice(jmeno):
    """Výplaty zaměstnance seskupené po měsících."""
    with get_db() as conn:
        if _USE_PG:
            rows = conn.execute("""
                SELECT TO_CHAR(NULLIF(datum,'')::date,'YYYY-MM') as mesic,
                       COALESCE(SUM(castka),0) as castka,
                       COUNT(*) as pocet
                FROM vyplaty WHERE jmeno=?
                GROUP BY mesic ORDER BY mesic DESC
            """, (jmeno,)).fetchall()
        else:
            rows = conn.execute("""
                SELECT strftime('%Y-%m', datum) as mesic,
                       COALESCE(SUM(castka),0) as castka,
                       COUNT(*) as pocet
                FROM vyplaty WHERE jmeno=?
                GROUP BY mesic ORDER BY mesic DESC
            """, (jmeno,)).fetchall()
        detail = conn.execute(
            "SELECT * FROM vyplaty WHERE jmeno=? ORDER BY datum DESC, id DESC", (jmeno,)
        ).fetchall()
    return jsonify({
        "mesice": [dict(r) for r in rows],
        "vyplaty": [dict(r) for r in detail],
    })

@app.route("/api/vyplaty/zamestnanci", methods=["GET"])
@vyzaduj_prihlaseni
def api_vyplaty_zamestnanci():
    with get_db() as conn:
        rows = conn.execute("SELECT DISTINCT jmeno FROM vyplaty ORDER BY jmeno").fetchall()
    return jsonify({"jmena": [r["jmeno"] for r in rows]})

@app.route("/api/vyplaty", methods=["GET"])
@vyzaduj_prihlaseni
def api_vyplaty():
    try:
        firma = request.args.get("firma", "")
        jmeno = request.args.get("jmeno", "")
        od    = request.args.get("od", "")
        do_   = request.args.get("do", "")
        clauses, params = [], []
        if firma: clauses.append("firma_zkratka=?"); params.append(firma)
        if jmeno: clauses.append("jmeno=?"); params.append(jmeno)
        if od:    clauses.append("datum>=?"); params.append(od)
        if do_:   clauses.append("datum<=?"); params.append(do_)
        where = ("WHERE " + " AND ".join(clauses)) if clauses else ""
        with get_db() as conn:
            rows = conn.execute(f"""
                SELECT * FROM vyplaty {where} ORDER BY datum DESC, created_at DESC
            """, params).fetchall()
            total_row = conn.execute(
                f"SELECT COALESCE(SUM(castka),0) as total FROM vyplaty {where}", params
            ).fetchone()
            total = _first_val(total_row)
        return jsonify({"vyplaty": [dict(r) for r in rows], "celkem": round(total, 2)})
    except Exception as e:
        import traceback
        app.logger.error(f"api_vyplaty GET error: {traceback.format_exc()}")
        return jsonify({"error": str(e)}), 500

@app.route("/api/vyplaty", methods=["POST"])
@vyzaduj_prihlaseni
def api_vyplata_ulozit():
    try:
        data = request.json
        if not data.get("jmeno") or not data.get("datum") or data.get("castka") is None:
            return jsonify({"error": "Chybí povinná pole"}), 400
        with get_db() as conn:
            cur = conn.execute("""
                INSERT INTO vyplaty (jmeno, datum, castka, poznamka, firma_zkratka, obdobi_od, obdobi_do)
                VALUES (?,?,?,?,?,?,?)
            """, (
                data["jmeno"],
                data["datum"],
                float(data["castka"]),
                data.get("poznamka", ""),
                data.get("firma_zkratka", ""),
                data.get("obdobi_od") or None,
                data.get("obdobi_do") or None,
            ))
        return jsonify({"ok": True, "id": cur.lastrowid})
    except Exception as e:
        import traceback
        app.logger.error(f"api_vyplata_ulozit error: {traceback.format_exc()}")
        return jsonify({"error": str(e)}), 500

@app.route("/api/vyplaty/<int:vid>", methods=["DELETE"])
@vyzaduj_prihlaseni
def api_vyplata_delete(vid):
    with get_db() as conn:
        conn.execute("DELETE FROM vyplaty WHERE id=?", (vid,))
    return jsonify({"ok": True})

@app.route("/api/vyplaty/<int:vid>", methods=["PUT"])
@vyzaduj_prihlaseni
def api_vyplata_update(vid):
    data = request.json
    fields = ["jmeno", "datum", "castka", "poznamka", "firma_zkratka", "obdobi_od", "obdobi_do"]
    set_parts = [f"{f}=?" for f in fields if f in data]
    vals = [data[f] for f in fields if f in data]
    if not set_parts:
        return jsonify({"ok": True})
    vals.append(vid)
    with get_db() as conn:
        conn.execute(f"UPDATE vyplaty SET {','.join(set_parts)} WHERE id=?", vals)
    return jsonify({"ok": True})

@app.route("/api/vyplaty/souhrn/<jmeno>")
@vyzaduj_prihlaseni
def api_vyplaty_souhrn(jmeno):
    from datetime import date as _date
    dnes = _date.today()
    mesic_od = f"{dnes.year}-{dnes.month:02d}-01"
    rok_od   = f"{dnes.year}-01-01"
    rok_do   = f"{dnes.year}-12-31"
    with get_db() as conn:
        r_mesic = conn.execute(
            "SELECT COALESCE(SUM(castka),0) as total FROM vyplaty WHERE jmeno=? AND datum>=?",
            (jmeno, mesic_od)
        ).fetchone()
        r_rok = conn.execute(
            "SELECT COALESCE(SUM(castka),0) as total FROM vyplaty WHERE jmeno=? AND datum>=? AND datum<=?",
            (jmeno, rok_od, rok_do)
        ).fetchone()
        odvody = conn.execute(
            "SELECT nazev, castka FROM pausalni_odvody WHERE jmeno=? ORDER BY poradi, nazev",
            (jmeno,)
        ).fetchall()
    celkem_mesic = _first_val(r_mesic)
    celkem_rok   = _first_val(r_rok)
    odvody_list  = [dict(r) for r in odvody]
    odvody_suma  = sum(float(r["castka"]) for r in odvody_list)
    return jsonify({
        "celkem_mesic": round(celkem_mesic, 2),
        "celkem_rok":   round(celkem_rok, 2),
        "odvody":       odvody_list,
        "odvody_suma":  round(odvody_suma, 2),
    })

@app.route("/api/nastaveni/odvody")
@vyzaduj_prihlaseni
def api_nastaveni_odvody_get():
    with get_db() as conn:
        rows = conn.execute("SELECT id, jmeno, nazev, castka FROM pausalni_odvody ORDER BY jmeno, poradi, nazev").fetchall()
    odvody = [dict(r) for r in rows]
    suma = sum(float(r["castka"]) for r in odvody)
    return jsonify({"odvody": odvody, "odvody_suma": round(suma, 2)})

@app.route("/api/nastaveni/odvody", methods=["POST"])
@vyzaduj_prihlaseni
def api_nastaveni_odvody_post():
    data = request.json or {}
    nazev  = str(data.get("nazev","")).strip()
    castka = float(data.get("castka", 0) or 0)
    jmeno  = str(data.get("jmeno","admin")).strip() or "admin"
    platnost_od = data.get("platnost_od", "") or date.today().strftime("%Y-%m") + "-01"
    if not nazev:
        return jsonify({"error": "Chybí název"}), 400
    with get_db() as conn:
        conn.execute(
            "INSERT INTO pausalni_odvody (jmeno, nazev, castka, poradi, platnost_od) VALUES (?,?,?,0,?)",
            (jmeno, nazev, castka, platnost_od)
        )
    return jsonify({"ok": True})

@app.route("/api/nastaveni/odvody/<int:oid>", methods=["DELETE"])
@vyzaduj_prihlaseni
def api_nastaveni_odvody_delete(oid):
    with get_db() as conn:
        conn.execute("DELETE FROM pausalni_odvody WHERE id=?", (oid,))
    return jsonify({"ok": True})


@app.route("/api/pausalni-odvody/<jmeno>", methods=["GET"])
@vyzaduj_prihlaseni
def api_pausalni_get(jmeno):
    with get_db() as conn:
        rows = conn.execute(
            "SELECT id, nazev, castka FROM pausalni_odvody WHERE jmeno=? ORDER BY poradi, nazev",
            (jmeno,)
        ).fetchall()
    return jsonify([dict(r) for r in rows])

@app.route("/api/pausalni-odvody/<jmeno>", methods=["POST"])
@vyzaduj_prihlaseni
def api_pausalni_save(jmeno):
    data = request.json or []
    with get_db() as conn:
        conn.execute("DELETE FROM pausalni_odvody WHERE jmeno=?", (jmeno,))
        for i, item in enumerate(data):
            nazev  = str(item.get("nazev","")).strip()
            castka = float(item.get("castka", 0) or 0)
            if nazev:
                conn.execute(
                    "INSERT INTO pausalni_odvody (jmeno, nazev, castka, poradi) VALUES (?,?,?,?)",
                    (jmeno, nazev, castka, i)
                )
    return jsonify({"ok": True})


@app.route("/api/vyplaty/nahrat-pasku", methods=["POST"])
@vyzaduj_prihlaseni
def api_nahrat_pasku():
    jmeno = request.form.get("jmeno", "").strip()
    soubor = request.files.get("soubor")
    if not jmeno or not soubor:
        return jsonify({"chyba": "Chybí jméno nebo soubor"}), 400
    try:
        import tempfile, os as _os
        from datetime import date as _date
        mesic = _date.today().strftime("%Y-%m")
        fname = f"paska_{jmeno}_{mesic}.pdf"
        with tempfile.NamedTemporaryFile(delete=False, suffix=".pdf") as tmp:
            soubor.save(tmp.name)
            gcs_url = upload_to_gcs(tmp.name, f"vyplaty/{fname}")
            _os.unlink(tmp.name)
        if not gcs_url:
            return jsonify({"chyba": "Nahrávání do GCS selhalo"}), 500
        # Uložit URL ke všem výplatám daného zaměstnance v daném měsíci
        mesic = request.form.get("mesic", _date.today().strftime("%Y-%m"))
        od_m  = f"{mesic}-01"
        import calendar as _cal
        rok_m, mes_m = mesic.split("-")
        posl = _cal.monthrange(int(rok_m), int(mes_m))[1]
        do_m  = f"{mesic}-{posl:02d}"
        with get_db() as conn:
            conn.execute("""
                UPDATE vyplaty SET paska_url=?
                WHERE jmeno=? AND datum>=? AND datum<=?
            """, (gcs_url, jmeno, od_m, do_m))
            # Pokud není žádná výplata v tom měsíci, ulož k poslední
            if conn.execute("SELECT COUNT(*) as c FROM vyplaty WHERE jmeno=? AND datum>=? AND datum<=?",
                           (jmeno, od_m, do_m)).fetchone()["c"] == 0:
                conn.execute("""
                    UPDATE vyplaty SET paska_url=?
                    WHERE id = (SELECT id FROM vyplaty WHERE jmeno=? ORDER BY datum DESC, id DESC LIMIT 1)
                """, (gcs_url, jmeno))
        return jsonify({"url": gcs_url})
    except Exception as e:
        return jsonify({"chyba": str(e)}), 500


@app.route("/api/vyplaty/paska-pdf")
@vyzaduj_prihlaseni
def api_vyplatni_paska_pdf():
    from reportlab.lib.pagesizes import A5
    from reportlab.lib import colors
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import mm
    import io

    jmeno = request.args.get("jmeno", "")
    od    = request.args.get("od", "")
    do_   = request.args.get("do", "")
    if not jmeno:
        return jsonify({"chyba": "Chybí jméno"}), 400

    with get_db() as conn:
        vyplaty = conn.execute(
            "SELECT datum, castka, poznamka, firma_zkratka FROM vyplaty WHERE jmeno=? AND datum>=? AND datum<=? ORDER BY datum",
            (jmeno, od, do_)
        ).fetchall()
        odvody = conn.execute(
            "SELECT nazev, castka FROM pausalni_odvody WHERE jmeno=? ORDER BY poradi, nazev",
            (jmeno,)
        ).fetchall()

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A5,
        leftMargin=15*mm, rightMargin=15*mm, topMargin=15*mm, bottomMargin=15*mm)

    styles = getSampleStyleSheet()
    nadpis = ParagraphStyle("nadpis", parent=styles["Normal"], fontSize=13, fontName="Helvetica-Bold", spaceAfter=4*mm)
    normal = ParagraphStyle("normal", parent=styles["Normal"], fontSize=9, fontName="Helvetica")
    maly   = ParagraphStyle("maly",   parent=styles["Normal"], fontSize=8, fontName="Helvetica", textColor=colors.grey)

    mesic_label = od[:7] if od else ""
    celkem_vyplata = sum(float(v["castka"] if isinstance(v, dict) else v[1]) for v in vyplaty)
    celkem_odvody  = sum(float(o["castka"] if isinstance(o, dict) else o[1]) for o in odvody)

    story = []
    story.append(Paragraph(f"Výplatní páska – {jmeno}", nadpis))
    story.append(Paragraph(f"Období: {mesic_label}", maly))
    story.append(Spacer(1, 4*mm))

    # Výplaty
    story.append(Paragraph("Výplaty / zálohy:", ParagraphStyle("h", parent=normal, fontName="Helvetica-Bold", spaceAfter=2*mm)))
    vdata = [["Datum", "Firma", "Částka", "Poznámka"]]
    for v in vyplaty:
        if isinstance(v, dict):
            vdata.append([v.get("datum",""), v.get("firma_zkratka",""), f"{czInt_py(v.get('castka',0))} Kč", v.get("poznamka","") or ""])
        else:
            vdata.append([v[0], v[3] or "", f"{czInt_py(v[1])} Kč", v[2] or ""])
    vdata.append(["", "CELKEM", f"{czInt_py(celkem_vyplata)} Kč", ""])
    t = Table(vdata, colWidths=[22*mm, 18*mm, 28*mm, 50*mm])
    t.setStyle(TableStyle([
        ("FONTNAME", (0,0), (-1,0), "Helvetica-Bold"),
        ("FONTSIZE", (0,0), (-1,-1), 8),
        ("BACKGROUND", (0,0), (-1,0), colors.HexColor("#f3f4f6")),
        ("BACKGROUND", (0,-1), (-1,-1), colors.HexColor("#fef9c3")),
        ("FONTNAME", (0,-1), (-1,-1), "Helvetica-Bold"),
        ("GRID", (0,0), (-1,-1), 0.3, colors.grey),
        ("ALIGN", (2,0), (2,-1), "RIGHT"),
    ]))
    story.append(t)
    story.append(Spacer(1, 5*mm))

    # Odvody
    if odvody:
        story.append(Paragraph("Paušální odvody:", ParagraphStyle("h", parent=normal, fontName="Helvetica-Bold", spaceAfter=2*mm)))
        odata = [["Položka", "Měsíčně"]]
        for o in odvody:
            if isinstance(o, dict):
                odata.append([o.get("nazev",""), f"{czInt_py(o.get('castka',0))} Kč"])
            else:
                odata.append([o[0], f"{czInt_py(o[1])} Kč"])
        odata.append(["CELKEM odvody", f"{czInt_py(celkem_odvody)} Kč"])
        ot = Table(odata, colWidths=[80*mm, 28*mm])
        ot.setStyle(TableStyle([
            ("FONTNAME", (0,0), (-1,0), "Helvetica-Bold"),
            ("FONTSIZE", (0,0), (-1,-1), 8),
            ("BACKGROUND", (0,0), (-1,0), colors.HexColor("#f3f4f6")),
            ("BACKGROUND", (0,-1), (-1,-1), colors.HexColor("#fee2e2")),
            ("FONTNAME", (0,-1), (-1,-1), "Helvetica-Bold"),
            ("GRID", (0,0), (-1,-1), 0.3, colors.grey),
            ("ALIGN", (1,0), (1,-1), "RIGHT"),
        ]))
        story.append(ot)
        story.append(Spacer(1, 4*mm))
        story.append(Paragraph(f"Celkem náklady: {czInt_py(celkem_vyplata + celkem_odvody)} Kč",
            ParagraphStyle("total", parent=normal, fontName="Helvetica-Bold", fontSize=10)))

    doc.build(story)
    buf.seek(0)
    fname = f"vyplatni_paska_{jmeno}_{mesic_label}.pdf"
    return send_file(buf, mimetype="application/pdf",
        as_attachment=False, download_name=fname)


def czInt_py(v):
    try:
        return f"{int(float(v or 0)):,}".replace(",", " ")
    except Exception:
        return "0"


# ── API: VÝDAJE ───────────────────────────────────────────────────────────────
@app.route("/api/vydaje")
@vyzaduj_prihlaseni
def api_vydaje_list():
    firma = request.args.get("firma", "")
    od    = request.args.get("od", "")
    do_   = request.args.get("do", "")
    stav  = request.args.get("stav", "")
    typ   = request.args.get("typ", "provozni")
    clauses, params = [], []
    if firma: clauses.append("firma_zkratka=?"); params.append(firma)
    if od:    clauses.append("datum>=?"); params.append(od)
    if do_:   clauses.append("datum<=?"); params.append(do_)
    if stav:  clauses.append("stav=?"); params.append(stav)
    clauses.append("COALESCE(typ,'provozni')=?"); params.append(typ)
    where = ("WHERE " + " AND ".join(clauses)) if clauses else ""
    with get_db() as conn:
        rows = conn.execute(
            f"SELECT * FROM vydaje {where} ORDER BY datum DESC, id DESC", params
        ).fetchall()
        total = conn.execute(
            f"SELECT COALESCE(SUM(castka),0) as t FROM vydaje {where}", params
        ).fetchone()
        result = []
        for r in rows:
            d = dict(r)
            polozky = conn.execute(
                "SELECT * FROM vydaje_polozky WHERE vydaj_id=? ORDER BY nazev", (d["id"],)
            ).fetchall()
            d["polozky"] = [dict(p) for p in polozky]
            result.append(d)
    return jsonify({"vydaje": result, "celkem": round(_first_val(total), 2)})

@app.route("/api/vydaje", methods=["POST"])
@vyzaduj_prihlaseni
def api_vydaje_ulozit():
    data = request.json or {}
    if not data.get("firma_zkratka"):
        return jsonify({"error": "Chybí firma"}), 400
    polozky = data.get("polozky", [])
    try:
        conn = psycopg2.connect(os.environ["DATABASE_URL"])
        cur = conn.cursor()
        cur.execute("""
            INSERT INTO vydaje
                (firma_zkratka, dodavatel, datum, datum_splatnosti, castka,
                 zpusob_uhrady, stav, popis, poznamka, soubor_cesta,
                 soubor_url, zdroj, typ, stitky)
            VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
            RETURNING id
        """, (
            data.get("firma_zkratka", ""),
            data.get("dodavatel", ""),
            data.get("datum", ""),
            data.get("datum_splatnosti", ""),
            float(data.get("castka", 0)),
            data.get("zpusob_uhrady", "hotovost"),
            data.get("stav", "nezaplaceno"),
            data.get("popis", ""),
            data.get("poznamka", ""),
            data.get("soubor_cesta", ""),
            data.get("soubor_url", ""),
            data.get("zdroj", "rucni"),
            data.get("typ", "provozni"),
            data.get("stitky", ""),
        ))
        vid = cur.fetchone()[0]
        for p in polozky:
            nazev = (p.get("nazev") or "").strip()
            if not nazev:
                continue
            cur.execute(
                "INSERT INTO vydaje_polozky (vydaj_id, nazev, castka) VALUES (%s,%s,%s)",
                (vid, nazev, float(p.get("castka", 0)))
            )
        conn.commit()
        cur.close()
        conn.close()
        return jsonify({"ok": True, "id": vid, "duplicita": False, "duplicita_id": None})
    except Exception as e:
        try: conn.rollback(); conn.close()
        except: pass
        return jsonify({"error": str(e)}), 500

@app.route("/api/vydaje/<int:vid>", methods=["PUT"])
@vyzaduj_prihlaseni
def api_vydaje_edit(vid):
    data = request.json
    polozky = data.pop("polozky", None)
    with get_db() as conn:
        conn.execute("""
            UPDATE vydaje SET dodavatel=?, datum=?, datum_splatnosti=?, castka=?,
                zpusob_uhrady=?, stav=?, popis=?, poznamka=?, firma_zkratka=?,
                datum_uhrady=?, banka_uhrady=?, stitky=?
            WHERE id=?
        """, (
            data.get("dodavatel", ""),
            data.get("datum", ""),
            data.get("datum_splatnosti", ""),
            float(data.get("castka", 0)),
            data.get("zpusob_uhrady", "hotovost"),
            data.get("stav", "nezaplaceno"),
            data.get("popis", ""),
            data.get("poznamka", ""),
            data.get("firma_zkratka", ""),
            data.get("datum_uhrady", ""),
            data.get("banka_uhrady", ""),
            data.get("stitky", ""),
            vid,
        ))
        if polozky is not None:
            conn.execute("DELETE FROM vydaje_polozky WHERE vydaj_id=?", (vid,))
            for p in polozky:
                nazev = (p.get("nazev") or "").strip()
                if not nazev: continue
                conn.execute("INSERT INTO vydaje_polozky (vydaj_id, nazev, castka) VALUES (?,?,?)",
                    (vid, nazev, float(p.get("castka", 0))))
    return jsonify({"ok": True})

@app.route("/api/vydaje/<int:vid>/stav", methods=["POST"])
@vyzaduj_prihlaseni
def api_vydaje_stav(vid):
    d = request.json or {}
    stav = d.get("stav", "zaplaceno")
    datum_uhrady = d.get("datum_uhrady", "")
    banka_uhrady = d.get("banka_uhrady", "")
    with get_db() as conn:
        conn.execute(
            "UPDATE vydaje SET stav=?, datum_uhrady=?, banka_uhrady=? WHERE id=?",
            (stav, datum_uhrady, banka_uhrady, vid)
        )
    return jsonify({"ok": True})

@app.route("/api/vydaje/<int:vid>", methods=["DELETE"])
@vyzaduj_prihlaseni
def api_vydaje_delete(vid):
    with get_db() as conn:
        conn.execute("DELETE FROM vydaje_polozky WHERE vydaj_id=?", (vid,))
        conn.execute("DELETE FROM vydaje WHERE id=?", (vid,))
    return jsonify({"ok": True})


@app.route("/api/vydaje/oznac-duplicity", methods=["POST"])
@vyzaduj_prihlaseni
def api_vydaje_oznac_duplicity():
    """Zpětně označí duplicitní výdaje — alespoň 2 ze 3: datum, částka, dodavatel."""
    try:
        oznaceno = 0
        with get_db() as conn:
            vydaje = conn.execute(
                "SELECT id, datum, dodavatel, castka FROM vydaje ORDER BY id ASC"
            ).fetchall()
            for v in vydaje:
                vid    = v["id"] if isinstance(v, dict) else v[0]
                datum  = (v["datum"] if isinstance(v, dict) else v[1]) or ""
                dodav  = ((v["dodavatel"] if isinstance(v, dict) else v[2]) or "").strip().lower()
                castka = float(v["castka"] if isinstance(v, dict) else v[3])
                # Přeskočit pokud již označen
                dup = conn.execute("SELECT duplicita_id FROM vydaje WHERE id=?", (vid,)).fetchone()
                if dup and (dup["duplicita_id"] if isinstance(dup, dict) else dup[0]):
                    continue
                # Hledej starší záznam s alespoň 2 shodnými kritérii
                kandidati = conn.execute(
                    "SELECT id, datum, dodavatel, castka FROM vydaje WHERE id < ?", (vid,)
                ).fetchall()
                for k in kandidati:
                    kid    = k["id"] if isinstance(k, dict) else k[0]
                    kdatum = (k["datum"] if isinstance(k, dict) else k[1]) or ""
                    kdodav = ((k["dodavatel"] if isinstance(k, dict) else k[2]) or "").strip().lower()
                    kcastka = float(k["castka"] if isinstance(k, dict) else k[3])
                    skore = sum([
                        datum and kdatum and datum == kdatum,
                        abs(castka - kcastka) < 1.0,
                        bool(dodav and kdodav and dodav == kdodav),
                    ])
                    if skore >= 2:
                        conn.execute(
                            "UPDATE vydaje SET duplicita_id=? WHERE id=? AND duplicita_id IS NULL",
                            (kid, vid)
                        )
                        oznaceno += 1
                        break
        return jsonify({"ok": True, "oznaceno": oznaceno})
    except Exception as e:
        return jsonify({"ok": False, "chyba": str(e)}), 500

@app.route("/api/vydaje/nahrat", methods=["POST"])
@vyzaduj_prihlaseni
def api_vydaje_nahrat():
    if "soubor" not in request.files:
        return jsonify({"error": "Žádný soubor"}), 400
    f = request.files["soubor"]
    firma = request.form.get("firma_zkratka", "")
    fname = secure_filename(f.filename or "vydaj")
    fpath = os.path.join(UPLOAD_DIR, fname)
    f.save(fpath)
    gcs_url = upload_to_gcs(fpath, f"vydaje/{fname}")
    return _vydaje_ocr(fpath, fname, gcs_url or "", firma)

@app.route("/api/vydaje/nahrat-path", methods=["POST"])
@vyzaduj_prihlaseni
def api_vydaje_nahrat_path():
    d = request.json or {}
    fpath = d.get("path", "")
    soubor_url = d.get("soubor_url", "")
    filename = d.get("filename", "vydaj.pdf")
    firma = d.get("firma_zkratka", "")
    if not fpath or not os.path.exists(fpath):
        return jsonify({"error": "Soubor nenalezen"}), 400
    return _vydaje_ocr(fpath, filename, soubor_url, firma)

def _vydaje_ocr(fpath, fname, gcs_url, firma):
    try:
        with open(fpath, "rb") as fh:
            raw = fh.read()
        b64 = base64.b64encode(raw).decode()
        ext = fname.rsplit(".", 1)[-1].lower()
        mt = "application/pdf" if ext == "pdf" else f"image/{ext if ext in ['jpeg','jpg','png','gif','webp'] else 'jpeg'}"
        if mt == "image/jpg": mt = "image/jpeg"
        msg_content = [
            {"type": "image" if not mt.startswith("application") else "document",
             "source": {"type": "base64", "media_type": mt, "data": b64}},
            {"type": "text", "text": """Analyzuj tento doklad/účtenku a extrahuj:
- dodavatel: název obchodu/firmy
- datum: datum nákupu ve formátu YYYY-MM-DD
- castka: celková částka v Kč (číslo bez měny)
- poznamka: krátký popis co bylo nakoupeno (max 80 znaků)
Odpověz POUZE jako JSON: {"dodavatel":"...","datum":"...","castka":0,"poznamka":"..."}"""}
        ]
        resp = anthropic.Anthropic(api_key=os.environ.get("ANTHROPIC_API_KEY","")).messages.create(
            model="claude-sonnet-4-20250514", max_tokens=300,
            messages=[{"role": "user", "content": msg_content}]
        )
        import json as _json
        text = resp.content[0].text.strip()
        text = text.replace("```json","").replace("```","").strip()
        parsed = _json.loads(text)
    except Exception as e:
        parsed = {}
    return jsonify({
        "dodavatel":      parsed.get("dodavatel", ""),
        "datum":          parsed.get("datum", ""),
        "castka":         parsed.get("castka", 0),
        "poznamka":       parsed.get("poznamka", ""),
        "soubor_cesta":   fname,
        "soubor_gcs_url": gcs_url,
        "firma_zkratka":  firma,
    })

# ── API: VYSTAVENÉ FAKTURY ────────────────────────────────────────────────────

@app.route("/api/vystavene-faktury")
@vyzaduj_prihlaseni
def api_vystavene_list():
    if session.get("role") == "verunka":
        return jsonify({"error": "Přístup zamítnut"}), 403
    firma = request.args.get("firma", "")
    od    = request.args.get("od", "")
    do_   = request.args.get("do", "")
    clauses, params = [], []
    if firma: clauses.append("firma_zkratka=?"); params.append(firma)
    if od:    clauses.append("datum>=?"); params.append(od)
    if do_:   clauses.append("datum<=?"); params.append(do_)
    where = ("WHERE " + " AND ".join(clauses)) if clauses else ""
    with get_db() as conn:
        rows = conn.execute(
            f"SELECT * FROM vystavene_faktury {where} ORDER BY datum DESC, id DESC", params
        ).fetchall()
    return jsonify([dict(r) for r in rows])

@app.route("/api/vystavene-faktury/zkontroluj", methods=["POST"])
@vyzaduj_prihlaseni
def api_vystavene_zkontroluj():
    d = request.json or {}
    duplicita = None
    if d.get("cislo_faktury") and d.get("datum"):
        with get_db() as conn:
            row = conn.execute(
                """SELECT id, firma_zkratka, datum, castka FROM vystavene_faktury
                   WHERE cislo_faktury=? AND datum=? AND ABS(castka-?)<0.01""",
                (d.get("cislo_faktury"), d.get("datum"), float(d.get("castka", 0)))
            ).fetchone()
            if row:
                duplicita = {"id": row["id"], "firma": row["firma_zkratka"],
                             "datum": row["datum"], "castka": row["castka"]}
    return jsonify({"duplicita": duplicita})

@app.route("/api/vystavene-faktury", methods=["POST"])
@vyzaduj_prihlaseni
def api_vystavene_ulozit():
    if session.get("role") != "admin":
        return jsonify({"error": "Přístup zamítnut"}), 403
    d = request.json or {}
    duplicita = None
    try:
        conn = psycopg2.connect(os.environ["DATABASE_URL"])
        cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
        if d.get("cislo_faktury") and d.get("datum"):
            cur.execute(
                """SELECT id, firma_zkratka, datum, castka FROM vystavene_faktury
                   WHERE cislo_faktury=%s AND datum=%s AND ABS(castka-%s)< 0.01""",
                (d.get("cislo_faktury"), d.get("datum"), float(d.get("castka",0)))
            )
            row = cur.fetchone()
            if row:
                duplicita = {"id": row["id"], "firma": row["firma_zkratka"],
                             "datum": row["datum"], "castka": row["castka"]}
        cur.execute(
            """INSERT INTO vystavene_faktury
               (firma_zkratka, cislo_faktury, datum, datum_splatnosti, odberatel, popis, castka, stav, soubor_url, duplicita_id)
               VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s) RETURNING id""",
            (d.get("firma_zkratka",""), d.get("cislo_faktury",""),
             d.get("datum",""), d.get("datum_splatnosti",""),
             d.get("odberatel",""), d.get("popis",""),
             float(d.get("castka",0)),
             "duplikat" if duplicita else d.get("stav","nezaplaceno"),
             d.get("soubor_url",""),
             duplicita["id"] if duplicita else None)
        )
        conn.commit()
        cur.close()
        conn.close()
    except Exception as e:
        return jsonify({"error": str(e)}), 500
    return jsonify({"ok": True, "duplicita": duplicita})

@app.route("/api/vystavene-faktury/<int:fid>", methods=["PUT"])
@vyzaduj_prihlaseni
def api_vystavene_edit(fid):
    if session.get("role") != "admin":
        return jsonify({"error": "Přístup zamítnut"}), 403
    d = request.json or {}
    with get_db() as conn:
        conn.execute(
            """UPDATE vystavene_faktury SET firma_zkratka=?, cislo_faktury=?, datum=?,
               datum_splatnosti=?, odberatel=?, popis=?, castka=?, stav=?, soubor_url=? WHERE id=?""",
            (d.get("firma_zkratka",""), d.get("cislo_faktury",""),
             d.get("datum",""), d.get("datum_splatnosti",""),
             d.get("odberatel",""), d.get("popis",""),
             float(d.get("castka",0)), d.get("stav","nezaplaceno"), d.get("soubor_url",""), fid)
        )
    return jsonify({"ok": True})

@app.route("/api/vystavene-faktury/<int:fid>", methods=["DELETE"])
@vyzaduj_prihlaseni
def api_vystavene_delete(fid):
    if session.get("role") != "admin":
        return jsonify({"error": "Přístup zamítnut"}), 403
    with get_db() as conn:
        conn.execute("DELETE FROM vystavene_faktury WHERE id=?", (fid,))
    return jsonify({"ok": True})

@app.route("/api/vystavene-faktury/<int:fid>/stav", methods=["POST"])
@vyzaduj_prihlaseni
def api_vystavene_stav(fid):
    if session.get("role") != "admin":
        return jsonify({"error": "Přístup zamítnut"}), 403
    stav = request.json.get("stav", "zaplaceno")
    with get_db() as conn:
        conn.execute("UPDATE vystavene_faktury SET stav=? WHERE id=?", (stav, fid))
    return jsonify({"ok": True})

@app.route("/api/vystavene-faktury/nahrat-path", methods=["POST"])
@vyzaduj_prihlaseni
def api_vystavene_nahrat_path():
    if session.get("role") != "admin":
        return jsonify({"error": "Přístup zamítnut"}), 403
    d = request.json or {}
    fpath = d.get("path", "")
    soubor_url = d.get("soubor_url", "")
    if not fpath or not os.path.exists(fpath):
        return jsonify({"error": "Soubor nenalezen"}), 400
    return _vystavene_ocr(fpath, soubor_url)

def _vystavene_ocr(fpath, soubor_url=""):
    api_key = os.environ.get("ANTHROPIC_API_KEY", "")
    if not api_key:
        return jsonify({"error": "ANTHROPIC_API_KEY není nastaven", "soubor_url": soubor_url}), 200
    try:
        ext = fpath.rsplit(".", 1)[-1].lower()
        with open(fpath, "rb") as fh:
            raw = fh.read()
        b64 = base64.standard_b64encode(raw).decode("utf-8")
        if ext == "pdf":
            content_block = {"type": "document", "source": {"type": "base64", "media_type": "application/pdf", "data": b64}}
        else:
            media_map = {"jpg": "image/jpeg", "jpeg": "image/jpeg", "png": "image/png", "webp": "image/webp"}
            content_block = {"type": "image", "source": {"type": "base64", "media_type": media_map.get(ext, "image/jpeg"), "data": b64}}
        client = anthropic.Anthropic(api_key=api_key)
        msg = client.messages.create(
            model="claude-sonnet-4-20250514", max_tokens=500,
            messages=[{"role": "user", "content": [
                content_block,
                {"type": "text", "text": """Analyzuj tuto vystavenou fakturu a extrahuj tyto hodnoty.
Odpověz POUZE platným JSON objektem, žádný jiný text ani backticky.
{
  "cislo_faktury": "číslo faktury (text)",
  "datum": "datum vystavení YYYY-MM-DD nebo null",
  "datum_splatnosti": "datum splatnosti YYYY-MM-DD nebo null",
  "castka": číslo (celková částka v Kč bez symbolu),
  "odberatel": "název odběratele",
  "popis": "stručný popis předmětu plnění max 100 znaků",
  "firma_zkratka": "zkratka vystavitele: pokud vidíš Food Plus → FP, MR plus nebo MRplus → MR, Clever food factory → CFF, jinak prázdný string"
}"""}
            ]}]
        )
        text = msg.content[0].text.strip()
        text = re.sub(r"^```json\s*", "", text)
        text = re.sub(r"```$", "", text).strip()
        parsed = json.loads(text)
    except Exception as e:
        app.logger.warning(f"OCR vystavene failed: {e}")
        return jsonify({"error": str(e), "soubor_url": soubor_url}), 200
    return jsonify({
        "cislo_faktury":    parsed.get("cislo_faktury") or "",
        "datum":            parsed.get("datum") or "",
        "datum_splatnosti": parsed.get("datum_splatnosti") or "",
        "castka":           float(parsed.get("castka") or 0),
        "odberatel":        parsed.get("odberatel") or "",
        "popis":            parsed.get("popis") or "",
        "firma_zkratka":    parsed.get("firma_zkratka") or "",
        "soubor_url":       soubor_url,
    })

@app.route("/api/vystavene-faktury/nahrat", methods=["POST"])
@vyzaduj_prihlaseni
def api_vystavene_nahrat():
    if session.get("role") != "admin":
        return jsonify({"error": "Přístup zamítnut"}), 403
    if "soubor" not in request.files:
        return jsonify({"error": "Žádný soubor"}), 400
    f = request.files["soubor"]
    fname = secure_filename(f.filename or "faktura.pdf")
    ts = datetime.now().strftime("%Y%m%d_%H%M%S_")
    fname = ts + fname
    fpath = os.path.join(UPLOAD_DIR, fname)
    f.save(fpath)
    gcs_url = upload_to_gcs(fpath, f"vystavene/{fname}")
    return _vystavene_ocr(fpath, gcs_url or "")


# ── API: BANKOVNÍ VÝPISY ──────────────────────────────────────────────────────
def parse_csv_airbank(content_bytes):
    import csv, io
    # Detekce kódování
    for enc in ["utf-8-sig", "cp1250", "utf-8"]:
        try:
            text = content_bytes.decode(enc)
            break
        except Exception:
            continue

    # Detekce oddělovače — tabulátor (osobní) nebo středník (firemní)
    first_line = text.split("\n")[0]
    delimiter = "\t" if "\t" in first_line else ";"

    reader = csv.DictReader(io.StringIO(text), delimiter=delimiter)
    pohyby = []
    for row in reader:
        datum_raw = row.get("Datum provedení", "").strip().strip('"')
        castka_raw = row.get("Částka v měně účtu", "").strip().strip('"').replace("\xa0", "").replace(" ", "").replace(",", ".")
        id_transakce = row.get("Referenční číslo", "").strip().strip('"')
        if not datum_raw or not castka_raw:
            continue
        try:
            # Podpora DD.MM.YYYY i DD/MM/YYYY
            if "." in datum_raw:
                d, m, y = datum_raw.split(".")
            else:
                d, m, y = datum_raw.split("/")
            datum = f"{y.strip()[:4]}-{m.zfill(2)}-{d.zfill(2)}"
            castka = float(castka_raw)
        except Exception:
            continue
        zprava = row.get("Obchodní místo", "").strip().strip('"') or row.get("Zpráva pro příjemce", "").strip().strip('"') or row.get("Poznámka pro mne", "").strip().strip('"')
        # Parsuj VS ze zprávy — hledej číslo 4-10 číslic
        import re as _re
        vs_match = _re.search(r'\b(\d{4,10})\b', zprava)
        var_sym = vs_match.group(1) if vs_match else ''
        pohyby.append({
            "banka":           "AirBank",
            "datum":           datum,
            "castka":          castka,
            "protiucet":       row.get("Číslo účtu protistrany", "").strip().strip('"'),
            "nazev_protiucet": row.get("Název protistrany", "").strip().strip('"'),
            "typ_transakce":   row.get("Typ úhrady", "").strip().strip('"'),
            "zprava":          zprava,
            "var_sym":         var_sym,
            "id_transakce":    f"AIR_{id_transakce}" if id_transakce else None,
        })
    return pohyby

def parse_csv_kb(content_bytes):
    """Parser pro Komerční banku KB+ CSV formát."""
    import csv, io
    for enc in ["utf-8-sig", "cp1250", "utf-8"]:
        try:
            text = content_bytes.decode(enc)
            break
        except Exception:
            continue

    # KB má metadata nahoře - najdeme řádek s hlavičkou dat
    lines = text.splitlines()
    header_idx = None
    for i, line in enumerate(lines):
        if "Datum zauctovani" in line or "Datum zaúčtování" in line:
            header_idx = i
            break

    if header_idx is None:
        return []

    # Sestavíme CSV jen od hlavičky dál
    csv_text = "\n".join(lines[header_idx:])
    delimiter = "\t" if "\t" in lines[header_idx] else ";"
    reader = csv.DictReader(io.StringIO(csv_text), delimiter=delimiter)

    pohyby = []
    for row in reader:
        datum_raw = (row.get("Datum zauctovani") or row.get("Datum zaúčtování") or "").strip()
        castka_raw = (row.get("Castka") or row.get("Částka") or "").strip().replace("\xa0", "").replace(" ", "").replace(",", ".")
        id_transakce = (row.get("Identifikace transakce") or "").strip()
        if not datum_raw or not castka_raw:
            continue
        try:
            d, m, y = datum_raw.split(".")
            datum = f"{y.strip()[:4]}-{m.zfill(2)}-{d.zfill(2)}"
            castka = float(castka_raw)
        except Exception:
            continue
        zprava_kb = (row.get("Zprava pro prijemce") or row.get("Zpráva pro příjemce") or row.get("Popis pro me") or "").strip()
        import re as _re
        vs_match_kb = _re.search(r'\b(\d{4,10})\b', zprava_kb)
        pohyby.append({
            "banka":           "KB",
            "datum":           datum,
            "castka":          castka,
            "protiucet":       (row.get("Protistrana") or "").strip(),
            "nazev_protiucet": (row.get("Nazev protiuctu") or row.get("Název protiúčtu") or "").strip(),
            "typ_transakce":   (row.get("Typ transakce") or "").strip(),
            "zprava":          zprava_kb,
            "var_sym":         vs_match_kb.group(1) if vs_match_kb else '',
            "id_transakce":    f"KB_{id_transakce}" if id_transakce else None,
        })
    return pohyby


def parse_csv_rb(content_bytes):
    import csv, io, re as _re
    for enc in ["utf-8-sig", "cp1250", "utf-8"]:
        try:
            text = content_bytes.decode(enc)
            break
        except Exception:
            continue

    first_line = text.split("\n")[0]
    delimiter = "\t" if "\t" in first_line else ";"

    reader = csv.DictReader(io.StringIO(text), delimiter=delimiter)
    pohyby = []
    for row in reader:
        # Datum — zkus oba sloupce
        datum_raw = (row.get("Datum provedení") or row.get("Datum zaúčtování") or "").strip().strip('"')
        # Částka — Raifka má "Zaúčtovaná částka" nebo "Zaúčtovaná částka"
        castka_raw = (row.get("Zaúčtovaná částka") or row.get("Zaúčtovaná částka") or row.get("Castka") or row.get("Částka") or "").strip().strip('"').replace("\xa0", "").replace(" ", "").replace(",", ".")
        # ID transakce
        id_transakce = (row.get("Id transakce") or row.get("ID transakce") or row.get("Identifikace transakce") or "").strip().strip('"')
        # VS — Raifka má samostatný sloupec VS!
        var_sym = (row.get("VS") or "").strip().strip('"')
        # Zpráva
        zprava_rb = (row.get("Zpráva") or row.get("Poznámka") or row.get("Vlastní poznámka") or "").strip().strip('"')
        # Pokud VS prázdné, zkus ho najít ve zprávě
        if not var_sym:
            vs_match_rb = _re.search(r'\b(\d{4,10})\b', zprava_rb)
            var_sym = vs_match_rb.group(1) if vs_match_rb else ''

        if not datum_raw or not castka_raw:
            continue
        try:
            datum_raw = datum_raw.split(" ")[0]  # odstraň čas pokud je přítomen
            if "." in datum_raw:
                d, m, y = datum_raw.split(".")
            else:
                d, m, y = datum_raw.split("/")
            datum = f"{y.strip()[:4]}-{m.zfill(2)}-{d.zfill(2)}"
            castka = float(castka_raw)
        except Exception:
            continue
        pohyby.append({
            "banka":           "RB",
            "datum":           datum,
            "castka":          castka,
            "protiucet":       (row.get("Číslo protiúčtu") or row.get("Číslo protiúčtu") or "").strip().strip('"'),
            "nazev_protiucet": (row.get("Název protiúčtu") or row.get("Název obchodníka") or "").strip().strip('"'),
            "typ_transakce":   (row.get("Typ transakce") or row.get("Kategorie transakce") or "").strip().strip('"'),
            "zprava":          zprava_rb,
            "var_sym":         var_sym,
            "id_transakce":    f"RB_{id_transakce}" if id_transakce else None,
        })
    return pohyby

@app.route("/api/banky/import", methods=["POST"])
@vyzaduj_prihlaseni
def api_banky_import():
    if "soubor" not in request.files:
        return jsonify({"error": "Žádný soubor"}), 400
    f = request.files["soubor"]
    firma = request.form.get("firma_zkratka", "")
    banka_hint = request.form.get("banka_hint", "")
    content = f.read()
    fname = (f.filename or "").lower()

    try:
        if banka_hint == "KB":
            pohyby = parse_csv_kb(content)
            banka = "KB"
        elif banka_hint == "AirBank" or "airbank" in fname or "air_bank" in fname:
            pohyby = parse_csv_airbank(content)
            banka = "AirBank"
        elif banka_hint == "RB" or "pohyby_" in fname:
            pohyby = parse_csv_rb(content)
            banka = "RB"
        else:
            if content[:3] == b'\xef\xbb\xbf':
                pohyby = parse_csv_rb(content)
                banka = "RB"
            else:
                pohyby = parse_csv_airbank(content)
                banka = "AirBank"
    except Exception as e:
        return jsonify({"error": f"Chyba parsování: {str(e)}"}), 400

    # Debug: pokud 0 řádků, vrátíme info o souboru
    if not pohyby:
        try:
            for enc in ["utf-8-sig", "cp1250", "utf-8"]:
                try:
                    preview = content.decode(enc)
                    break
                except Exception:
                    preview = ""
            lines = preview.splitlines()
            return jsonify({
                "ok": True, "banka": banka, "naimportovano": 0, "duplicity": 0,
                "debug": {
                    "radku_celkem": len(lines),
                    "prvni_radek": lines[0][:200] if lines else "",
                    "druhy_radek": lines[1][:200] if len(lines) > 1 else "",
                    "banka_hint": banka_hint,
                    "fname": fname,
                }
            })
        except Exception:
            pass

    naimportovano = 0
    duplicity = 0
    prvni_chyba = ""
    try:
        import psycopg2 as _pg2
        db_url = os.environ.get("DATABASE_URL", "")
        pg_conn = _pg2.connect(db_url)
        pg_cur = pg_conn.cursor()

        # Zjisti jaké sloupce existují
        pg_cur.execute("SELECT column_name FROM information_schema.columns WHERE table_name='bankovni_pohyby'")
        existujici_sloupce = [r[0] for r in pg_cur.fetchall()]
        ma_var_sym = 'var_sym' in existujici_sloupce

        # Přidej chybějící sloupce
        if not ma_var_sym:
            try:
                pg_cur.execute("ALTER TABLE bankovni_pohyby ADD COLUMN var_sym TEXT DEFAULT ''")
                pg_conn.commit()
                ma_var_sym = True
            except Exception: pg_conn.rollback()
        if 'sparovano' not in existujici_sloupce:
            try:
                pg_cur.execute("ALTER TABLE bankovni_pohyby ADD COLUMN sparovano INTEGER DEFAULT 0")
                pg_conn.commit()
            except Exception: pg_conn.rollback()
        if 'sparovano_typ' not in existujici_sloupce:
            try:
                pg_cur.execute("ALTER TABLE bankovni_pohyby ADD COLUMN sparovano_typ TEXT DEFAULT ''")
                pg_conn.commit()
            except Exception: pg_conn.rollback()
        if 'sparovano_id' not in existujici_sloupce:
            try:
                pg_cur.execute("ALTER TABLE bankovni_pohyby ADD COLUMN sparovano_id INTEGER DEFAULT NULL")
                pg_conn.commit()
            except Exception: pg_conn.rollback()

        # Načti existující id_transakce pro deduplikaci
        pg_cur.execute("SELECT id_transakce FROM bankovni_pohyby WHERE id_transakce IS NOT NULL")
        existujici_ids = set(r[0] for r in pg_cur.fetchall())

        for p in pohyby:
            id_tr = p["id_transakce"]
            if id_tr and id_tr in existujici_ids:
                duplicity += 1
                continue
            try:
                pg_cur.execute("""
                    INSERT INTO bankovni_pohyby
                        (banka, datum, castka, protiucet, nazev_protiucet, typ_transakce, zprava, var_sym, id_transakce, firma_zkratka)
                    VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
                """, (
                    p["banka"], p["datum"], p["castka"],
                    p["protiucet"], p["nazev_protiucet"],
                    p["typ_transakce"], p["zprava"],
                    p.get("var_sym", ""),
                    id_tr, firma
                ))
                if id_tr:
                    existujici_ids.add(id_tr)
                naimportovano += 1
            except Exception as row_err:
                if not prvni_chyba:
                    prvni_chyba = str(row_err)
                try: pg_conn.rollback()
                except: pass
                duplicity += 1
        pg_conn.commit()
        pg_conn.close()
    except Exception as e:
        return jsonify({"error": f"Chyba DB: {str(e)}"}), 500
    return jsonify({"ok": True, "banka": banka, "naimportovano": naimportovano, "duplicity": duplicity, "prvni_chyba": prvni_chyba})

@app.route("/api/banky/pohyby")
@vyzaduj_prihlaseni
def api_banky_pohyby():
    banka  = request.args.get("banka", "")
    firma  = request.args.get("firma", "")
    od     = request.args.get("od", "")
    do_    = request.args.get("do", "")
    typ    = request.args.get("typ", "")
    clauses, params = [], []
    if banka: clauses.append("banka=%s"); params.append(banka)
    if firma:
        if firma == "_soukrome":
            clauses.append("(firma_zkratka=%s OR firma_zkratka='' OR firma_zkratka IS NULL)")
            params.append(firma)
        else:
            clauses.append("firma_zkratka=%s"); params.append(firma)
    if od:    clauses.append("datum>=%s"); params.append(od)
    if do_:   clauses.append("datum<=%s"); params.append(do_)
    if typ == "prichozi":  clauses.append("castka>0")
    if typ == "odchozi":   clauses.append("castka<0")
    where = ("WHERE " + " AND ".join(clauses)) if clauses else ""
    try:
        import psycopg2 as _pg2
        db_url = os.environ.get("DATABASE_URL", "")
        pg_conn = _pg2.connect(db_url)
        pg_cur = pg_conn.cursor()
        pg_cur.execute(f"SELECT * FROM bankovni_pohyby {where} ORDER BY datum DESC, id DESC", params)
        cols = [d[0] for d in pg_cur.description]
        rows = [dict(zip(cols, r)) for r in pg_cur.fetchall()]
        pg_cur.execute(f"SELECT COALESCE(SUM(castka),0) FROM bankovni_pohyby {where}", params)
        total = pg_cur.fetchone()[0] or 0
        pg_conn.close()
        return jsonify({"pohyby": rows, "celkem": round(float(total), 2)})
    except Exception as e:
        return jsonify({"pohyby": [], "celkem": 0, "error": str(e)})

@app.route("/api/banky/export")
@vyzaduj_prihlaseni
def api_banky_export():
    banka  = request.args.get("banka", "")
    mesic  = request.args.get("mesic", "")
    fmt    = request.args.get("format", "csv")
    if not banka or not mesic:
        return jsonify({"error": "Chybí parametry"}), 400
    od = mesic + "-01"
    import calendar
    rok, mes = int(mesic[:4]), int(mesic[5:7])
    posledni = calendar.monthrange(rok, mes)[1]
    do_ = f"{mesic}-{posledni:02d}"
    with get_db() as conn:
        rows = conn.execute(
            "SELECT * FROM bankovni_pohyby WHERE banka=? AND datum>=? AND datum<=? ORDER BY datum",
            (banka, od, do_)
        ).fetchall()
    if fmt == "csv":
        import csv, io
        out = io.StringIO()
        w = csv.writer(out)
        w.writerow(["Datum","Protistrana","Číslo účtu","Typ transakce","Zpráva","Částka"])
        for r in rows:
            w.writerow([r["datum"], r["nazev_protiucet"], r["protiucet"], r["typ_transakce"], r["zprava"], r["castka"]])
        from flask import make_response
        resp = make_response(out.getvalue().encode("utf-8-sig"))
        resp.headers["Content-Type"] = "text/csv; charset=utf-8"
        resp.headers["Content-Disposition"] = f'attachment; filename="{banka}_{mesic}.csv"'
        return resp
    else:
        from reportlab.lib.pagesizes import A4
        from reportlab.lib import colors
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.styles import getSampleStyleSheet
        from reportlab.lib.units import mm
        import io as _io
        from flask import make_response

        nazev_banky = "Air Bank" if banka == "AirBank" else "Raiffeisenbank"
        prichozi = sum(r["castka"] for r in rows if r["castka"] > 0)
        odchozi  = sum(r["castka"] for r in rows if r["castka"] < 0)
        saldo    = prichozi + odchozi

        buf = _io.BytesIO()
        doc = SimpleDocTemplate(buf, pagesize=A4,
            leftMargin=15*mm, rightMargin=15*mm,
            topMargin=15*mm, bottomMargin=15*mm)
        styles = getSampleStyleSheet()
        story = []
        story.append(Paragraph(f"<b>{nazev_banky}</b> – výpis {mesic}", styles["Title"]))
        story.append(Spacer(1, 4*mm))

        souhrn = [
            ["Příchozí", "Odchozí", "Saldo"],
            [f"{prichozi:,.2f} Kč", f"{abs(odchozi):,.2f} Kč", f"{saldo:,.2f} Kč"],
        ]
        ts = TableStyle([
            ("BACKGROUND", (0,0), (-1,0), colors.HexColor("#f0f0f0")),
            ("FONTNAME",   (0,0), (-1,0), "Helvetica-Bold"),
            ("ALIGN",      (0,0), (-1,-1), "CENTER"),
            ("GRID",       (0,0), (-1,-1), 0.5, colors.grey),
            ("FONTSIZE",   (0,0), (-1,-1), 9),
        ])
        t = Table(souhrn, colWidths=[55*mm, 55*mm, 55*mm])
        t.setStyle(ts)
        story.append(t)
        story.append(Spacer(1, 5*mm))

        hlavicka = ["Datum", "Protistrana", "Typ transakce", "Zpráva", "Částka"]
        data_rows = [hlavicka] + [
            [r["datum"], (r["nazev_protiucet"] or "")[:35],
             (r["typ_transakce"] or "")[:25], (r["zprava"] or "")[:30],
             f"{r['castka']:,.2f}"]
            for r in rows
        ]
        col_w = [22*mm, 55*mm, 38*mm, 40*mm, 25*mm]
        tbl = Table(data_rows, colWidths=col_w, repeatRows=1)
        tbl_style = TableStyle([
            ("BACKGROUND",  (0,0), (-1,0), colors.HexColor("#1e3a2f")),
            ("TEXTCOLOR",   (0,0), (-1,0), colors.white),
            ("FONTNAME",    (0,0), (-1,0), "Helvetica-Bold"),
            ("FONTSIZE",    (0,0), (-1,-1), 8),
            ("ALIGN",       (4,0), (4,-1), "RIGHT"),
            ("GRID",        (0,0), (-1,-1), 0.3, colors.HexColor("#dddddd")),
            ("ROWBACKGROUNDS", (0,1), (-1,-1), [colors.white, colors.HexColor("#f9f9f9")]),
            ("TOPPADDING",  (0,0), (-1,-1), 3),
            ("BOTTOMPADDING",(0,0), (-1,-1), 3),
        ])
        tbl.setStyle(tbl_style)
        story.append(tbl)

        doc.build(story)
        buf.seek(0)
        resp = make_response(buf.read())
        resp.headers["Content-Type"] = "application/pdf"
        resp.headers["Content-Disposition"] = f'attachment; filename="{banka}_{mesic}.pdf"'
        return resp

@app.route("/api/banky/debug")
@vyzaduj_prihlaseni
def api_banky_debug():
    try:
        import psycopg2
        db_url = os.environ.get("DATABASE_URL", "")
        conn = psycopg2.connect(db_url)
        cur = conn.cursor()
        cur.execute("SELECT COUNT(*) FROM bankovni_pohyby")
        celkem = cur.fetchone()[0]
        cur.execute("SELECT DISTINCT banka FROM bankovni_pohyby ORDER BY banka")
        banky = [r[0] for r in cur.fetchall()]
        cur.execute("SELECT DISTINCT firma_zkratka FROM bankovni_pohyby ORDER BY firma_zkratka")
        firmy = [r[0] for r in cur.fetchall()]
        cur.execute("SELECT banka, firma_zkratka, COUNT(*) FROM bankovni_pohyby GROUP BY banka, firma_zkratka ORDER BY banka, firma_zkratka")
        skupiny = [{"banka": r[0], "firma": r[1], "pocet": r[2]} for r in cur.fetchall()]
        # Reporty a vydaje
        cur.execute("SELECT MIN(datum), MAX(datum), COUNT(*) FROM reporty")
        rep = cur.fetchone()
        cur.execute("SELECT MIN(datum), MAX(datum), COUNT(*) FROM vydaje")
        vyd = cur.fetchone()
        conn.close()
        return jsonify({
            "celkem": celkem, "banky": banky, "firmy": firmy, "skupiny": skupiny,
            "reporty": {"od": rep[0], "do": rep[1], "pocet": rep[2]},
            "vydaje": {"od": vyd[0], "do": vyd[1], "pocet": vyd[2]},
        })
    except Exception as e:
        return jsonify({"error": str(e), "type": type(e).__name__}), 500

@app.route("/api/banky/oprav-soukrome", methods=["POST"])
@vyzaduj_prihlaseni
def api_banky_oprav_soukrome():
    """Přepíše firma_zkratka na _soukrome pro záznamy dané banky."""
    if session.get("role") != "admin":
        return jsonify({"error": "Pouze admin"}), 403
    d = request.json or {}
    banka = d.get("banka", "")
    if not banka:
        return jsonify({"error": "Chybí banka"}), 400
    with get_db() as conn:
        cur = conn.execute(
            "UPDATE bankovni_pohyby SET firma_zkratka='_soukrome' WHERE banka=? AND firma_zkratka != '_soukrome'",
            (banka,)
        )
        opraveno = cur.rowcount
    return jsonify({"ok": True, "opraveno": opraveno})

@app.route("/api/banky/pohyby/<int:pid>", methods=["DELETE"])
@vyzaduj_prihlaseni
def api_banky_pohyb_delete(pid):
    with get_db() as conn:
        conn.execute("DELETE FROM bankovni_pohyby WHERE id=?", (pid,))
    return jsonify({"ok": True})

@app.route("/api/parovani/navrh")
@vyzaduj_prihlaseni
def api_parovani_navrh():
    """Vrátí nezaplacené doklady a k nim hledá platbu v bankovním výpisu."""
    import datetime as _dt
    navrhy = []
    try:
        import psycopg2 as _pg2
        db_url = os.environ.get("DATABASE_URL", "")
        pg_conn = _pg2.connect(db_url)
        pg_cur = pg_conn.cursor()

        auto_zaplaceno = 0

        # 1. Přijaté faktury (kromě MAKRO) — odchozí platba
        pg_cur.execute("""
            SELECT id, firma_zkratka, dodavatel, cislo_faktury, celkem_s_dph,
                   datum_vystaveni, datum_splatnosti
            FROM faktury
            WHERE stav NOT IN ('zaplaceno','duplikat')
            AND dodavatel NOT ILIKE '%MAKRO%'
            ORDER BY datum_splatnosti, datum_vystaveni
        """)
        for r in pg_cur.fetchall():
            fid, firma, dodavatel, cislo, castka, dat_vys, dat_spl = r
            shoda = _hledej_platbu_pg(pg_cur, castka, dat_vys, dat_spl, cislo, smer='odchozi')
            if shoda:
                # Automaticky označ jako zaplaceno
                pg_cur.execute("UPDATE faktury SET stav='zaplaceno', datum_zaplaceno=%s WHERE id=%s",
                               (shoda['datum'], fid))
                auto_zaplaceno += 1
            else:
                navrhy.append({
                    "typ": "faktura", "id": fid, "firma": firma,
                    "popis": f"Faktura č.{cislo or '?'} | {dodavatel}",
                    "castka": castka, "datum": dat_vys, "datum_splatnosti": dat_spl,
                    "smer": "odchozi", "shoda": None
                })

        # 2. Výdaje — odchozí platba
        pg_cur.execute("""
            SELECT id, firma_zkratka, dodavatel, castka, datum, datum_splatnosti,
                   COALESCE(var_sym,'') as var_sym, COALESCE(popis,'') as popis
            FROM vydaje
            WHERE stav = 'nezaplaceno'
            ORDER BY datum_splatnosti, datum
        """)
        for r in pg_cur.fetchall():
            vid, firma, dodavatel, castka, datum, dat_spl, var_sym, popis = r
            shoda = _hledej_platbu_pg(pg_cur, castka, datum, dat_spl, var_sym or None, smer='odchozi')
            if shoda:
                pg_cur.execute("UPDATE vydaje SET stav='zaplaceno', datum_zaplaceno=%s WHERE id=%s",
                               (shoda['datum'], vid))
                auto_zaplaceno += 1
            else:
                navrhy.append({
                    "typ": "vydaj", "id": vid, "firma": firma,
                    "popis": f"Výdaj | {dodavatel}",
                    "detail": popis, "var_sym": var_sym,
                    "castka": castka, "datum": datum, "datum_splatnosti": dat_spl,
                    "smer": "odchozi", "shoda": None
                })

        # 3. Vystavené faktury — příchozí platba
        pg_cur.execute("""
            SELECT id, firma_zkratka, odberatel, cislo_faktury, castka,
                   datum, datum_splatnosti,
                   COALESCE(var_sym,'') as var_sym, COALESCE(popis,'') as popis
            FROM vystavene_faktury
            WHERE stav = 'nezaplaceno'
            ORDER BY datum_splatnosti, datum
        """)
        for r in pg_cur.fetchall():
            fid, firma, odberatel, cislo, castka, datum, dat_spl, var_sym, popis = r
            shoda = _hledej_platbu_pg(pg_cur, castka, datum, dat_spl, cislo or var_sym or None, smer='prichozi')
            if shoda:
                pg_cur.execute("UPDATE vystavene_faktury SET stav='zaplaceno', datum_zaplaceno=%s WHERE id=%s",
                               (shoda['datum'], fid))
                auto_zaplaceno += 1
            else:
                navrhy.append({
                    "typ": "vystavena", "id": fid, "firma": firma,
                    "popis": f"Vystavená FA č.{cislo or '?'} | {odberatel}",
                    "detail": popis, "var_sym": cislo or var_sym,
                    "castka": castka, "datum": datum, "datum_splatnosti": dat_spl,
                    "smer": "prichozi", "shoda": None
                })

        if auto_zaplaceno:
            pg_conn.commit()
        pg_conn.close()
    except Exception as e:
        return jsonify({"navrhy": [], "error": str(e)})

    return jsonify({"navrhy": navrhy, "auto_zaplaceno": auto_zaplaceno})


def _hledej_platbu_pg(pg_cur, castka, datum, datum_splatnosti, cislo_faktury, smer):
    """Hledá platbu v bankovních pohybech odpovídající dokladu."""
    import datetime as _dt
    if not datum_splatnosti and not datum:
        return None
    try:
        ref = _dt.date.fromisoformat((datum_splatnosti or datum)[:10])
        d_od = (ref - _dt.timedelta(days=30)).isoformat()
        d_do = (ref + _dt.timedelta(days=30)).isoformat()
    except Exception:
        return None

    castka_abs = abs(float(castka))
    operator = "<" if smer == "odchozi" else ">"

    try:
        # Hledej podle VS + částka
        if cislo_faktury:
            pg_cur.execute(f"""
                SELECT id, banka, datum, castka, nazev_protiucet, var_sym, firma_zkratka
                FROM bankovni_pohyby
                WHERE castka {operator} 0
                AND ABS(ABS(castka) - %s) < 1.0
                AND datum BETWEEN %s AND %s
                AND (var_sym = %s OR zprava ILIKE %s)
                LIMIT 1
            """, (castka_abs, d_od, d_do, str(cislo_faktury), f"%{cislo_faktury}%"))
            row = pg_cur.fetchone()
            if row:
                return {"pohyb_id": row[0], "banka": row[1], "datum": row[2],
                        "castka": row[3], "nazev": row[4] or "", "istota": "vs"}

        # Hledej jen podle částky
        pg_cur.execute(f"""
            SELECT id, banka, datum, castka, nazev_protiucet, var_sym, firma_zkratka
            FROM bankovni_pohyby
            WHERE castka {operator} 0
            AND ABS(ABS(castka) - %s) < 1.0
            AND datum BETWEEN %s AND %s
            LIMIT 1
        """, (castka_abs, d_od, d_do))
        row = pg_cur.fetchone()
        if row:
            return {"pohyb_id": row[0], "banka": row[1], "datum": row[2],
                    "castka": row[3], "nazev": row[4] or "", "istota": "castka"}
    except Exception:
        pass
    return None

@app.route("/api/parovani/potvrdit", methods=["POST"])
@vyzaduj_prihlaseni
def api_parovani_potvrdit():
    """Potvrdí spárování pohybu s dokladem."""
    data = request.json
    pohyb_id = data.get("pohyb_id")
    typ = data.get("typ")        # faktura / vydaj / vystavena / bez_dokladu
    doklad_id = data.get("doklad_id")
    datum_zaplaceno = data.get("datum_zaplaceno", "")

    with get_db() as conn:
        # Označ pohyb jako spárovaný
        conn.execute("""
            UPDATE bankovni_pohyby
            SET sparovano=1, sparovano_typ=?, sparovano_id=?
            WHERE id=?
        """, (typ, doklad_id, pohyb_id))

        # Označ doklad jako zaplacený
        if typ == "faktura" and doklad_id:
            conn.execute("UPDATE faktury SET stav='zaplaceno', datum_zaplaceno=? WHERE id=?",
                        (datum_zaplaceno, doklad_id))
        elif typ == "vydaj" and doklad_id:
            conn.execute("UPDATE vydaje SET stav='zaplaceno', datum_zaplaceno=? WHERE id=?",
                        (datum_zaplaceno, doklad_id))
        elif typ == "vystavena" and doklad_id:
            conn.execute("UPDATE vystavene_faktury SET stav='zaplaceno', datum_zaplaceno=? WHERE id=?",
                        (datum_zaplaceno, doklad_id))

    return jsonify({"ok": True})

@app.route("/api/parovani/po-splatnosti")
@vyzaduj_prihlaseni
def api_parovani_po_splatnosti():
    """Vrátí všechny nezaplacené doklady po splatnosti ze všech firem."""
    import datetime as _dt, psycopg2 as _pg2
    dnes = _dt.date.today().isoformat()
    vysledky = []
    try:
        db_url = os.environ.get("DATABASE_URL", "")
        pg_conn = _pg2.connect(db_url)
        pg_cur = pg_conn.cursor()

        # Faktury (kromě MAKRO)
        pg_cur.execute("""
            SELECT id, firma_zkratka, dodavatel, cislo_faktury, celkem_s_dph,
                   datum_splatnosti, COALESCE(var_sym,'') as var_sym
            FROM faktury
            WHERE stav NOT IN ('zaplaceno','duplikat')
            AND dodavatel NOT ILIKE '%MAKRO%'
            AND datum_splatnosti != '' AND datum_splatnosti < %s
            ORDER BY datum_splatnosti
        """, (dnes,))
        for r in pg_cur.fetchall():
            fid, firma, dodavatel, cislo, castka, ds, var_sym = r
            vysledky.append({
                "typ": "faktura", "id": fid, "firma": firma,
                "popis": f"Faktura č.{cislo or '?'} | {dodavatel}",
                "detail": dodavatel, "var_sym": cislo or var_sym or "",
                "castka": castka, "datum_splatnosti": ds,
                "dnu_po": ((_dt.date.fromisoformat(dnes) - _dt.date.fromisoformat(ds)).days if ds else 0),
                "smer": "odchozi"
            })

        # Výdaje
        pg_cur.execute("""
            SELECT id, firma_zkratka, dodavatel, castka, datum_splatnosti,
                   COALESCE(var_sym,'') as var_sym, COALESCE(popis,'') as popis
            FROM vydaje
            WHERE stav = 'nezaplaceno'
            AND datum_splatnosti != '' AND datum_splatnosti < %s
            ORDER BY datum_splatnosti
        """, (dnes,))
        for r in pg_cur.fetchall():
            vid, firma, dodavatel, castka, ds, var_sym, popis = r
            vysledky.append({
                "typ": "vydaj", "id": vid, "firma": firma,
                "popis": f"Výdaj | {dodavatel}",
                "detail": popis, "var_sym": var_sym or "",
                "castka": castka, "datum_splatnosti": ds,
                "dnu_po": ((_dt.date.fromisoformat(dnes) - _dt.date.fromisoformat(ds)).days if ds else 0),
                "smer": "odchozi"
            })

        # Vystavené faktury
        pg_cur.execute("""
            SELECT id, firma_zkratka, odberatel, cislo_faktury, castka,
                   datum_splatnosti, COALESCE(var_sym,'') as var_sym, COALESCE(popis,'') as popis
            FROM vystavene_faktury
            WHERE stav = 'nezaplaceno'
            AND datum_splatnosti != '' AND datum_splatnosti < %s
            ORDER BY datum_splatnosti
        """, (dnes,))
        for r in pg_cur.fetchall():
            fid, firma, odberatel, cislo, castka, ds, var_sym, popis = r
            vysledky.append({
                "typ": "vystavena", "id": fid, "firma": firma,
                "popis": f"Vystavená FA č.{cislo or '?'} | {odberatel}",
                "detail": popis, "var_sym": cislo or var_sym or "",
                "castka": castka, "datum_splatnosti": ds,
                "dnu_po": ((_dt.date.fromisoformat(dnes) - _dt.date.fromisoformat(ds)).days if ds else 0),
                "smer": "prichozi"
            })

        pg_conn.close()
        vysledky.sort(key=lambda x: x.get("datum_splatnosti",""))
    except Exception as e:
        return jsonify({"po_splatnosti": [], "error": str(e)})
    return jsonify({"po_splatnosti": vysledky})

# ── API: REPORTY ──────────────────────────────────────────────────────────────
@app.route("/api/reporty/nahrat-foto", methods=["POST"])
@vyzaduj_prihlaseni
def api_report_nahrat_foto():
    if "soubor" not in request.files:
        return jsonify({"error": "Žádný soubor"}), 400
    f = request.files["soubor"]
    if not f.filename:
        return jsonify({"error": "Prázdný soubor"}), 400

    fname = secure_filename(f.filename)
    ts    = datetime.now().strftime("%Y%m%d_%H%M%S_")
    fname = "report_" + ts + fname
    fpath = os.path.join(UPLOAD_DIR, fname)
    f.save(fpath)

    parsed, err = parse_report_image_claude(fpath)
    if err:
        return jsonify({"error": err}), 200

    report = build_report_from_parsed(parsed)

    gcs_url = None
    try:
        gcs_url = upload_to_gcs(fpath, f"reporty/{fname}")
    except Exception as e:
        app.logger.warning(f"GCS upload reportu selhal: {e}")

    report["soubor_url"] = gcs_url
    return jsonify(report)


@app.route("/api/reporty/nahrat-foto-pouze", methods=["POST"])
@vyzaduj_prihlaseni
def api_report_nahrat_foto_pouze():
    """Nahraje fotku na GCS bez OCR — používá se při editaci existujícího reportu."""
    if "soubor" not in request.files:
        return jsonify({"error": "Žádný soubor"}), 400
    f = request.files["soubor"]
    if not f.filename:
        return jsonify({"error": "Prázdný soubor"}), 400

    fname = secure_filename(f.filename)
    ts    = datetime.now().strftime("%Y%m%d_%H%M%S_")
    fname = "report_" + ts + fname
    fpath = os.path.join(UPLOAD_DIR, fname)
    f.save(fpath)

    gcs_url = None
    try:
        gcs_url = upload_to_gcs(fpath, f"reporty/{fname}")
    except Exception as e:
        app.logger.warning(f"GCS upload reportu selhal: {e}")

    return jsonify({"ok": True, "soubor_url": gcs_url})


@app.route("/api/reporty/nahrat-text", methods=["POST"])
@vyzaduj_prihlaseni
def api_report_nahrat_text():
    text = request.json.get("text", "").strip()
    if not text:
        return jsonify({"error": "Prázdný text"}), 400

    parsed, err = parse_report_text(text)
    if err:
        return jsonify({"error": err}), 200

    report = build_report_from_parsed(parsed)
    return jsonify(report)


@app.route("/api/reporty", methods=["GET"])
@vyzaduj_prihlaseni
def api_reporty_list():
    od  = request.args.get("od", "")
    do_ = request.args.get("do", "")
    clauses, params = [], []
    if od:  clauses.append("datum>=?"); params.append(od)
    if do_: clauses.append("datum<=?"); params.append(do_)
    else:
        clauses.append("datum<=?"); params.append(date.today().isoformat())
    where = ("WHERE " + " AND ".join(clauses)) if clauses else ""
    with get_db() as conn:
        rows = conn.execute(f"""
            SELECT * FROM reporty {where} ORDER BY datum DESC LIMIT 500
        """, params).fetchall()
    return jsonify([dict(r) for r in rows])


@app.route("/api/reporty", methods=["POST"])
@vyzaduj_prihlaseni
def api_report_ulozit():
    data = request.json
    if not data.get("datum"):
        return jsonify({"error": "Chybí datum"}), 400

    karty    = float(data.get("karty", 0) or 0)
    kov      = float(data.get("kov", 0) or 0)
    papir    = float(data.get("papir", 0) or 0)
    vydaje   = float(data.get("vydaje", 0) or 0)
    hotovost = kov + papir
    trzba    = karty + hotovost + vydaje
    pk50_ks  = int(data.get("pk50_ks", 0) or 0)
    pk100_ks = int(data.get("pk100_ks", 0) or 0)
    pk_celkem  = pk50_ks * 50 + pk100_ks * 100
    trzba_vcpk = trzba + pk_celkem

    firma = data.get("firma_zkratka", "")
    with get_db() as conn:
        # Zjistit jestli existuje záznam se stejným datem
        existing = conn.execute("SELECT id FROM reporty WHERE datum=?", (data["datum"],)).fetchone()
        duplicita_id = None
        if existing:
            duplicita_id = existing["id"] if isinstance(existing, dict) else existing[0]

        soubor_url = data.get("soubor_url") or None
        cur = conn.execute("""
            INSERT INTO reporty (datum,den,smena,karty,kov,papir,hotovost,vydaje,
            trzba,trzba_vcpk,pk50_ks,pk100_ks,pk_celkem,
            pizza_cela,pizza_ctvrt,burger,talire,burtgulas,poznamka,firma_zkratka,soubor_url,duplicita_id)
            VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
        """, (
            data["datum"], data.get("den",""), data.get("smena",""),
            karty, kov, papir, hotovost, vydaje, trzba, trzba_vcpk,
            pk50_ks, pk100_ks, pk_celkem,
            int(data.get("pizza_cela",0) or 0), int(data.get("pizza_ctvrt",0) or 0),
            int(data.get("burger",0) or 0), int(data.get("talire",0) or 0),
            int(data.get("burtgulas",0) or 0),
            data.get("poznamka",""), firma, soubor_url, duplicita_id
        ))
        rid = cur.lastrowid
        # Označit i původní záznam jako duplicitu pokud ještě není
        if duplicita_id:
            conn.execute(
                "UPDATE reporty SET duplicita_id=? WHERE id=? AND duplicita_id IS NULL",
                (rid, duplicita_id)
            )

    return jsonify({"ok": True, "id": rid, "duplicita": duplicita_id is not None})


@app.route("/api/reporty/<int:rid>", methods=["GET"])
@vyzaduj_prihlaseni
def api_report_get(rid):
    with get_db() as conn:
        r = conn.execute("SELECT * FROM reporty WHERE id=?", (rid,)).fetchone()
    if not r:
        return jsonify({"error": "Nenalezen"}), 404
    return jsonify(dict(r))

@app.route("/api/reporty/<int:rid>", methods=["PUT"])
@vyzaduj_prihlaseni
def api_report_update(rid):
    data = request.json

    with get_db() as conn:
        existing = conn.execute("SELECT * FROM reporty WHERE id=?", (rid,)).fetchone()
        if not existing:
            return jsonify({"error": "Report nenalezen"}), 404

        # Speciální případ: jen smazat duplicita_id
        if data.get("_jen_duplicita_id"):
            conn.execute("UPDATE reporty SET duplicita_id=NULL WHERE id=?", (rid,))
            return jsonify({"ok": True})

        if not data.get("datum"):
            return jsonify({"error": "Chybí datum"}), 400

        karty    = float(data.get("karty", 0) or 0)
        kov      = float(data.get("kov", 0) or 0)
        papir    = float(data.get("papir", 0) or 0)
        vydaje   = float(data.get("vydaje", 0) or 0)
        hotovost = kov + papir
        trzba    = karty + hotovost + vydaje
        pk50_ks  = int(data.get("pk50_ks", 0) or 0)
        pk100_ks = int(data.get("pk100_ks", 0) or 0)
        pk_celkem  = pk50_ks * 50 + pk100_ks * 100
        trzba_vcpk = trzba + pk_celkem

        # Zachovat stávající soubor_url pokud nebylo nahráno nové
        soubor_url = data.get("soubor_url") or (existing.get("soubor_url") if hasattr(existing, "get") else None)

        conn.execute("""
            UPDATE reporty SET datum=?,den=?,smena=?,karty=?,kov=?,papir=?,hotovost=?,
            vydaje=?,trzba=?,trzba_vcpk=?,pk50_ks=?,pk100_ks=?,pk_celkem=?,
            pizza_cela=?,pizza_ctvrt=?,burger=?,talire=?,burtgulas=?,poznamka=?,
            firma_zkratka=?,soubor_url=?
            WHERE id=?
        """, (
            data["datum"], data.get("den",""), data.get("smena",""),
            karty, kov, papir, hotovost, vydaje, trzba, trzba_vcpk,
            pk50_ks, pk100_ks, pk_celkem,
            int(data.get("pizza_cela",0) or 0), int(data.get("pizza_ctvrt",0) or 0),
            int(data.get("burger",0) or 0), int(data.get("talire",0) or 0),
            int(data.get("burtgulas",0) or 0),
            data.get("poznamka",""), data.get("firma_zkratka",""),
            soubor_url, rid
        ))
    return jsonify({"ok": True, "id": rid})

@app.route("/api/reporty/<int:rid>", methods=["DELETE"])
@vyzaduj_prihlaseni
def api_report_delete(rid):
    with get_db() as conn:
        conn.execute("DELETE FROM reporty WHERE id=?", (rid,))
    return jsonify({"ok": True})


@app.route("/api/reporty/smaz-budouci", methods=["POST"])
@vyzaduj_prihlaseni
def api_reporty_smaz_budouci():
    dnes = date.today().isoformat()
    with get_db() as conn:
        cur = conn.execute("DELETE FROM reporty WHERE datum > ?", (dnes,))
        smazano = cur.rowcount
    return jsonify({"ok": True, "smazano": smazano})


@app.route("/api/reporty/import-xlsx", methods=["POST"])
@vyzaduj_prihlaseni
def api_report_import_xlsx():
    if "soubor" not in request.files:
        return jsonify({"error": "Žádný soubor"}), 400
    f = request.files["soubor"]
    fname = secure_filename(f.filename)
    fpath = os.path.join(UPLOAD_DIR, "import_" + fname)
    f.save(fpath)

    try:
        wb = openpyxl.load_workbook(fpath, data_only=True)
        imported = 0
        skipped  = 0
        errors   = []

        den_map = {
            "po": "Pondělí", "út": "Úterý", "st": "Středa",
            "čt": "Čtvrtek", "pá": "Pátek", "so": "Sobota", "ne": "Neděle"
        }
        mesic_map = {
            "LEDEN": 1, "ÚNOR": 2, "BŘEZEN": 3, "DUBEN": 4,
            "KVĚTEN": 5, "ČERVEN": 6, "ČERVENEC": 7, "SRPEN": 8,
            "ZÁŘÍ": 9, "ŘÍJEN": 10, "LISTOPAD": 11, "PROSINEC": 12,
            # Varianty bez diakritiky
            "UNOR": 2, "BREZEN": 3, "KVETEN": 5, "CERVEN": 6,
            "CERVENEC": 7, "ZARI": 9, "RIJEN": 10,
        }

        rows_to_insert = []
        for sheet_name in wb.sheetnames:
            if sheet_name not in ("2023", "2024", "2025", "2026"):
                continue
            year = int(sheet_name)
            ws = wb[sheet_name]

            current_mesic = None
            dnes = date.today()
            konec_import = date(dnes.year, dnes.month, dnes.day)
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not row:
                    continue
                # Řádek záhlaví měsíce — row[0] je None, row[1] je název měsíce
                if row[0] is None:
                    if row[1] and str(row[1]).upper() in mesic_map:
                        current_mesic = mesic_map[str(row[1]).upper()]
                    continue
                if str(row[0]).upper() in ("SOUČET", "DNÍ", "PRŮMĚR", "SOU\ČET"):
                    continue
                # Datový řádek — aktualizuj měsíc pokud je v sloupci B
                if row[1] and str(row[1]).upper() in mesic_map:
                    current_mesic = mesic_map[str(row[1]).upper()]

                try:
                    den_cislo = int(row[0])
                except (TypeError, ValueError):
                    continue

                if not current_mesic:
                    continue

                try:
                    datum_test = date(year, current_mesic, den_cislo)
                    if datum_test > konec_import:
                        skipped += 1
                        continue
                except ValueError:
                    continue

                try:
                    datum_iso = date(year, current_mesic, den_cislo).isoformat()
                except ValueError:
                    errors.append(f"Neplatné datum: {year}-{current_mesic}-{den_cislo}")
                    continue

                den_str = den_map.get(str(row[2] or "").lower(), str(row[2] or ""))
                trzba_vcpk = float(row[3] or 0)
                karty      = float(row[4] or 0)
                hotovost   = float(row[5] or 0)
                vydaje     = float(row[6] or 0)
                trzba      = float(row[7] or 0)
                pk50_ks    = int(row[8] or 0)
                pk100_ks   = int(row[9] or 0)
                pk_celkem  = float(row[10] or 0)
                pizza_cela = int(row[11] or 0)
                pizza_ctvrt= int(row[12] or 0)
                burger     = int(row[13] or 0)
                talire     = int(row[14] or 0)
                burtgulas  = int(row[15] or 0)
                smena      = normalize_jmena(str(row[16] or ""))

                kov   = 0
                papir = hotovost

                rows_to_insert.append((
                    datum_iso, den_str, smena, karty, kov, papir, hotovost,
                    vydaje, trzba, trzba_vcpk, pk50_ks, pk100_ks, pk_celkem,
                    pizza_cela, pizza_ctvrt, burger, talire, burtgulas
                ))

        with get_db() as conn:
            for params in rows_to_insert:
                existing = conn.execute("SELECT id FROM reporty WHERE datum=?", (params[0],)).fetchone()
                if existing:
                    skipped += 1
                    continue
                conn.execute("""
                    INSERT INTO reporty (datum,den,smena,karty,kov,papir,hotovost,vydaje,
                    trzba,trzba_vcpk,pk50_ks,pk100_ks,pk_celkem,
                    pizza_cela,pizza_ctvrt,burger,talire,burtgulas)
                    VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
                """, params)
                imported += 1

        return jsonify({"ok": True, "imported": imported, "skipped": skipped, "errors": errors[:10]})

    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/api/reporty/karty-alert")
@vyzaduj_prihlaseni
def api_karty_alert():
    with get_db() as conn:
        total_row = conn.execute("""
            SELECT COALESCE(SUM(karty),0) as total
            FROM reporty
            WHERE datum >= date('now','-12 months')
        """).fetchone()
        total = _first_val(total_row)
        per_firma = conn.execute("""
            SELECT firma_zkratka, COALESCE(SUM(karty),0) as karty_12m
            FROM reporty
            WHERE datum >= date('now','-12 months')
            GROUP BY firma_zkratka
            ORDER BY karty_12m DESC
        """).fetchall()
    LIMIT = 1500000
    firmy_alert = []
    for r in per_firma:
        firma = r["firma_zkratka"] or "—"
        k = round(r["karty_12m"], 2)
        firmy_alert.append({
            "firma": firma,
            "karty_12m": k,
            "procent": round(k / LIMIT * 100, 1),
            "alert": k >= LIMIT,
            "varovani": k >= 1200000,
        })
    return jsonify({
        "karty_12m": round(total, 2),
        "limit": LIMIT,
        "procent": round(total / LIMIT * 100, 1),
        "alert": total >= LIMIT,
        "varovani": total >= 1200000,
        "per_firma": firmy_alert,
    })


@app.route("/api/statistiky/rucni-data")
@vyzaduj_prihlaseni
def api_stat_rucni_get():
    with get_db() as conn:
        rows = conn.execute("SELECT rok, mesic, hodnota, typ FROM stat_rucni_data ORDER BY rok, mesic").fetchall()
    return jsonify([dict(r) for r in rows])

@app.route("/api/statistiky/rucni-data", methods=["POST"])
@vyzaduj_prihlaseni
def api_stat_rucni_set():
    data = request.json
    rok     = str(data.get("rok","")).strip()
    mesic   = str(data.get("mesic","")).strip().zfill(2)
    hodnota = float(data.get("hodnota", 0) or 0)
    typ     = data.get("typ", "trzba_vcpk_prumer")
    if not rok or not mesic:
        return jsonify({"error": "Chybí rok nebo měsíc"}), 400
    with get_db() as conn:
        conn.execute("""
            INSERT INTO stat_rucni_data (rok, mesic, hodnota, typ)
            VALUES (?,?,?,?)
            ON CONFLICT (rok, mesic, typ) DO UPDATE SET hodnota=EXCLUDED.hodnota
        """, (rok, mesic, hodnota, typ))
    return jsonify({"ok": True})

@app.route("/api/statistiky/rucni-data", methods=["DELETE"])
@vyzaduj_prihlaseni
def api_stat_rucni_delete():
    data = request.json
    rok   = str(data.get("rok","")).strip()
    mesic = str(data.get("mesic","")).strip().zfill(2)
    typ   = data.get("typ", "trzba_vcpk_prumer")
    with get_db() as conn:
        conn.execute("DELETE FROM stat_rucni_data WHERE rok=? AND mesic=? AND typ=?", (rok, mesic, typ))
    return jsonify({"ok": True})


@app.route("/api/statistiky/prehled-pl")
@vyzaduj_prihlaseni
def api_statistiky_prehled_pl():
    """Vrátí data pro 4 tabulky: náklady, průměry po letech, marže, P&L."""
    import datetime as _dt, calendar as _cal
    firma = request.args.get("firma", "")
    rok   = request.args.get("rok", "")
    fw  = "AND firma_zkratka=?" if firma else ""
    fp  = [firma] if firma else []
    ffw = ("AND firma_zkratka=?" if firma else "")

    # Rozsah
    if rok:
        od, do = f"{rok}-01-01", f"{rok}-12-31"
    else:
        od, do = "2020-01-01", _dt.date.today().isoformat()

    from decimal import Decimal
    def _f(v): return float(v) if isinstance(v, Decimal) else (float(v) if v is not None else 0.0)

    with get_db() as conn:
        # 1. Tržby po měsících (pro průměry a P&L)
        trzby = conn.execute(f"""
            SELECT TO_CHAR(datum::date,'YYYY') as rok,
                   TO_CHAR(datum::date,'MM') as mesic,
                   COUNT(*) as dni,
                   ROUND(SUM(trzba_vcpk)::numeric,0) as trzba_vcpk,
                   ROUND(SUM(karty+hotovost+vydaje)::numeric,0) as trzba
            FROM reporty
            WHERE datum IS NOT NULL AND datum != ''
              AND datum >= ? AND datum <= ? {fw} AND trzba_vcpk > 0
            GROUP BY rok, mesic ORDER BY rok, mesic
        """, [od, do] + fp).fetchall()

        # 2. Faktury po měsících
        faktury = conn.execute(f"""
            SELECT TO_CHAR(datum_vystaveni::date,'YYYY') as rok,
                   TO_CHAR(datum_vystaveni::date,'MM') as mesic,
                   ROUND(SUM(celkem_s_dph)::numeric,0) as castka
            FROM faktury
            WHERE datum_vystaveni IS NOT NULL AND datum_vystaveni != ''
              AND datum_vystaveni >= ? AND datum_vystaveni <= ? {ffw}
            GROUP BY rok, mesic ORDER BY rok, mesic
        """, [od, do] + fp).fetchall()

        # 3. Ruční výdaje po měsících (provozní)
        vydaje = conn.execute(f"""
            SELECT TO_CHAR(datum::date,'YYYY') as rok,
                   TO_CHAR(datum::date,'MM') as mesic,
                   ROUND(SUM(castka)::numeric,0) as castka
            FROM vydaje
            WHERE datum IS NOT NULL AND datum != ''
              AND datum >= ? AND datum <= ?
              AND COALESCE(typ,'provozni')='provozni' {ffw}
            GROUP BY rok, mesic ORDER BY rok, mesic
        """, [od, do] + fp).fetchall()

        # 4. Výplaty po měsících
        vyplaty = conn.execute(f"""
            SELECT TO_CHAR(datum::date,'YYYY') as rok,
                   TO_CHAR(datum::date,'MM') as mesic,
                   ROUND(SUM(castka)::numeric,0) as castka
            FROM vyplaty
            WHERE datum IS NOT NULL AND datum != ''
              AND datum >= ? AND datum <= ? {ffw}
            GROUP BY rok, mesic ORDER BY rok, mesic
        """, [od, do] + fp).fetchall()

        # 5. Paušální odvody – suma všech zaměstnanců (měsíční fix)
        odvody_row = conn.execute("SELECT COALESCE(SUM(castka),0) as suma FROM pausalni_odvody").fetchone()
        odvody_mesic = _f(odvody_row["suma"] if isinstance(odvody_row, dict) else odvody_row[0])

    # Sestavit dict rok-mesic
    def _to_dict(rows, key="castka"):
        d = {}
        for r in rows:
            rm = r["rok"] if isinstance(r, dict) else r[0]
            mm = r["mesic"] if isinstance(r, dict) else r[1]
            d[(rm, mm)] = _f(r[key] if isinstance(r, dict) else r[2])
        return d

    def _to_dict2(rows):
        d = {}
        for r in rows:
            rm = r["rok"] if isinstance(r, dict) else r[0]
            mm = r["mesic"] if isinstance(r, dict) else r[1]
            d[(rm, mm)] = {
                "dni": int(_f(r["dni"] if isinstance(r, dict) else r[2])),
                "trzba_vcpk": _f(r["trzba_vcpk"] if isinstance(r, dict) else r[3]),
                "trzba": _f(r["trzba"] if isinstance(r, dict) else r[4]),
            }
        return d

    t_dict  = _to_dict2(trzby)
    f_dict  = _to_dict(faktury)
    v_dict  = _to_dict(vydaje)
    p_dict  = _to_dict(vyplaty)

    # Unikátní roky
    roky = sorted(set(k[0] for k in list(t_dict.keys())+list(f_dict.keys())+list(p_dict.keys())))

    # Sestavit výsledek po měsících
    result = []
    for mi in range(1, 13):
        m = f"{mi:02d}"
        row = {"mesic": m}
        for r in roky:
            td = t_dict.get((r, m), {})
            fakt = f_dict.get((r, m), 0)
            vyda = v_dict.get((r, m), 0)
            vypl = p_dict.get((r, m), 0)
            odv  = odvody_mesic if td.get("dni", 0) > 0 else 0
            naklady = fakt + vyda + vypl + odv
            trzba_vcpk = td.get("trzba_vcpk", 0)
            trzba = td.get("trzba", 0)
            row[r] = {
                "dni": td.get("dni", 0),
                "trzba_vcpk": trzba_vcpk,
                "trzba": trzba,
                "faktury": fakt,
                "vydaje": vyda,
                "vyplaty": vypl,
                "odvody": odv,
                "naklady": naklady,
                "marze_czk": trzba - fakt,
                "marze_pct": round((trzba - fakt) / fakt * 100, 1) if fakt > 0 else None,
                "pl": trzba_vcpk - naklady,
            }
        result.append(row)

    return jsonify({"mesice": result, "roky": roky, "odvody_mesic": odvody_mesic})


@app.route("/api/statistiky/trzby-mesice")
@vyzaduj_prihlaseni
def api_statistiky_trzby_mesice():
    firma = request.args.get("firma", "")
    rok   = request.args.get("rok", "")
    clauses = ["datum <= ?"]
    params  = [date.today().isoformat()]
    if firma:
        clauses.append("firma_zkratka=?"); params.append(firma)
    if rok:
        clauses.append("datum >= ?"); params.append(f"{rok}-01-01")
        clauses.append("datum <= ?"); params.append(f"{rok}-12-31")
    where = "WHERE " + " AND ".join(clauses)
    with get_db() as conn:
        rows = conn.execute(f"""
            SELECT
                TO_CHAR(NULLIF(datum,'')::date,'YYYY') as rok,
                TO_CHAR(NULLIF(datum,'')::date,'MM')   as mesic,
                COUNT(*) as dni,
                ROUND(SUM(karty+hotovost+vydaje)::numeric,0)  as trzba,
                ROUND(SUM(trzba_vcpk)::numeric,0)             as trzba_vcpk,
                ROUND(SUM(karty)::numeric,0)                  as karty,
                ROUND(SUM(hotovost)::numeric,0)               as hotovost,
                COALESCE(SUM(pk50_ks),0)                      as pk50,
                COALESCE(SUM(pk100_ks),0)                     as pk100,
                COALESCE(SUM(pizza_cela),0)                   as pizza,
                COALESCE(SUM(pizza_ctvrt),0)                  as pizza_ctvrt,
                COALESCE(SUM(burger),0)                       as burger,
                COALESCE(SUM(burtgulas),0)                    as bgulas
            FROM reporty {where}
            AND trzba_vcpk > 0
            GROUP BY rok, mesic
            ORDER BY rok DESC, mesic DESC
        """, params).fetchall()
    from decimal import Decimal
    def _f(v): return int(v) if isinstance(v, Decimal) else (float(v) if v is not None else 0)
    return jsonify([{k: _f(v) for k,v in dict(r).items()} for r in rows])


@app.route("/api/statistiky/mesice")
@vyzaduj_prihlaseni
def api_statistiky_mesice():
    firma = request.args.get("firma", "")
    clauses = ["datum <= ?"]
    params  = [date.today().isoformat()]
    if firma:
        clauses.append("firma_zkratka=?")
        params.append(firma)
    where = "WHERE " + " AND ".join(clauses)
    with get_db() as conn:
        rows = conn.execute(f"""
            SELECT
                strftime('%Y', datum) as rok,
                strftime('%m', datum) as mesic,
                COUNT(*) as dni,
                ROUND((SUM(trzba_vcpk))::numeric,2)  as trzba_vcpk_sum,
                ROUND((AVG(trzba_vcpk))::numeric,2)  as trzba_vcpk_avg,
                ROUND((SUM(karty))::numeric,2)       as karty_sum,
                ROUND((AVG(karty))::numeric,2)       as karty_avg,
                ROUND((SUM(hotovost))::numeric,2)    as hotovost_sum,
                ROUND((AVG(hotovost))::numeric,2)    as hotovost_avg,
                ROUND((SUM(vydaje))::numeric,2)      as vydaje_sum,
                ROUND((SUM(pk_celkem))::numeric,2)   as pk_celkem_sum,
                SUM(pizza_cela)           as pizza_cela_sum,
                SUM(pizza_ctvrt)          as pizza_ctvrt_sum,
                SUM(burger)               as burger_sum,
                SUM(talire)               as talire_sum,
                SUM(burtgulas)            as burtgulas_sum,
                ROUND((AVG(pizza_cela))::numeric,1)  as pizza_cela_avg,
                ROUND((AVG(pizza_ctvrt))::numeric,1) as pizza_ctvrt_avg,
                ROUND((AVG(burger))::numeric,1)      as burger_avg,
                ROUND((AVG(talire))::numeric,1)      as talire_avg,
                ROUND((AVG(burtgulas))::numeric,1)   as burtgulas_avg
            FROM reporty {where}
            AND trzba_vcpk > 0
            GROUP BY rok, mesic
            ORDER BY rok DESC, mesic DESC
        """, params).fetchall()
    return jsonify([dict(r) for r in rows])


@app.route("/api/statistiky/roky")
@vyzaduj_prihlaseni
def api_statistiky_roky():
    with get_db() as conn:
        rows = conn.execute("""
            SELECT
                strftime('%Y', datum) as rok,
                strftime('%m', datum) as mesic,
                ROUND((AVG(trzba_vcpk))::numeric,0) as prumer_den
            FROM reporty
            WHERE datum <= ? AND trzba_vcpk > 0
            GROUP BY rok, mesic
            ORDER BY rok, mesic
        """, (date.today().isoformat(),)).fetchall()
    return jsonify([dict(r) for r in rows])


@app.route("/api/statistiky/prehled")
@vyzaduj_prihlaseni
def api_statistiky_prehled():
    import datetime as _dt
    rok  = request.args.get("rok", str(_dt.date.today().year))
    firma = request.args.get("firma", "")
    od = f"{rok}-01-01"
    do = f"{rok}-12-31"
    fw = "AND r.firma_zkratka=?" if firma else ""
    fp = [firma] if firma else []
    with get_db() as conn:
        # Tržby + karty + hotovost + výdaje + poukazky z reportů
        rows_r = conn.execute(f"""
            SELECT
                TO_CHAR(NULLIF(datum,'')::date, 'MM') as mesic,
                ROUND(SUM(karty)::numeric,0)      as karty,
                ROUND(SUM(hotovost)::numeric,0)   as hotovost,
                ROUND(SUM(trzba_vcpk)::numeric,0) as trzba,
                ROUND(SUM(vydaje)::numeric,0)     as vydaje_rep,
                ROUND(SUM(pk_celkem)::numeric,0)  as poukazky
            FROM reporty r
            WHERE datum >= ? AND datum <= ? {fw}
            GROUP BY mesic ORDER BY mesic
        """, [od, do] + fp).fetchall()
        # Náklady z faktur
        rows_f = conn.execute(f"""
            SELECT
                TO_CHAR(NULLIF(datum_vystaveni,'')::date, 'MM') as mesic,
                ROUND(SUM(celkem_s_dph)::numeric,0) as faktury
            FROM faktury
            WHERE datum_vystaveni >= ? AND datum_vystaveni <= ?
            {"AND firma_zkratka=?" if firma else ""}
            GROUP BY mesic ORDER BY mesic
        """, [od, do] + fp).fetchall()
    # Sloučit do dict mesic→data
    data = {}
    for m in range(1, 13):
        data[f"{m:02d}"] = {"karty":0,"hotovost":0,"trzba":0,"vydaje_rep":0,"poukazky":0,"faktury":0}
    for r in rows_r:
        m = r["mesic"] if isinstance(r, dict) else r[0]
        d = dict(r) if isinstance(r, dict) else {"mesic":r[0],"karty":r[1],"hotovost":r[2],"trzba":r[3],"vydaje_rep":r[4],"poukazky":r[5]}
        if m in data:
            data[m].update({k: float(v or 0) for k,v in d.items() if k != "mesic"})
    for r in rows_f:
        m = r["mesic"] if isinstance(r, dict) else r[0]
        v = float((r["faktury"] if isinstance(r, dict) else r[1]) or 0)
        if m in data:
            data[m]["faktury"] = v
    result = []
    for m, d in data.items():
        d["mesic"] = m
        d["naklady"] = d["vydaje_rep"] + d["faktury"]
        result.append(d)
    return jsonify(result)


@app.route("/api/statistiky/mesic-detail")
@vyzaduj_prihlaseni
def api_statistiky_mesic_detail():
    rok   = request.args.get("rok", "")
    mesic = request.args.get("mesic", "")  # "01" až "12"
    firma = request.args.get("firma", "")
    if not rok or not mesic:
        return jsonify([])
    od = f"{rok}-{mesic}-01"
    # poslední den měsíce
    import calendar as _cal
    posledni = _cal.monthrange(int(rok), int(mesic))[1]
    do = f"{rok}-{mesic}-{posledni:02d}"
    fw = "AND firma_zkratka=?" if firma else ""
    fp = [firma] if firma else []
    with get_db() as conn:
        rows = conn.execute(f"""
            SELECT datum, den, smena, firma_zkratka,
                karty, hotovost, trzba_vcpk as trzba,
                vydaje, pk_celkem, pk50_ks, pk100_ks,
                burger, burtgulas, pizza_cela, pizza_ctvrt, talire
            FROM reporty
            WHERE datum >= ? AND datum <= ? {fw}
            ORDER BY datum
        """, [od, do] + fp).fetchall()
    return jsonify([dict(r) for r in rows])


@app.route("/api/ai-dotaz", methods=["POST"])
@vyzaduj_prihlaseni
def api_ai_dotaz():
    import datetime as _dt
    data   = request.json or {}
    dotaz  = data.get("dotaz", "").strip()
    rok_raw = data.get("rok", str(_dt.date.today().year))
    if rok_raw in ("", "vše", "vse", "Vše", "all"):
        rok = None
        rok_od = "2020-01-01"
        rok_do = "2099-12-31"
        rok_label = "všechny roky"
    else:
        rok = str(rok_raw)
        rok_od = f"{rok}-01-01"
        rok_do = f"{rok}-12-31"
        rok_label = rok
    firma  = data.get("firma", "")
    if not dotaz:
        return jsonify({"chyba": "Prázdný dotaz"}), 400
    api_key = os.environ.get("ANTHROPIC_API_KEY", "")
    if not api_key:
        return jsonify({"chyba": "ANTHROPIC_API_KEY není nastaven"}), 500

    # Zjistit práva přihlášeného uživatele
    role = session.get("role", "")
    je_admin = (role == "admin")
    def ma_pravo(sekce):
        if je_admin: return True
        prava = get_prava_z_db()
        return prava.get(role, {}).get(sekce, False)

    # Detekce relevantních sekcí z dotazu
    sekce_override = data.get("sekce", [])  # manuální override z frontendu

    def _je_relevantni(klicova_slova, dotaz_lower):
        return any(k in dotaz_lower for k in klicova_slova)

    dotaz_lower = dotaz.lower()
    if sekce_override:
        # Uživatel ručně vybral sekce
        nactist = set(sekce_override)
    else:
        # Automatická detekce
        nactist = set()
        if _je_relevantni(["tržb","trzb","burger","pizza","kart","hotov","report","směn","smen","talíř","talir","burtgul","pk","poukazk","výdaj v report"], dotaz_lower):
            nactist.add("reporty")
        if _je_relevantni(["faktur","nákup","náklad","dodavat","makro","účet","utrat"], dotaz_lower):
            nactist.add("faktury")
        if _je_relevantni(["zboží","zbozi","položk","polozk","klobás","klobas","maso","sýr","syr","zelenin","nápoj","napoj","pivo","víno","vino","chléb","chleb","uzenin","párk","park","sekan","parek"], dotaz_lower):
            nactist.add("zbozi")
        if _je_relevantni(["výplat","vyplat","mzd","plat","zaměstnan","zamestnan"], dotaz_lower):
            nactist.add("vyplaty")
        if _je_relevantni(["výdaj","vydaj","náklad","naklad","soukrom"], dotaz_lower):
            nactist.add("vydaje")
        if _je_relevantni(["dokument","smlouv","pojistk","faktur","certif","list","ppas","plynáren","elektr","energie"], dotaz_lower):
            nactist.add("dokumenty")
        if _je_relevantni(["peněženk","penezenk","hotovost","sporeni","majetek","úspor"], dotaz_lower):
            nactist.add("penezenka")
        if _je_relevantni(["vystaven","odberatel","bauhaus","fakturace"], dotaz_lower):
            nactist.add("vystavene")
        if _je_relevantni(["bank","výpis","pohyb","platb","převod","prevod","transakc","raiffa","airbank","utrat","výdaj","vydaj","prosin","říjen","rijen","zari","srpen","cerven","kveten","duben","brezen","unor","leden"], dotaz_lower):
            nactist.add("banky")
        # Pokud nic nedetekováno — načti základní sekce
        if not nactist:
            nactist = {"reporty", "faktury"}
        # Při výběru Vše vždy přidej banky
        if rok_raw in ("", "vše", "vse", "Vše", "all"):
            nactist.add("banky")
            nactist.add("reporty")

    try:
        import psycopg2 as _pg2
        db_url = os.environ.get("DATABASE_URL", "")
        conn = _pg2.connect(db_url)
        pg = conn.cursor()
        fw = "AND firma_zkratka=%s" if firma else ""
        fp = [firma] if firma else []
        def _q(sql, params=()):
            pg.execute(sql, params)
            cols = [d[0] for d in pg.description]
            return [dict(zip(cols, r)) for r in pg.fetchall()]
        if True:
            kontext_casti = [f"""Jsi analytik restaurace/bistra. Máš přístup k datům za {rok_label}{' pro firmu '+firma if firma else ''}.
Načtené sekce: {', '.join(sorted(nactist))}.
DŮLEŽITÉ: Pokud se tě ptají na konkrétní období (měsíc, rok), hledej v VŠECH dostupných datech níže — reporty, faktury, výdaje, bankovní výpisy.
Pokud data pro dané období nejsou v kontextu, řekni to jasně a navrhni co udělat (např. nahrát výpis).
Nikdy nevymýšlej data která nemáš."""]

            # Reporty
            if "reporty" in nactist and (ma_pravo("statistiky") or ma_pravo("reporty")):
                rep = _q(f"""
                    SELECT TO_CHAR(NULLIF(datum,'')::date,'YYYY-MM') as mesic,
                        ROUND(SUM(karty)::numeric,0) as karty,
                        ROUND(SUM(hotovost)::numeric,0) as hotovost,
                        ROUND(SUM(trzba_vcpk)::numeric,0) as trzba,
                        ROUND(SUM(vydaje)::numeric,0) as vydaje,
                        ROUND(SUM(pk_celkem)::numeric,0) as poukazky,
                        SUM(burger) as burger, SUM(burtgulas) as burtgulas,
                        SUM(pizza_cela) as pizza_cela, SUM(pizza_ctvrt) as pizza_ctvrt,
                        SUM(talire) as talire, COUNT(*) as dni
                    FROM reporty WHERE datum >= %s AND datum <= %s {fw}
                    GROUP BY mesic ORDER BY mesic
                """, [rok_od, rok_do] + fp)
                dny = _q(f"""
                    SELECT datum, firma_zkratka, karty, hotovost, trzba_vcpk as trzba,
                        vydaje, pk_celkem, burger, burtgulas, pizza_cela, pizza_ctvrt, talire, smena
                    FROM reporty
                    WHERE datum >= %s {fw}
                    ORDER BY datum DESC LIMIT 60
                """, [(_dt.date.today() - _dt.timedelta(days=60)).isoformat()] + fp)
                kontext_casti.append(f"\nMĚSÍČNÍ PŘEHLED REPORTŮ:\n{_safe_json(rep)}")
                kontext_casti.append(f"\nDENNÍ DATA (posledních 90 dní):\n{_safe_json(dny)}")

            # Faktury
            if "faktury" in nactist and ma_pravo("faktury"):
                fakt = _q(f"""
                    SELECT TO_CHAR(NULLIF(datum_vystaveni,'')::date,'YYYY-MM') as mesic,
                        dodavatel, ROUND(SUM(celkem_s_dph)::numeric,0) as castka, COUNT(*) as pocet
                    FROM faktury WHERE datum_vystaveni >= %s AND datum_vystaveni <= %s
                    {"AND firma_zkratka=%s" if firma else ""}
                    GROUP BY mesic, dodavatel ORDER BY mesic, castka DESC
                """, [rok_od, rok_do] + fp)
                kontext_casti.append(f"\nFAKTURY:\n{_safe_json(fakt)}")

            # Výplaty
            if "vyplaty" in nactist and ma_pravo("vyplaty"):
                vypl = _q(f"""
                    SELECT TO_CHAR(NULLIF(datum,'')::date,'YYYY-MM') as mesic,
                        jmeno, ROUND(SUM(castka)::numeric,0) as castka
                    FROM vyplaty WHERE datum >= %s AND datum <= %s
                    {"AND firma_zkratka=%s" if firma else ""}
                    GROUP BY mesic, jmeno ORDER BY mesic, jmeno
                """, [rok_od, rok_do] + fp)
                kontext_casti.append(f"\nVÝPLATY:\n{_safe_json(vypl)}")

            # Výdaje
            if "vydaje" in nactist:
                if ma_pravo("vydaje_zobrazit"):
                    vyd = _q(f"""
                        SELECT datum, dodavatel, castka, popis, stav
                        FROM vydaje WHERE typ='provozni'
                        AND datum >= %s AND datum <= %s {fw}
                        ORDER BY datum DESC LIMIT 200
                    """, [rok_od, rok_do] + fp)
                    kontext_casti.append(f"\nPROVOZNÍ VÝDAJE:\n{_safe_json(vyd)}")
                if je_admin:
                    svyd = _q("""
                        SELECT datum, dodavatel, castka, popis, stav
                        FROM vydaje WHERE typ='soukrome'
                        ORDER BY datum DESC LIMIT 100
                    """)
                    kontext_casti.append(f"\nSOUKROMÉ VÝDAJE:\n{_safe_json(svyd)}")

            # Peněženka
            if "penezenka" in nactist and je_admin:
                pw = _q(f"""
                    SELECT datum, hotovost, sporeni, poznamka
                    FROM penezenka ORDER BY datum DESC LIMIT 24
                """)
                kontext_casti.append(f"\nPENĚŽENKA:\n{_safe_json(pw)}")

            # Dokumenty
            if "dokumenty" in nactist and je_admin:
                dok = _q(f"""
                    SELECT datum, nazev, misto, kategorie
                    FROM dokumenty ORDER BY datum DESC
                """)
                kontext_casti.append(f"\nDOKUMENTY:\n{_safe_json(dok)}")

            # Vystavené faktury
            if "vystavene" in nactist and ma_pravo("faktury_zobrazit"):
                vf = _q(f"""
                    SELECT datum, odberatel, castka, stav, popis
                    FROM vystavene_faktury
                    WHERE datum >= %s AND datum <= %s
                    ORDER BY datum DESC
                """, [rok_od, rok_do])
                kontext_casti.append(f"\nVYSTAVENÉ FAKTURY:\n{_safe_json(vf)}")

            # Zboží — při výběru sekce zbozi načti vše, jinak top 100
            if "zbozi" in nactist and ma_pravo("zbozi_zobrazit"):
                limit_zbozi = "" if "zbozi" in sekce_override else "LIMIT 100"
                zbz = _q(f"""
                    SELECT z.nazev_canonical,
                        ROUND(SUM(p.celkem_s_dph)::numeric,0) as utraceno,
                        COUNT(DISTINCT p.faktura_id) as nakupu,
                        ROUND(SUM(p.mnozstvi)::numeric,2) as mnozstvi
                    FROM zbozi z
                    JOIN polozky p ON p.zbozi_id = z.id
                    GROUP BY z.nazev_canonical
                    ORDER BY utraceno DESC {limit_zbozi}
                """)
                kontext_casti.append(f"\nZBOŽÍ{'(vše)' if not limit_zbozi else '(top 100)'}:\n{_safe_json(zbz)}")

            # Bankovní výpisy
            if "banky" in nactist or je_admin:
                try:
                    import psycopg2 as _pg2b
                    conn2 = _pg2b.connect(db_url)
                    cur2 = conn2.cursor()
                    fw2 = "AND firma_zkratka=%s" if firma else ""
                    cur2.execute(f"""
                        SELECT datum, banka, firma_zkratka, ROUND(castka::numeric,2) as castka, nazev_protiucet, var_sym
                        FROM bankovni_pohyby
                        WHERE datum >= %s AND datum <= %s {fw2}
                        ORDER BY datum DESC LIMIT 800
                    """, [rok_od, rok_do] + ([firma] if firma else []))
                    cols2 = [d[0] for d in cur2.description]
                    banky_rows = [dict(zip(cols2, r)) for r in cur2.fetchall()]
                    conn2.close()
                    kontext_casti.append(f"\nBANKOVNÍ VÝPISY (celkem {len(banky_rows)} pohybů):\n{_safe_json(banky_rows)}")
                except Exception as be:
                    kontext_casti.append(f"\nBANKOVNÍ VÝPISY: chyba načtení ({be})")

        conn.close()
        kontext_casti.append("\nOdpovídej stručně a konkrétně v češtině.\nPokud uživatel žádá export dat (CSV, tabulka, seznam), vrať odpověď ve formátu:\nEXPORT_CSV:nazev_souboru.csv\ndatum,hodnota1,hodnota2\nřádek1...\n\nJinak odpovídej normálně jako text.")
        kontext = "\n".join(kontext_casti)

        # Sestavit messages — systémový kontext v první zprávě + správně střídající se role
        historie = data.get("historie", [])
        if not historie:
            messages = [{"role": "user", "content": kontext + "\n\nDotaz: " + dotaz}]
        else:
            # Kontext přidáme k první user zprávě v historii
            messages = []
            for i, h in enumerate(historie):
                if i == 0 and h["role"] == "user":
                    messages.append({"role": "user", "content": kontext + "\n\nDotaz: " + h["content"]})
                else:
                    messages.append({"role": h["role"], "content": h["content"]})
            messages.append({"role": "user", "content": dotaz})

        client = anthropic.Anthropic(api_key=api_key)
        # Ořez kontextu pokud je příliš dlouhý (max ~150k znaků = ~50k tokenů)
        if len(kontext) > 150000:
            kontext = kontext[:150000] + "\n\n[Data zkrácena kvůli délce — zobrazena část]"
        msg = client.messages.create(
            model="claude-sonnet-4-20250514",
            max_tokens=2000,
            messages=messages
        )
        odpoved = msg.content[0].text.strip()
        # Detekce CSV exportu
        export = None
        if odpoved.startswith("EXPORT_CSV:"):
            lines = odpoved.split("\n")
            fname = lines[0].replace("EXPORT_CSV:", "").strip()
            csv_data = "\n".join(lines[1:])
            export = {"nazev": fname, "data": csv_data}
            odpoved = f"Připravil jsem export: **{fname}**"
        return jsonify({"odpoved": odpoved, "export": export})
    except Exception as e:
        import traceback
        tb = traceback.format_exc()
        app.logger.error(f"api_ai_dotaz error: {tb}")
        return jsonify({"chyba": tb}), 500


@app.route("/api/export/reporty")
@vyzaduj_prihlaseni
def export_reporty():
    fmt = request.args.get("format", "xlsx")
    od  = request.args.get("od", "")
    do_ = request.args.get("do", "")
    clauses, params = [], []
    if od:  clauses.append("datum>=?"); params.append(od)
    if do_: clauses.append("datum<=?"); params.append(do_)
    where = ("WHERE " + " AND ".join(clauses)) if clauses else ""

    with get_db() as conn:
        rows = conn.execute(f"""
            SELECT datum,den,trzba_vcpk,karty,hotovost,vydaje,trzba,
                   pk50_ks,pk100_ks,pk_celkem,pizza_cela,pizza_ctvrt,
                   burger,talire,burtgulas,smena
            FROM reporty {where} ORDER BY datum
        """, params).fetchall()

    headers = ["datum","měsíc","den","TRŽBA vč. PK","karty","hotovost","výdaje","tržba",
               "pk50 ks","pk100 ks","poukaz Kč","pizza celá","čtvrt","burger","talíře","buřtguláš","KDO"]

    def r_val(r, key, idx):
        return r[key] if isinstance(r, dict) else r[idx]

    if fmt == "csv":
        buf = io.StringIO()
        w   = csv.writer(buf, delimiter=";")
        w.writerow(headers)
        for r in rows:
            d = date.fromisoformat(r_val(r,"datum",0)) if r_val(r,"datum",0) else None
            mesic = d.strftime("%B").upper() if d else ""
            w.writerow([d.day if d else "", mesic, r_val(r,"den",1), r_val(r,"trzba_vcpk",2),
                        r_val(r,"karty",3), r_val(r,"hotovost",4), r_val(r,"vydaje",5),
                        r_val(r,"trzba",6), r_val(r,"pk50_ks",7), r_val(r,"pk100_ks",8),
                        r_val(r,"pk_celkem",9), r_val(r,"pizza_cela",10), r_val(r,"pizza_ctvrt",11),
                        r_val(r,"burger",12), r_val(r,"talire",13), r_val(r,"burtgulas",14),
                        r_val(r,"smena",15)])
        buf.seek(0)
        return send_file(io.BytesIO(buf.getvalue().encode("utf-8-sig")),
                         mimetype="text/csv", download_name="reporty.csv", as_attachment=True)
    else:
        wb_out = openpyxl.Workbook()
        ws_out = wb_out.active; ws_out.title = str(date.today().year)
        _xlsx_header(ws_out, headers)
        mesice_cs = ["","LEDEN","ÚNOR","BŘEZEN","DUBEN","KVĚTEN","ČERVEN",
                     "ČERVENEC","SRPEN","ZÁŘÍ","ŘÍJEN","LISTOPAD","PROSINEC"]
        for r in rows:
            d = date.fromisoformat(r_val(r,"datum",0)) if r_val(r,"datum",0) else None
            mesic = mesice_cs[d.month] if d else ""
            ws_out.append([d.day if d else "", mesic, r_val(r,"den",1), r_val(r,"trzba_vcpk",2),
                           r_val(r,"karty",3), r_val(r,"hotovost",4), r_val(r,"vydaje",5),
                           r_val(r,"trzba",6), r_val(r,"pk50_ks",7), r_val(r,"pk100_ks",8),
                           r_val(r,"pk_celkem",9), r_val(r,"pizza_cela",10), r_val(r,"pizza_ctvrt",11),
                           r_val(r,"burger",12), r_val(r,"talire",13), r_val(r,"burtgulas",14),
                           r_val(r,"smena",15)])
        buf = io.BytesIO(); wb_out.save(buf); buf.seek(0)
        return send_file(buf, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                         download_name="reporty.xlsx", as_attachment=True)


# ── API: nahrání souboru ──────────────────────────────────────────────────────
@app.route("/api/nahrat-text", methods=["POST"])
@vyzaduj_prihlaseni
def api_nahrat_text():
    text = request.json.get("text", "")
    if not text.strip():
        return jsonify({"error": "Prázdný text"}), 400
    data = _parse_makro_text(text)
    return jsonify(data)


def _parse_makro_text(text):
    lines = text.splitlines()
    result = {
        "cislo_faktury":   "",
        "datum_vystaveni": "",
        "datum_splatnosti":"",
        "zpusob_uhrady":   "Hotovost",
        "stav":            "zaplaceno",
        "dodavatel":       "MAKRO Cash & Carry ČR s.r.o.",
        "celkem_s_dph":    0,
        "polozky":         []
    }

    items = []
    sleva_kw = ["urceno pro konecnou", "kup vice", "kup více", "věrnostní"]

    for line in lines:
        ls = line.strip()
        if not ls: continue
        ll = ls.lower()

        m = re.search(r"Faktura\s*[čc\.]\s*/\s*VS\s*:\s*(\S+)", ls, re.IGNORECASE)
        if m and not result["cislo_faktury"]: result["cislo_faktury"] = m.group(1)
        m = re.search(r"Datum\s+vystavení\s*:\s*(\d{2}[-\.]\d{2}[-\.]\d{4})", ls, re.IGNORECASE)
        if m and not result["datum_vystaveni"]: result["datum_vystaveni"] = _makro_date(m.group(1).replace(".", "-") if "." in m.group(1) else m.group(1))
        m = re.search(r"Datum\s+splatnosti\s*:\s*(\d{2}[-\.]\d{2}[-\.]\d{4})", ls, re.IGNORECASE)
        if m and not result["datum_splatnosti"]: result["datum_splatnosti"] = _makro_date(m.group(1).replace(".", "-") if "." in m.group(1) else m.group(1))
        m = re.search(r"Celková\s+částka\s+([\d\s]{1,10}[,\.]\d{2})", ls, re.IGNORECASE)
        if m: result["celkem_s_dph"] = _parse_money(m.group(1))

        is_sleva = any(kw in ll for kw in sleva_kw)
        if is_sleva and items:
            neg = re.findall(r"-\s*(\d[\d\s]*[,\.]\d{2})", ls)
            if neg:
                sleva = _parse_money(neg[-1])
                items[-1]["celkem_s_dph"] = round(max(0, items[-1]["celkem_s_dph"] - sleva), 2)
                mn = items[-1]["mnozstvi"]
                if mn: items[-1]["cena_za_jednotku_s_dph"] = round(items[-1]["celkem_s_dph"] / mn, 4)
            continue

        mm = re.match(r"^(\d{6,14})\s+\*?(.+?)\s+(PC|KG|BG|KS|BX|CA|SW)\s+(.+)$", ls, re.IGNORECASE)
        if not mm: continue

        nazev    = mm.group(2).strip().rstrip("*")
        jednotka = mm.group(3).upper()
        rest     = mm.group(4)

        cisla = re.findall(r"\d+[,\.]\d+|\d+", rest)
        cf = [_parse_money(c) for c in cisla]
        cf = [c for c in cf if c > 0]

        if len(cf) < 2: continue

        if cf[-1] == int(cf[-1]) and cf[-1] <= 25:
            idx_dph = len(cf) - 1
        else:
            idx_dph = len(cf)

        idx_cs  = idx_dph - 1
        idx_mn  = idx_dph - 3
        celkem  = cf[idx_cs] if 0 <= idx_cs < len(cf) else 0
        pocet   = cf[idx_mn] if 0 <= idx_mn < len(cf) else 1.0
        if pocet <= 0 or pocet > 10000: pocet = 1.0
        cena_j  = round(celkem / pocet, 4) if pocet else celkem

        if not nazev or celkem <= 0: continue
        items.append({
            "nazev":                  _format_nazev(nazev),
            "mnozstvi":               pocet,
            "jednotka":               _map_unit(jednotka),
            "cena_za_jednotku_s_dph": cena_j,
            "celkem_s_dph":           round(celkem, 2)
        })

    result["polozky"] = items
    if result["celkem_s_dph"] == 0 and items:
        result["celkem_s_dph"] = round(sum(p["celkem_s_dph"] for p in items), 2)
    return result


@app.route("/api/nahrat", methods=["POST"])
@vyzaduj_prihlaseni
def api_nahrat():
    if "soubor" not in request.files:
        return jsonify({"error": "Žádný soubor"}), 400
    f = request.files["soubor"]
    if not f.filename or not allowed_file(f.filename):
        return jsonify({"error": "Nepodporovaný formát"}), 400

    fname  = secure_filename(f.filename)
    ts     = datetime.now().strftime("%Y%m%d_%H%M%S_")
    fname  = ts + fname
    fpath  = os.path.join(UPLOAD_DIR, fname)
    f.save(fpath)

    gcs_url = upload_to_gcs(fpath, fname)

    ext = fname.rsplit(".", 1)[1].lower()
    api_key = os.environ.get("ANTHROPIC_API_KEY", "")
    if api_key:
        if ext == "pdf":
            data, err = parse_makro_pdf(fpath)
            if err or not data:
                data, err = parse_faktura_claude(fpath)
        else:
            data, err = parse_faktura_claude(fpath)
    else:
        if ext == "pdf":
            data, err = parse_makro_pdf(fpath)
        else:
            data, err = parse_makro_image(fpath)

    if err:
        return jsonify({"error": err, "soubor_cesta": fname}), 200

    data["soubor_cesta"] = fname
    if gcs_url:
        data["soubor_gcs_url"] = gcs_url

    if data.get("cislo_faktury"):
        with get_db() as conn:
            row = conn.execute("""
                SELECT id, firma_zkratka, datum_vystaveni, celkem_s_dph
                FROM faktury
                WHERE cislo_faktury = ?
                AND datum_vystaveni = ?
                AND ABS(celkem_s_dph - ?) < 1.0
            """, (data["cislo_faktury"], data.get("datum_vystaveni",""), float(data.get("celkem_s_dph", 0)))).fetchone()
            if row:
                data["duplicita"] = {
                    "id": row["id"],
                    "firma": row["firma_zkratka"],
                    "datum": row["datum_vystaveni"],
                    "celkem": row["celkem_s_dph"]
                }

    # Obsahová duplicita (stejné datum + částka + položky, ignoruje číslo faktury)
    if not data.get("duplicita") and data.get("datum_vystaveni") and data.get("polozky"):
        with get_db() as conn:
            obs_dup = _najdi_obsahovou_duplicitu(
                conn, data.get("datum_vystaveni",""),
                float(data.get("celkem_s_dph", 0)),
                data.get("polozky", [])
            )
            if obs_dup:
                data["duplicita"] = obs_dup
                data["duplicita"]["typ"] = "obsahova"

    return jsonify(data)

@app.route("/api/faktury", methods=["POST"])
@vyzaduj_prihlaseni
def api_faktura_ulozit():
    data = request.json
    required = ["firma_zkratka", "dodavatel"]
    for r in required:
        if not data.get(r):
            return jsonify({"error": f"Chybí pole: {r}"}), 400

    polozky = data.pop("polozky", [])

    # MAKRO faktury jsou vždy zaplaceny okamžitě při nákupu
    dodavatel = data.get("dodavatel", "")
    if "MAKRO" in dodavatel.upper():
        data["stav"] = "zaplaceno"
        if not data.get("datum_zaplaceno"):
            data["datum_zaplaceno"] = data.get("datum_vystaveni", "")

    # Detekce obsahové duplicity před uložením
    obsahova_duplicita_id = None
    with get_db() as conn:
        obs_dup = _najdi_obsahovou_duplicitu(
            conn, data.get("datum_vystaveni",""),
            float(data.get("celkem_s_dph", 0)),
            polozky
        )
        if obs_dup:
            obsahova_duplicita_id = obs_dup["id"]

    with get_db() as conn:
        cur = conn.execute("""
            INSERT INTO faktury (firma_zkratka, dodavatel, cislo_faktury, datum_vystaveni,
                datum_splatnosti, zpusob_uhrady, stav, celkem_s_dph, soubor_cesta, soubor_url, zdroj, duplicita_id)
            VALUES (?,?,?,?,?,?,?,?,?,?,?,?)
        """, (
            data.get("firma_zkratka"),
            data.get("dodavatel"),
            data.get("cislo_faktury",""),
            data.get("datum_vystaveni",""),
            data.get("datum_splatnosti",""),
            data.get("zpusob_uhrady",""),
            data.get("stav","ceka"),
            data.get("celkem_s_dph", 0),
            data.get("soubor_cesta",""),
            data.get("soubor_url",""),
            data.get("zdroj","rucni"),
            obsahova_duplicita_id or data.get("duplicita_id", None)
        ))
        faktura_id = cur.lastrowid

        for p in polozky:
            nazev = p.get("nazev","").strip()
            if not nazev: continue
            mnozstvi = float(p.get("mnozstvi", 1) or 1)
            celkem   = float(p.get("celkem_s_dph", 0) or 0)
            cena_j   = float(p.get("cena_za_jednotku_s_dph", 0) or 0)
            if cena_j == 0 and mnozstvi:
                cena_j = celkem / mnozstvi
            jed = p.get("jednotka","ks")
            zbozi_id = _get_or_create_zbozi(conn, nazev)
            conn.execute("""
                INSERT INTO polozky (faktura_id, nazev, mnozstvi, jednotka,
                    cena_za_jednotku_s_dph, celkem_s_dph, zbozi_id)
                VALUES (?,?,?,?,?,?,?)
            """, (faktura_id, nazev, mnozstvi, jed, round(cena_j,4), round(celkem,2), zbozi_id))

        recalc_faktura_total(conn, faktura_id)

    return jsonify({"ok": True, "id": faktura_id})


def _najdi_obsahovou_duplicitu(conn, datum, celkem_s_dph, polozky, vynechat_id=None):
    """Hledá fakturu se stejným datem, celkovou částkou a stejnými položkami (bez ohledu na číslo faktury).
    Vrátí dict s info o duplicitě nebo None."""
    import re as _re
    if not datum or not polozky:
        return None

    def _norm_nazev(n):
        n = str(n).strip().upper()
        n = _re.sub(r'^(ARO|MC|FL|CBA)\s+', '', n)  # odstranit prefix
        n = _re.sub(r'\s+(KG|G|L|ML|KS|PC|BG|SW|CA)$', '', n)  # odstranit příponu
        return n

    kandidati = conn.execute("""
        SELECT id, firma_zkratka, datum_vystaveni, celkem_s_dph, cislo_faktury
        FROM faktury
        WHERE datum_vystaveni = ? AND ABS(celkem_s_dph - ?) < 1.0
    """, (datum, float(celkem_s_dph))).fetchall()

    for k in kandidati:
        kid = k["id"] if isinstance(k, dict) else k[0]
        if vynechat_id and kid == vynechat_id:
            continue
        kpol = conn.execute(
            "SELECT nazev, mnozstvi FROM polozky WHERE faktura_id = ? ORDER BY nazev", (kid,)
        ).fetchall()
        if len(kpol) != len(polozky):
            continue
        nove = sorted([(_norm_nazev(p.get("nazev","")), float(p.get("mnozstvi",1) or 1)) for p in polozky])
        existujici = sorted([
            (_norm_nazev(p["nazev"] if isinstance(p, dict) else p[0]),
             float(p["mnozstvi"] if isinstance(p, dict) else p[1]))
            for p in kpol
        ])
        if nove == existujici:
            return {
                "id": kid,
                "firma": k["firma_zkratka"] if isinstance(k, dict) else k[1],
                "datum": k["datum_vystaveni"] if isinstance(k, dict) else k[2],
                "celkem": k["celkem_s_dph"] if isinstance(k, dict) else k[3],
                "cislo_faktury": k["cislo_faktury"] if isinstance(k, dict) else k[4],
            }
    return None


def _get_or_create_zbozi(conn, nazev):
    row = conn.execute("SELECT zbozi_id FROM zbozi_aliasy WHERE alias=?", (nazev,)).fetchone()
    if row: return row["zbozi_id"]
    row = conn.execute("SELECT id FROM zbozi WHERE nazev_canonical=?", (nazev,)).fetchone()
    if row: return row["id"]
    cur = conn.execute("INSERT INTO zbozi (nazev_canonical) VALUES (?)", (nazev,))
    return cur.lastrowid

@app.route("/api/polozky")
@vyzaduj_prihlaseni
def api_polozky():
    firma = request.args.get("firma", "")
    od    = request.args.get("od", "")
    do_   = request.args.get("do", "")

    f_cond  = "AND fakt.firma_zkratka=%s" if firma else ""
    od_c    = "AND fakt.datum_vystaveni>=%s" if od else ""
    do_c    = "AND fakt.datum_vystaveni<=%s" if do_ else ""
    params = tuple(v for v in [firma, od, do_] if v)

    with get_db() as conn:
        rows = conn.execute(f"""
            SELECT
                COALESCE(al.alias, z.nazev_canonical, p.nazev) AS zbozi_nazev,
                MIN(z.id) AS zbozi_id,
                ROUND(CAST(SUM(p.mnozstvi) AS NUMERIC), 3)               AS celkove_mnozstvi,
                ROUND(CAST(SUM(p.celkem_s_dph) AS NUMERIC), 2)           AS celkem_utraceno,
                ROUND(CAST(AVG(p.cena_za_jednotku_s_dph) AS NUMERIC), 4) AS prumerna_cena,
                COUNT(DISTINCT p.faktura_id)                              AS pocet_nakupu,
                STRING_AGG(DISTINCT
                    REPLACE(REPLACE(fakt.dodavatel, ' CR s', ' ČR s'), ' cr s', ' ČR s'),
                ', ')                                                     AS dodavatele,
                MIN(p.jednotka)                                           AS jednotka,
                al.alias                                                  AS skupina
            FROM polozky p
            JOIN faktury fakt ON fakt.id = p.faktura_id
            LEFT JOIN zbozi z ON z.id = p.zbozi_id
            LEFT JOIN zbozi_aliasy al ON al.zbozi_id = z.id
            WHERE 1=1 {f_cond} {od_c} {do_c}
            GROUP BY COALESCE(al.alias, z.nazev_canonical, p.nazev), al.alias
            ORDER BY celkem_utraceno DESC
        """, params).fetchall()
    return jsonify([dict(r) for r in rows])
@app.route("/api/zbozi/alias-detail/<path:alias>")
@vyzaduj_prihlaseni
def api_alias_detail(alias):
    with get_db() as conn:
        nakupy = conn.execute("""
            SELECT DISTINCT ON (p.id) p.*, f.dodavatel, f.datum_vystaveni, f.firma_zkratka, f.id as faktura_id,
                   f.soubor_url, f.cislo_faktury, z.nazev_canonical, z.id as zbozi_id_orig
            FROM polozky p
            JOIN faktury f ON f.id = p.faktura_id
            JOIN zbozi z ON z.id = p.zbozi_id
            JOIN zbozi_aliasy al ON al.zbozi_id = z.id
            WHERE LOWER(al.alias) = LOWER(%s)
            ORDER BY p.id, f.datum_vystaveni DESC
        """, (alias,)).fetchall()
    return jsonify({
        "alias": alias,
        "nakupy": [dict(r) for r in nakupy]
    })
@app.route("/api/polozky/detail/<int:zbozi_id>")
@vyzaduj_prihlaseni
def api_zbozi_detail(zbozi_id):
    with get_db() as conn:
        zbozi = conn.execute("SELECT * FROM zbozi WHERE id=?", (zbozi_id,)).fetchone()
        if not zbozi:
            return jsonify({"error": "Nenalezeno"}), 404
        aliasy = conn.execute("SELECT alias FROM zbozi_aliasy WHERE zbozi_id=?", (zbozi_id,)).fetchall()
        nakupy = conn.execute("""
            SELECT p.*, f.dodavatel, f.datum_vystaveni, f.firma_zkratka, f.id as faktura_id,
                   f.soubor_url, f.cislo_faktury
            FROM polozky p
            JOIN faktury f ON f.id = p.faktura_id
            WHERE p.zbozi_id=?
            ORDER BY f.datum_vystaveni DESC
        """, (zbozi_id,)).fetchall()
    return jsonify({
        "zbozi": dict(zbozi),
        "aliasy": [r["alias"] for r in aliasy],
        "nakupy": [dict(r) for r in nakupy]
    })

@app.route("/api/zbozi-search")
@vyzaduj_prihlaseni
def api_zbozi_search():
    import unicodedata
    q = request.args.get("q", "").strip()
    unaccent = request.args.get("unaccent", "0") == "1"
    if not q:
        return jsonify([])

    def _strip(s):
        return ''.join(c for c in unicodedata.normalize('NFD', s) if unicodedata.category(c) != 'Mn').lower()

    with get_db() as conn:
        # Zkusit nejdřív přímý LIKE
        rows = conn.execute("""
            SELECT DISTINCT nazev_canonical
            FROM zbozi
            WHERE LOWER(nazev_canonical) LIKE LOWER(?)
            ORDER BY nazev_canonical
            LIMIT 20
        """, (f"%{q}%",)).fetchall()
        result = [r["nazev_canonical"] if isinstance(r, dict) else r[0] for r in rows]

        # Pokud unaccent=1 a nenašli jsme dost, doplníme Python filtrací
        if unaccent and len(result) < 10:
            q_stripped = _strip(q)
            all_rows = conn.execute("SELECT DISTINCT nazev_canonical FROM zbozi ORDER BY nazev_canonical").fetchall()
            for r in all_rows:
                n = r["nazev_canonical"] if isinstance(r, dict) else r[0]
                if n not in result and q_stripped in _strip(n):
                    result.append(n)
                if len(result) >= 10:
                    break

    return jsonify([{"nazev_canonical": n} for n in result[:10]])


@app.route("/api/zbozi")
@vyzaduj_prihlaseni
def api_zbozi_list():
    with get_db() as conn:
        rows = conn.execute("SELECT id, nazev_canonical FROM zbozi ORDER BY nazev_canonical").fetchall()
    return jsonify([dict(r) for r in rows])

@app.route("/api/zbozi/alias/<int:zbozi_id>/<path:alias>", methods=["DELETE"])
@vyzaduj_prihlaseni
def api_zbozi_alias_delete(zbozi_id, alias):
    with get_db() as conn:
        conn.execute("DELETE FROM zbozi_aliasy WHERE zbozi_id=%s AND alias=%s", (zbozi_id, alias))
    return jsonify({"ok": True})

@app.route("/api/zbozi/aliasy-seznam")
def api_zbozi_aliasy_seznam():
    q = request.args.get("q", "").strip().lower()
    with get_db() as conn:
        rows = conn.execute(
            "SELECT DISTINCT alias FROM zbozi_aliasy WHERE LOWER(alias) LIKE %s ORDER BY alias LIMIT 10",
            (f"%{q}%",)
        ).fetchall()
    return jsonify([r["alias"] if isinstance(r, dict) else r[0] for r in rows])

@app.route("/api/zbozi/alias", methods=["POST"])
@vyzaduj_prihlaseni
def api_zbozi_alias():
    data = request.json
    zbozi_id   = data.get("zbozi_id")
    alias_text = data.get("alias", "").strip()
    polozka_id = data.get("polozka_id")
    if not zbozi_id or not alias_text:
        return jsonify({"error": "Chybí zbozi_id nebo alias"}), 400
    with get_db() as conn:
        # Smaž starý alias pro toto zboží
        conn.execute("DELETE FROM zbozi_aliasy WHERE zbozi_id=%s", (zbozi_id,))
        # Vlož nový alias (bez ON CONFLICT — více zboží může mít stejný alias)
        conn.execute(
            "INSERT INTO zbozi_aliasy (zbozi_id, alias) VALUES (%s,%s)",
            (zbozi_id, alias_text)
        )
    return jsonify({"ok": True})

@app.route("/api/zbozi", methods=["POST"])
@vyzaduj_prihlaseni
def api_zbozi_create():
    nazev = request.json.get("nazev_canonical", "").strip()
    if not nazev:
        return jsonify({"error": "Chybí název"}), 400
    with get_db() as conn:
        try:
            cur = conn.execute("INSERT INTO zbozi (nazev_canonical) VALUES (?)", (nazev,))
            return jsonify({"ok": True, "id": cur.lastrowid})
        except Exception:
            row = conn.execute("SELECT id FROM zbozi WHERE nazev_canonical=?", (nazev,)).fetchone()
            return jsonify({"ok": True, "id": row["id"]})

@app.route("/api/statistiky")
@vyzaduj_prihlaseni
def api_statistiky():
    firma = request.args.get("firma", "")
    od    = request.args.get("od", date.today().replace(day=1).isoformat())
    do_   = request.args.get("do", date.today().isoformat())

    f_cond  = "AND firma_zkratka=?" if firma else ""
    f_params = (firma,) if firma else ()

    with get_db() as conn:
        mesice = conn.execute(f"""
            SELECT strftime('%Y-%m', datum_vystaveni) m, ROUND((SUM(celkem_s_dph))::numeric,2) castka
            FROM faktury
            WHERE datum_vystaveni>=? AND datum_vystaveni<=? {f_cond}
            GROUP BY m ORDER BY m
        """, (od, do_) + f_params).fetchall()

        dodavatele = conn.execute(f"""
            SELECT dodavatel, ROUND((SUM(celkem_s_dph))::numeric,2) castka, COUNT(*) pocet
            FROM faktury
            WHERE datum_vystaveni>=? AND datum_vystaveni<=? {f_cond}
            GROUP BY dodavatel ORDER BY castka DESC LIMIT 10
        """, (od, do_) + f_params).fetchall()

        zbozi_top = conn.execute(f"""
            SELECT COALESCE(z.nazev_canonical, p.nazev) zbozi, ROUND((SUM(p.celkem_s_dph))::numeric,2) castka,
                   ROUND((SUM(p.mnozstvi))::numeric,2) mnozstvi, MAX(p.jednotka) jednotka
            FROM polozky p
            JOIN faktury f ON f.id=p.faktura_id
            LEFT JOIN zbozi z ON z.id=p.zbozi_id
            WHERE f.datum_vystaveni>=? AND f.datum_vystaveni<=? {f_cond}
            GROUP BY COALESCE(z.nazev_canonical, p.nazev) ORDER BY castka DESC LIMIT 20
        """, (od, do_) + f_params).fetchall()

        zbozi_id = request.args.get("zbozi_id")
        cena_vyvoj = []
        if zbozi_id:
            cena_vyvoj = conn.execute(f"""
                SELECT f.datum_vystaveni dat, ROUND(p.cena_za_jednotku_s_dph,4) cena, f.dodavatel
                FROM polozky p JOIN faktury f ON f.id=p.faktura_id
                WHERE p.zbozi_id=? AND f.datum_vystaveni>=? AND f.datum_vystaveni<=? {f_cond}
                ORDER BY f.datum_vystaveni
            """, (zbozi_id, od, do_) + f_params).fetchall()

    return jsonify({
        "mesice": [dict(r) for r in mesice],
        "dodavatele": [dict(r) for r in dodavatele],
        "zbozi_top": [dict(r) for r in zbozi_top],
        "cena_vyvoj": [dict(r) for r in cena_vyvoj]
    })

@app.route("/api/export/faktury")
@vyzaduj_prihlaseni
def export_faktury():
    fmt   = request.args.get("format", "xlsx")
    firma = request.args.get("firma", "")
    stav  = request.args.get("stav", "")
    od    = request.args.get("od", "")
    do_   = request.args.get("do", "")

    clauses, params = [], []
    if firma: clauses.append("firma_zkratka=?"); params.append(firma)
    if stav:  clauses.append("stav=?"); params.append(stav)
    if od:    clauses.append("datum_vystaveni>=?"); params.append(od)
    if do_:   clauses.append("datum_vystaveni<=?"); params.append(do_)
    where = ("WHERE " + " AND ".join(clauses)) if clauses else ""

    with get_db() as conn:
        rows = conn.execute(f"""
            SELECT firma_zkratka, dodavatel, cislo_faktury, datum_vystaveni,
                   datum_splatnosti, zpusob_uhrady, stav, celkem_s_dph
            FROM faktury {where} ORDER BY datum_vystaveni DESC
        """, params).fetchall()

    headers = ["Firma", "Dodavatel", "Číslo faktury", "Datum vystavení",
               "Datum splatnosti", "Způsob úhrady", "Stav", "Celkem s DPH"]

    if fmt == "csv":
        buf = io.StringIO()
        w   = csv.writer(buf, delimiter=";")
        w.writerow(headers)
        for r in rows: w.writerow(list(r))
        buf.seek(0)
        return send_file(io.BytesIO(buf.getvalue().encode("utf-8-sig")),
                         mimetype="text/csv", download_name="faktury.csv", as_attachment=True)
    else:
        wb_out = openpyxl.Workbook()
        ws_out = wb_out.active; ws_out.title = "Faktury"
        _xlsx_header(ws_out, headers)
        for r in rows: ws_out.append(list(r))
        buf = io.BytesIO(); wb_out.save(buf); buf.seek(0)
        return send_file(buf, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                         download_name="faktury.xlsx", as_attachment=True)
@app.route("/api/export/vydaje")
@vyzaduj_prihlaseni
def export_vydaje():
    fmt   = request.args.get("format", "xlsx")
    firma = request.args.get("firma", "")
    od    = request.args.get("od", "")
    do_   = request.args.get("do", "")
    stav  = request.args.get("stav", "")
    clauses, params = [], []
    if firma: clauses.append("firma_zkratka=?"); params.append(firma)
    if od:    clauses.append("datum>=?"); params.append(od)
    if do_:   clauses.append("datum<=?"); params.append(do_)
    if stav:  clauses.append("stav=?"); params.append(stav)
    where = ("WHERE " + " AND ".join(clauses)) if clauses else ""
    with get_db() as conn:
        rows = conn.execute(f"""
            SELECT firma_zkratka, datum, dodavatel, popis, castka, zpusob_uhrady, stav
            FROM vydaje {where} ORDER BY datum DESC
        """, params).fetchall()
    headers = ["Firma", "Datum", "Dodavatel", "Popis/účel", "Částka", "Způsob úhrady", "Stav"]
    if fmt == "csv":
        buf = io.StringIO()
        w = csv.writer(buf, delimiter=";")
        w.writerow(headers)
        for r in rows: w.writerow(list(r))
        buf.seek(0)
        return send_file(io.BytesIO(buf.getvalue().encode("utf-8-sig")),
                         mimetype="text/csv", download_name="vydaje.csv", as_attachment=True)
    else:
        wb_out = openpyxl.Workbook()
        ws_out = wb_out.active; ws_out.title = "Výdaje"
        _xlsx_header(ws_out, headers)
        for r in rows: ws_out.append(list(r))
        buf = io.BytesIO(); wb_out.save(buf); buf.seek(0)
        return send_file(buf, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                         download_name="vydaje.xlsx", as_attachment=True)

@app.route("/api/export/polozky")
@vyzaduj_prihlaseni
def export_polozky():
    fmt   = request.args.get("format", "xlsx")
    firma = request.args.get("firma", "")
    od    = request.args.get("od", "")
    do_   = request.args.get("do", "")

    f_cond = "AND f.firma_zkratka=?" if firma else ""
    od_c   = "AND f.datum_vystaveni>=?" if od else ""
    do_c   = "AND f.datum_vystaveni<=?" if do_ else ""
    params = tuple(v for v in [firma, od, do_] if v)

    with get_db() as conn:
        rows = conn.execute(f"""
            SELECT COALESCE(z.nazev_canonical, p.nazev), p.jednotka,
                   ROUND((SUM(p.mnozstvi))::numeric,3), ROUND((SUM(p.celkem_s_dph))::numeric,2),
                   ROUND((AVG(p.cena_za_jednotku_s_dph))::numeric,4),
                   COUNT(DISTINCT p.faktura_id),
                   STRING_AGG(DISTINCT f.dodavatel, ', ')
            FROM polozky p JOIN faktury f ON f.id=p.faktura_id
            LEFT JOIN zbozi z ON z.id=p.zbozi_id
            WHERE 1=1 {f_cond} {od_c} {do_c}
            GROUP BY COALESCE(z.id::text, p.nazev)
            ORDER BY SUM(p.celkem_s_dph) DESC
        """, params).fetchall()

    headers = ["Zboží", "Jednotka", "Celkové množství", "Celkem s DPH",
               "Průměrná cena/jedn.", "Počet nákupů", "Dodavatelé"]

    if fmt == "csv":
        buf = io.StringIO()
        w   = csv.writer(buf, delimiter=";")
        w.writerow(headers)
        for r in rows: w.writerow(list(r))
        buf.seek(0)
        return send_file(io.BytesIO(buf.getvalue().encode("utf-8-sig")),
                         mimetype="text/csv", download_name="polozky.csv", as_attachment=True)
    else:
        wb_out = openpyxl.Workbook()
        ws_out = wb_out.active; ws_out.title = "Položky"
        _xlsx_header(ws_out, headers)
        for r in rows: ws_out.append(list(r))
        buf = io.BytesIO(); wb_out.save(buf); buf.seek(0)
        return send_file(buf, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                         download_name="polozky.xlsx", as_attachment=True)

@app.route("/api/export/vyplaty")
@vyzaduj_prihlaseni
def export_vyplaty():
    fmt   = request.args.get("format", "xlsx")
    firma = request.args.get("firma", "")
    od    = request.args.get("od", "")
    do_   = request.args.get("do", "")
    clauses, params = [], []
    if firma: clauses.append("firma_zkratka=?"); params.append(firma)
    if od:    clauses.append("datum>=?"); params.append(od)
    if do_:   clauses.append("datum<=?"); params.append(do_)
    where = ("WHERE " + " AND ".join(clauses)) if clauses else ""
    with get_db() as conn:
        rows = conn.execute(f"SELECT firma_zkratka, jmeno, datum, castka, poznamka FROM vyplaty {where} ORDER BY datum DESC", params).fetchall()
    headers = ["Firma", "Jméno", "Datum", "Částka", "Poznámka"]
    if fmt == "csv":
        buf = io.StringIO()
        w = csv.writer(buf, delimiter=";")
        w.writerow(headers)
        for r in rows: w.writerow(list(r))
        buf.seek(0)
        return send_file(io.BytesIO(buf.getvalue().encode("utf-8-sig")), mimetype="text/csv", download_name="vyplaty.csv", as_attachment=True)
    else:
        wb_out = openpyxl.Workbook()
        ws_out = wb_out.active; ws_out.title = "Výplaty"
        _xlsx_header(ws_out, headers)
        for r in rows: ws_out.append(list(r))
        buf = io.BytesIO(); wb_out.save(buf); buf.seek(0)
        return send_file(buf, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", download_name="vyplaty.xlsx", as_attachment=True)

def _xlsx_header(ws, headers):
    green = "2D6A4F"
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor=green)
        cell.alignment = Alignment(horizontal="center")

@app.route("/uploads/<path:filename>")
def uploaded_file(filename):
    return send_from_directory(UPLOAD_DIR, filename)


init_db()
migrate_db()


@app.route("/api/drive-config")
@vyzaduj_prihlaseni
def api_drive_config():
    client_id = os.environ.get("GOOGLE_CLIENT_ID", "")
    return jsonify({"client_id": client_id})

@app.route("/api/drive-download", methods=["POST"])
@vyzaduj_prihlaseni
def api_drive_download():
    import requests as _req
    d = request.json or {}
    file_id    = d.get("file_id", "")
    access_token = d.get("access_token", "")
    filename   = d.get("filename", "dokument.pdf")
    if not file_id or not access_token:
        return jsonify({"error": "Chybí file_id nebo access_token"}), 400
    url = f"https://www.googleapis.com/drive/v3/files/{file_id}?alt=media"
    headers = {"Authorization": f"Bearer {access_token}"}
    resp = _req.get(url, headers=headers, timeout=30)
    if resp.status_code != 200:
        return jsonify({"error": f"Chyba stahování z Drive: {resp.status_code}"}), 400
    import tempfile, os as _os
    suffix = ".pdf" if filename.lower().endswith(".pdf") else ""
    with tempfile.NamedTemporaryFile(delete=False, suffix=suffix) as tmp:
        tmp.write(resp.content)
        tmp_path = tmp.name
    try:
        from werkzeug.datastructures import FileStorage
        import io
        fs = FileStorage(
            stream=io.BytesIO(resp.content),
            filename=filename,
            content_type="application/pdf"
        )
        safe = filename.replace(" ", "_")
        dest = os.path.join(UPLOAD_DIR, safe)
        fs.save(dest)
        gcs_url = upload_to_gcs(dest, safe)
        return jsonify({"ok": True, "tmp_path": dest, "soubor_url": gcs_url or "", "filename": safe})
    except Exception as e:
        return jsonify({"error": str(e)}), 500
    finally:
        try: _os.unlink(tmp_path)
        except: pass


# ── GOOGLE DRIVE WEBHOOK ──────────────────────────────────────────────────────
DRIVE_FOLDER_ID = "1Oopnqi_IDwqWOKb--u9gGQ3ds1RwhjKh"
DRIVE_CHANNEL_ID = "faktury-makro-channel-1"

def get_drive_service():
    """Vrátí Google Drive service a credentials pomocí service account."""
    creds_json = os.environ.get("GCS_CREDENTIALS_JSON", "")
    if not creds_json:
        return None, None
    try:
        from google.oauth2 import service_account
        from googleapiclient.discovery import build
        import google.auth.transport.requests
        creds_info = json.loads(creds_json)
        scopes = ["https://www.googleapis.com/auth/drive"]
        creds = service_account.Credentials.from_service_account_info(creds_info, scopes=scopes)
        creds.refresh(google.auth.transport.requests.Request())
        service = build("drive", "v3", credentials=creds)
        return service, creds
    except Exception as e:
        print(f"⚠ Drive service error: {e}")
        return None, None

@app.route("/api/drive-registruj", methods=["POST"])
@vyzaduj_prihlaseni
def api_drive_registruj():
    if session.get("role") != "admin":
        return jsonify({"error": "Pouze admin"}), 403
    import uuid
    try:
        from googleapiclient.discovery import build
        service, creds = get_drive_service()
        if not service:
            return jsonify({"error": "Drive service není dostupný"}), 500
        webhook_url = f"{os.environ.get('APP_URL', 'https://faktury-makro-git-904528626460.europe-west1.run.app')}/api/drive-webhook"
        channel_id = str(uuid.uuid4())
        body = {
            "id": channel_id,
            "type": "web_hook",
            "address": webhook_url,
            "expiration": str(int((__import__("time").time() + 604800) * 1000))
        }
        result = service.files().watch(
            fileId=DRIVE_FOLDER_ID,
            body=body
        ).execute()
        with get_db() as conn:
            try:
                conn.execute("""CREATE TABLE IF NOT EXISTS drive_channels (
                    id SERIAL PRIMARY KEY, channel_id TEXT, resource_id TEXT, expiration TEXT)""")
            except: pass
            conn.execute("INSERT INTO drive_channels (channel_id, resource_id, expiration) VALUES (?,?,?)",
                (channel_id, result.get("resourceId",""), result.get("expiration","")))
        return jsonify({"ok": True, "channel_id": channel_id, "expiration": result.get("expiration")})
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route("/api/drive-zkontrolovat", methods=["POST"])
@vyzaduj_prihlaseni
def api_drive_zkontrolovat():
    """Ručně spustí stažení nových souborů z Google Drive."""
    print("DRIVE_ZKONTROLOVAT_SPUSTENO")
    stats = _zpracuj_nove_faktury_z_drive()
    return jsonify({"ok": True, "stazeno": stats.get("stazeno", 0), "preskoceno": stats.get("preskoceno", 0), "chyby": stats.get("chyby", 0)})

@app.route("/api/drive-webhook", methods=["POST"])
def api_drive_webhook():
    """Příjem notifikací od Google Drive."""
    resource_state = request.headers.get("X-Goog-Resource-State", "")
    if resource_state == "sync":
        return "", 200
    import threading
    t = threading.Thread(target=_zpracuj_nove_faktury_z_drive)
    t.daemon = True
    t.start()
    return "", 200

def _zpracuj_nove_faktury_z_drive():
    """Stáhne nové PDF ze složky faktury-nahrat a zpracuje OCR."""
    import google.auth.transport.requests
    stats = {"stazeno": 0, "preskoceno": 0, "chyby": 0}
    print("DRIVE_START")
    print(f"DRIVE_FOLDER_ID={DRIVE_FOLDER_ID}")
    try:
        service, creds = get_drive_service()
        if not service:
            print("⚠ Drive: service není dostupný")
            return stats
        print("✓ Drive: service OK, načítám soubory")
        result = service.files().list(
            q=f"'{DRIVE_FOLDER_ID}' in parents and mimeType='application/pdf' and trashed=false",
            orderBy="createdTime desc",
            fields="files(id,name,createdTime)",
            pageSize=20
        ).execute()
        files = result.get("files", [])
        print(f"✓ Drive: nalezeno {len(files)} souborů ve složce")

        with get_db() as conn:
            try:
                conn.execute("""CREATE TABLE IF NOT EXISTS drive_zpracovane (
                    id SERIAL PRIMARY KEY, file_id TEXT UNIQUE, zpracovano_at TEXT)""")
            except: pass
            rows = conn.execute("SELECT file_id FROM drive_zpracovane").fetchall()
            # OPRAVA: funguje pro PostgreSQL (dict) i SQLite (tuple)
            zpracovane = {r["file_id"] if isinstance(r, dict) else r[0] for r in rows}

        for f in files:
            if f["id"] in zpracovane:
                print(f"⏭ Drive: přeskakuji {f['name']} (již zpracováno)")
                stats["preskoceno"] += 1
                continue
            print(f"📥 Drive: stahuji {f['name']} ({f['id']})")
            try:
                # Obnovit token před každým stažením
                creds.refresh(google.auth.transport.requests.Request())
                print(f"DEBUG_TOKEN: {creds.token[:20] if creds.token else 'PRAZDNY'}")
                import requests as _req
                headers = {"Authorization": f"Bearer {creds.token}"}
                dl_url = f"https://www.googleapis.com/drive/v3/files/{f['id']}?alt=media"
                resp = _req.get(dl_url, headers=headers, timeout=60)
                if resp.status_code != 200:
                    print(f"⚠ Drive stahování chyba {resp.status_code}: {resp.text[:200]}")
                    continue
                content = resp.content
                if not content:
                    print(f"⚠ Drive: prázdný obsah pro {f['name']}, přeskakuji")
                    continue
                safe_name = f["name"].replace(" ", "_")
                ts = __import__("datetime").datetime.now().strftime("%Y%m%d_%H%M%S_")
                fname = ts + safe_name
                fpath = os.path.join(UPLOAD_DIR, fname)
                with open(fpath, "wb") as fh:
                    fh.write(content)
                print(f"✓ Drive: soubor uložen jako {fname}")
                gcs_url = upload_to_gcs(fpath, f"faktury/{fname}")
                ocr_data = _ocr_faktura(fpath)
                print(f"✓ Drive: OCR dokončeno pro {fname}")
                with get_db() as conn:
                    # Kontrola duplicity podle čísla faktury - jen označíme, nesmažeme
                    cislo = ocr_data.get("cislo_faktury", "")
                    duplicita_id = None
                    if cislo:
                        dup = conn.execute(
                        "SELECT id FROM faktury WHERE cislo_faktury=? AND dodavatel LIKE ?",
                        (cislo, "%MAKRO%")
                    ).fetchone()
                    if dup:
                        print(f"⏭ Drive: přeskakuji duplicitu č. {cislo}, již existuje")
                        conn.execute("INSERT INTO drive_zpracovane (file_id, zpracovano_at) VALUES (?,?)",
                            (f["id"], __import__("datetime").datetime.now().isoformat()))
                        stats["preskoceno"] += 1
                        continue
                    conn.execute("""
                        INSERT INTO faktury (firma_zkratka, dodavatel, cislo_faktury,
                            datum_vystaveni, datum_splatnosti, celkem_s_dph,
                            stav, soubor_cesta, soubor_url, zdroj)
                        VALUES (?,?,?,?,?,?,?,?,?,?)
                    """, (
                        ocr_data.get("firma_zkratka", ""),
                        ocr_data.get("dodavatel", ""),
                        ocr_data.get("cislo_faktury", ""),
                        ocr_data.get("datum_vystaveni", ""),
                        ocr_data.get("datum_splatnosti", ""),
                        float(ocr_data.get("celkem_s_dph", 0)),
                        "zaplaceno",
                        fname,
                        gcs_url or "",
                        "drive_auto"
                    ))
                    if duplicita_id:
                        try:
                            conn.execute("UPDATE faktury SET duplicita_id=? WHERE soubor_cesta=?", (duplicita_id, fname))
                        except Exception:
                            pass
                    fid = conn.execute("SELECT id FROM faktury WHERE soubor_cesta=? ORDER BY id DESC LIMIT 1", (fname,)).fetchone()
                    if fid:
                        fid_val = fid["id"] if isinstance(fid, dict) else fid[0]
                        for p in ocr_data.get("polozky", []):
                            nazev = (p.get("nazev") or "").strip()
                            if not nazev: continue
                            conn.execute("""
                                INSERT INTO polozky (faktura_id, nazev, mnozstvi, jednotka,
                                    cena_za_jednotku_s_dph, celkem_s_dph)
                                VALUES (?,?,?,?,?,?)
                            """, (
                                fid_val,
                                nazev,
                                float(p.get("mnozstvi") or 1),
                                p.get("jednotka") or "ks",
                                float(p.get("cena_za_jednotku_s_dph") or 0),
                                float(p.get("celkem_s_dph") or 0),
                            ))
                    conn.execute(
                        "INSERT INTO drive_zpracovane (file_id, zpracovano_at) VALUES (?,?)",
                        (f["id"], __import__("datetime").datetime.now().isoformat())
                    )
                print(f"✅ Drive auto: zpracována FA {fname}")
                stats["stazeno"] += 1
            except Exception as e:
                import traceback
                print(f"⚠ Drive auto error pro {f['name']}: {e}")
                print(traceback.format_exc())
                stats["chyby"] += 1
    except Exception as e:
        import traceback
        print(f"⚠ Drive webhook error: {e}")
        print(traceback.format_exc())
    return stats

def _ocr_faktura(fpath):
    """OCR faktury — vrátí dict s daty."""
    api_key = os.environ.get("ANTHROPIC_API_KEY", "")
    if not api_key:
        return {}
    try:
        ext = fpath.rsplit(".", 1)[-1].lower()
        with open(fpath, "rb") as fh:
            raw = fh.read()
        b64 = base64.standard_b64encode(raw).decode("utf-8")
        if ext == "pdf":
            block = {"type": "document", "source": {"type": "base64", "media_type": "application/pdf", "data": b64}}
        else:
            mt = {"jpg":"image/jpeg","jpeg":"image/jpeg","png":"image/png"}.get(ext,"image/jpeg")
            block = {"type": "image", "source": {"type": "base64", "media_type": mt, "data": b64}}
        client = anthropic.Anthropic(api_key=api_key)
        ico_map = json.loads(os.environ.get("ICO_MAP_JSON", "{}"))
        msg = client.messages.create(
            model="claude-sonnet-4-20250514", max_tokens=2000,
            messages=[{"role": "user", "content": [block, {"type": "text", "text": f"""Analyzuj tuto MAKRO fakturu (daňový doklad).
Odpověz POUZE platným JSON, žádný jiný text.

Důležité pro číslo faktury: hledej POUZE pole "Faktura č. / VS" — hodnota je číslo ve formátu 0415000291 (10 číslic, pouze číslice). IGNORUJ číslo vpravo nahoře (formát 0015/0135 — to je číslo stránky), IGNORUJ "č. zákazníka" a IGNORUJ "Technické ID".

Důležité pro položky — KRITICKÁ PRAVIDLA:
1. MAKRO faktura má SEKCE (nadpisy kategorií) a pod nimi POLOŽKY ZBOŽÍ. Jsou to RŮZNÉ věci.
2. SEKCE poznáš takto: jsou to řádky POUZE s textem jako "OVOCE A ZELENINA", "MASO A DRŮBEŽ", "ZPRACOVANÉ MASO", "MRAŽENÉ POTRAVINY", "MLÉČNÉ VÝROBKY", "CHLÉB A PEČIVO", "JEDLÉ POTRAVINY KOLONIAL", "DROGERIE", "NÁPOJE" atd. Nemají číslo zboží ani cenu. SEKCE VŮBEC NEZAPISUJ do položek.
3. POLOŽKA ZBOŽÍ poznáš takto: má číslo zboží (dlouhé číslo např. 8435409365610 nebo kratší jako 24251105), má jednotku (PC/KG/CA/BAL), má cenu. Název položky je JEN samotný název zboží (např. "JABLKA ČER.BAL. TAŠKA 1kg") BEZ názvu sekce.
4. NIKDY nespojuj název sekce s názvem zboží. "OVOCE A ZELENINA" je sekce, "JABLKA ČER.BAL." je zboží — jsou to DVA oddělené řádky, zboží je vždy jen to druhé.
5. Řádky "Urceno pro konecnou spotrebu" jsou slevy — IGNORUJ jako položku.
6. Řádky "KUP VÍCE = PLAŤ MÉNĚ" jsou slevy — IGNORUJ jako položku.
7. IGNORUJ: "Strana celkem", "Poslední strana celkem", "Celková částka", "Spotřební daň celkem", "Platba kartou".
8. Zpracuj položky ze VŠECH stran faktury.

{{
  "dodavatel": "název dodavatele (obvykle MAKRO Cash & Carry ČR s.r.o.)",
  "cislo_faktury": "číslo POUZE z pole Faktura č. / VS (10 číslic, např. 0415000291)",
  "datum_vystaveni": "YYYY-MM-DD nebo null",
  "datum_splatnosti": "YYYY-MM-DD nebo null",
  "celkem_s_dph": číslo (celková částka včetně DPH),
  "ico_odberatele": "IČO odběratele nebo null",
  "polozky": [
    {{
      "nazev": "název zboží (BEZ názvu sekce/kategorie) — např. pokud vidíš řádek 'OVOCE A ZELENINA' a pak řádek 'JABLKA ČER.BAL. TAŠKA 1kg', název je POUZE 'JABLKA ČER.BAL. TAŠKA 1kg'",
      "mnozstvi": číslo,
      "jednotka": "PC/CA/KG atd.",
      "cena_za_jednotku_s_dph": číslo,
      "celkem_s_dph": číslo
    }}
  ]
}}
Známá IČO firem: {json.dumps(ico_map)}"""
            }]}]
        )
        text = msg.content[0].text.strip()
        text = re.sub(r"^```json\s*", "", text); text = re.sub(r"```$", "", text).strip()
        parsed = json.loads(text)
        ico_odb = parsed.get("ico_odberatele", "")
        firma = ico_map.get(str(ico_odb), "")
        parsed["firma_zkratka"] = firma
        return parsed
    except Exception as e:
        print(f"⚠ OCR error: {e}")
        return {}


@app.route("/api/auto-zaloha", methods=["POST", "GET"])
def api_auto_zaloha():
    """Endpoint pro Cloud Scheduler — automatická noční záloha."""
    import datetime as _dt_mod, psycopg2 as _pg
    # Ověření že volání přichází z Cloud Scheduler nebo má správný token
    token = request.headers.get("X-Zaloha-Token", "") or request.args.get("token", "")
    ocekavany = os.environ.get("ZALOHA_TOKEN", "auto-zaloha-2026")
    if token != ocekavany:
        return jsonify({"error": "Neautorizováno"}), 403

    ts = _dt_mod.datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"zaloha_{ts}.sql"
    try:
        db_url = os.environ.get("DATABASE_URL", "")
        conn = _pg.connect(db_url)
        cur = conn.cursor()
        cur.execute("SELECT tablename FROM pg_tables WHERE schemaname='public' ORDER BY tablename")
        tables = [r[0] for r in cur.fetchall()]
        lines = ["-- SQL záloha vygenerovaná automaticky", f"-- Datum: {ts}", ""]
        for tbl in tables:
            try:
                cur.execute(f"SELECT * FROM {tbl}")
                rows = cur.fetchall()
                cols = [d[0] for d in cur.description]
                if not rows: continue
                lines.append(f"-- Tabulka: {tbl}")
                for row in rows:
                    vals = []
                    for v in row:
                        if v is None: vals.append("NULL")
                        elif isinstance(v, bool): vals.append("TRUE" if v else "FALSE")
                        elif isinstance(v, (int, float)): vals.append(str(v))
                        else: vals.append("'" + str(v).replace("'", "''") + "'")
                    lines.append(f"INSERT INTO {tbl} ({', '.join(cols)}) VALUES ({', '.join(vals)}) ON CONFLICT DO NOTHING;")
                lines.append("")
            except Exception as e:
                lines.append(f"-- Chyba: {tbl}: {e}")
        conn.close()
        sql_data = "\n".join(lines).encode("utf-8")
    except Exception as e:
        return jsonify({"error": f"DB chyba: {str(e)}"}), 500

    # Uložit do GCS a smazat staré (ponechat posledních 5)
    try:
        bucket = get_gcs_client()
        if not bucket:
            return jsonify({"error": "GCS není dostupný"}), 500
        blob = bucket.blob(f"zalohy/{filename}")
        blob.upload_from_string(sql_data, content_type="application/sql")

        # Smaž zálohy starší než posledních 5
        vsechny = sorted(
            [b for b in bucket.list_blobs(prefix="zalohy/") if b.name.endswith(".sql")],
            key=lambda b: b.updated, reverse=True
        )
        smazano = 0
        for stara in vsechny[5:]:
            stara.delete()
            smazano += 1

        return jsonify({"ok": True, "soubor": filename, "smazano_starych": smazano})
    except Exception as e:
        return jsonify({"error": f"GCS chyba: {str(e)}"}), 500

@app.route("/api/zaloha-db")
@vyzaduj_prihlaseni
def api_zaloha_db():
    if session.get("role") != "admin":
        return jsonify({"error": "Pouze admin"}), 403
    import os as _os
    import datetime as _dt_mod
    ts = _dt_mod.datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"zaloha_{ts}.sql"
    try:
        import psycopg2 as _pg
        db_url = _os.environ.get("DATABASE_URL", "")
        if not db_url:
            return jsonify({"error": "DATABASE_URL není nastavena"}), 500
        conn = _pg.connect(db_url)
        cur = conn.cursor()
        cur.execute("SELECT tablename FROM pg_tables WHERE schemaname='public' ORDER BY tablename")
        tables = [r[0] for r in cur.fetchall()]
        lines = ["-- SQL záloha vygenerovaná aplikací", f"-- Datum: {ts}", ""]
        for tbl in tables:
            try:
                cur.execute(f"SELECT * FROM {tbl}")
                rows = cur.fetchall()
                cols = [d[0] for d in cur.description]
                if not rows:
                    continue
                lines.append(f"-- Tabulka: {tbl}")
                for row in rows:
                    vals = []
                    for v in row:
                        if v is None:
                            vals.append("NULL")
                        elif isinstance(v, bool):
                            vals.append("TRUE" if v else "FALSE")
                        elif isinstance(v, (int, float)):
                            vals.append(str(v))
                        else:
                            vals.append("'" + str(v).replace("'", "''") + "'")
                    col_str = ", ".join(cols)
                    val_str = ", ".join(vals)
                    lines.append(f"INSERT INTO {tbl} ({col_str}) VALUES ({val_str}) ON CONFLICT DO NOTHING;")
                lines.append("")
            except Exception as e:
                lines.append(f"-- Chyba při záloze tabulky {tbl}: {e}")
        conn.close()
        sql_data = "\n".join(lines).encode("utf-8")
    except Exception as e:
        return jsonify({"error": f"Záloha selhala: {str(e)}"}), 500

    gcs_url = None
    try:
        bucket = get_gcs_client()
        if bucket:
            blob = bucket.blob(f"zalohy/{filename}")
            blob.upload_from_string(sql_data, content_type="application/sql")
            gcs_url = f"gs://{os.environ.get('GCS_BUCKET_NAME','')}/zalohy/{filename}"
            print(f"✅ Záloha uložena do GCS: {gcs_url}")
    except Exception as e:
        print(f"⚠  GCS záloha error: {e}")

    from flask import Response
    resp = Response(
        sql_data,
        mimetype="application/sql",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )
    if gcs_url:
        resp.headers["X-GCS-URL"] = gcs_url
    return resp


@app.route("/api/admin/zaloha-export", methods=["POST"])
@vyzaduj_prihlaseni
def api_zaloha_export():
    import datetime as _dt, json as _json
    tabulky = [
        "faktury", "polozky", "reporty", "vydaje", "vyplaty",
        "zbozi", "kalkulace", "kalkulace_polozky", "stat_rucni_data",
        "pausalni_odvody", "bankovni_pohyby"
    ]
    export = {"datum": _dt.datetime.now().isoformat(), "tabulky": {}}
    with get_db() as conn:
        for t in tabulky:
            try:
                rows = conn.execute(f"SELECT * FROM {t}").fetchall()
                export["tabulky"][t] = [dict(r) for r in rows]
            except Exception:
                export["tabulky"][t] = []
    datum_str = _dt.datetime.now().strftime("%Y%m%d_%H%M")
    filename = f"zaloha_{datum_str}.json"
    json_bytes = _json.dumps(export, ensure_ascii=False, default=str).encode("utf-8")
    # Uložit do GCS
    bucket = get_gcs_client()
    if bucket:
        blob = bucket.blob(f"zalohy/{filename}")
        blob.upload_from_string(json_bytes, content_type="application/json")
        return jsonify({"ok": True, "soubor": filename, "ulozeno": "gcs"})
    else:
        # Fallback – stáhnout přímo
        from flask import Response
        return Response(
            json_bytes,
            status=200,
            mimetype="application/json",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )

@app.route("/api/admin/zalohy")
@vyzaduj_prihlaseni
def api_zalohy_seznam():
    bucket = get_gcs_client()
    if not bucket:
        return jsonify({"zalohy": [], "error": "GCS není nakonfigurováno"})
    blobs = sorted(bucket.list_blobs(prefix="zalohy/"), key=lambda b: b.updated, reverse=True)
    result = []
    for b in blobs:
        if b.name.endswith(".json"):
            result.append({
                "nazev": b.name.replace("zalohy/", ""),
                "velikost": b.size,
                "datum": b.updated.isoformat() if b.updated else "",
                "url": b.generate_signed_url(expiration=3600) if hasattr(b, 'generate_signed_url') else ""
            })
    return jsonify({"zalohy": result[:3]})
@app.route("/api/admin/zaloha-stahnout/<nazev>")
@vyzaduj_prihlaseni
def api_zaloha_stahnout(nazev):
    from flask import Response
    bucket = get_gcs_client()
    if not bucket:
        return jsonify({"error": "GCS není nakonfigurováno"}), 500
    blob = bucket.blob(f"zalohy/{nazev}")
    data = blob.download_as_bytes()
    return Response(
        data,
        status=200,
        mimetype="application/json",
        headers={"Content-Disposition": f"attachment; filename={nazev}"}
    )

@app.route("/api/reset-drive-zpracovane", methods=["POST"])
@vyzaduj_prihlaseni
def api_reset_drive_zpracovane():
    if session.get("role") != "admin":
        return jsonify({"error": "Pouze admin"}), 403
    with get_db() as conn:
        conn.execute("DELETE FROM drive_zpracovane")
    return jsonify({"ok": True})
@app.route("/api/makro-zaplaceno", methods=["POST"])
@vyzaduj_prihlaseni
def api_makro_zaplaceno():
    """Zpětně označí všechny MAKRO faktury jako zaplaceno."""
    if session.get("role") != "admin":
        return jsonify({"error": "Pouze admin"}), 403
    try:
        import psycopg2 as _pg2
        db_url = os.environ.get("DATABASE_URL", "")
        conn = _pg2.connect(db_url)
        cur = conn.cursor()
        cur.execute("""
            UPDATE faktury SET stav='zaplaceno', datum_zaplaceno=datum_vystaveni
            WHERE dodavatel ILIKE '%MAKRO%'
            AND stav NOT IN ('zaplaceno','duplikat')
        """)
        pocet = cur.rowcount
        conn.commit()
        conn.close()
        return jsonify({"ok": True, "oznaceno": pocet})
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route("/api/smazat-vse-faktury", methods=["POST"])
@vyzaduj_prihlaseni
def api_smazat_vse_faktury():
    with get_db() as conn:
        conn.execute("DELETE FROM polozky")
        cur = conn.execute("DELETE FROM faktury")
        smazano = cur.rowcount if hasattr(cur, 'rowcount') else 0
    return jsonify({"ok": True, "smazano": smazano})
@app.route("/api/normalizuj-dodavatele", methods=["POST"])
@vyzaduj_prihlaseni
def api_normalizuj_dodavatele():
    if session.get("role") != "admin":
        return jsonify({"error": "Pouze admin"}), 403
    opravy = [
        ("MAKRO Cash & Carry CR s.r.o.", "MAKRO Cash & Carry ČR s.r.o."),
        ("MAKRO Cash&Carry ČR s.r.o.", "MAKRO Cash & Carry ČR s.r.o."),
        ("MAKRO Cash&Carry CR s.r.o.", "MAKRO Cash & Carry ČR s.r.o."),
    ]
    opraveno = 0
    with get_db() as conn:
        for spatne, spravne in opravy:
            cur = conn.execute(
                "UPDATE faktury SET dodavatel=? WHERE dodavatel=?",
                (spravne, spatne)
            )
            opraveno += cur.rowcount
    return jsonify({"ok": True, "opraveno": opraveno})
@app.route("/api/normalizuj-nazvy", methods=["POST"])
@vyzaduj_prihlaseni
def api_normalizuj_nazvy():
    import re as _re
    prefix_re = _re.compile(r'^(ARO|MC|FL)\s+', _re.IGNORECASE)
    prejmenovano = 0
    slouceno = 0
    with get_db() as conn:
        # Krok 1: Odeber prefixy ARO/MC/FL
        zbozi = conn.execute("SELECT id, nazev_canonical FROM zbozi").fetchall()
        for z in zbozi:
            nazev = (z["nazev_canonical"] if isinstance(z, dict) else z[1]) or ""
            novy = prefix_re.sub("", nazev).strip()
            if novy == nazev:
                continue
            zid = z["id"] if isinstance(z, dict) else z[0]
            existujici = conn.execute(
                "SELECT id FROM zbozi WHERE LOWER(nazev_canonical)=LOWER(%s) AND id!=%s",
                (novy, zid)
            ).fetchone()
            if existujici:
                cil_id = existujici["id"] if isinstance(existujici, dict) else existujici[0]
                conn.execute("UPDATE polozky SET zbozi_id=%s WHERE zbozi_id=%s", (cil_id, zid))
                conn.execute("DELETE FROM zbozi WHERE id=%s", (zid,))
                slouceno += 1
            else:
                conn.execute("UPDATE zbozi SET nazev_canonical=%s WHERE id=%s", (novy, zid))
                prejmenovano += 1

        # Krok 2: Slouč záznamy se stejným názvem (různé jednotky, různá velikost písmen)
        # Najdi skupiny duplicit podle LOWER(nazev_canonical)
        skupiny = conn.execute("""
            SELECT LOWER(nazev_canonical) as nazev_low, COUNT(*) as pocet, MIN(id) as zachovat_id
            FROM zbozi
            GROUP BY LOWER(nazev_canonical)
            HAVING COUNT(*) > 1
        """).fetchall()
        for sk in skupiny:
            nazev_low = sk["nazev_low"] if isinstance(sk, dict) else sk[0]
            zachovat_id = sk["zachovat_id"] if isinstance(sk, dict) else sk[2]
            # Najdi všechny duplicity kromě toho co zachováme
            duplikaty = conn.execute(
                "SELECT id FROM zbozi WHERE LOWER(nazev_canonical)=%s AND id!=%s",
                (nazev_low, zachovat_id)
            ).fetchall()
            for dup in duplikaty:
                dup_id = dup["id"] if isinstance(dup, dict) else dup[0]
                conn.execute("UPDATE polozky SET zbozi_id=%s WHERE zbozi_id=%s", (zachovat_id, dup_id))
                conn.execute("DELETE FROM zbozi WHERE id=%s", (dup_id,))
                slouceno += 1

    return jsonify({"ok": True, "prejmenovano": prejmenovano, "slouceno": slouceno})

@app.route("/api/oprav-duplicity", methods=["POST"])
@vyzaduj_prihlaseni
def api_oprav_duplicity():
    try:
        with get_db() as conn:
            faktury = conn.execute(
                "SELECT id, cislo_faktury, datum_vystaveni, celkem_s_dph FROM faktury ORDER BY id ASC"
            ).fetchall()

        opraveno = 0
        with get_db() as conn:
            for f in faktury:
                original = conn.execute(
                    """SELECT id FROM faktury
                       WHERE cislo_faktury = ? AND datum_vystaveni = ? AND celkem_s_dph = ?
                       AND id < ? AND (duplicita_id IS NULL OR duplicita_id = 0)
                       ORDER BY id ASC LIMIT 1""",
                    (f["cislo_faktury"], f["datum_vystaveni"], f["celkem_s_dph"], f["id"])
                ).fetchone()

                if original:
                    conn.execute(
                        "UPDATE faktury SET duplicita_id = ? WHERE id = ? AND (duplicita_id IS NULL OR duplicita_id = 0)",
                        (original["id"], f["id"])
                    )
                    opraveno += 1

        return jsonify({"ok": True, "opraveno": opraveno})
    except Exception as e:
        return jsonify({"ok": False, "chyba": str(e)}), 500


@app.route("/api/debug-duplicita/<int:fid1>/<int:fid2>")
@vyzaduj_prihlaseni
def api_debug_duplicita(fid1, fid2):
    import re as _re
    def _norm_nazev(n):
        n = str(n).strip().upper()
        n = _re.sub(r'^(ARO|MC|FL|CBA)\s+', '', n)
        n = _re.sub(r'\s+(KG|G|L|ML|KS|PC|BG|SW|CA)$', '', n)
        return n
    with get_db() as conn:
        f1 = dict(conn.execute("SELECT id, datum_vystaveni, celkem_s_dph FROM faktury WHERE id=?", (fid1,)).fetchone())
        f2 = dict(conn.execute("SELECT id, datum_vystaveni, celkem_s_dph FROM faktury WHERE id=?", (fid2,)).fetchone())
        pol1 = conn.execute("SELECT nazev, mnozstvi FROM polozky WHERE faktura_id=? ORDER BY nazev", (fid1,)).fetchall()
        pol2 = conn.execute("SELECT nazev, mnozstvi FROM polozky WHERE faktura_id=? ORDER BY nazev", (fid2,)).fetchall()
    norm1 = sorted([(_norm_nazev(p["nazev"] if isinstance(p,dict) else p[0]), float(p["mnozstvi"] if isinstance(p,dict) else p[1])) for p in pol1])
    norm2 = sorted([(_norm_nazev(p["nazev"] if isinstance(p,dict) else p[0]), float(p["mnozstvi"] if isinstance(p,dict) else p[1])) for p in pol2])
    return jsonify({
        "f1": f1, "f2": f2,
        "pocet1": len(pol1), "pocet2": len(pol2),
        "datum_ok": f1["datum_vystaveni"] == f2["datum_vystaveni"],
        "castka_ok": abs(float(f1["celkem_s_dph"]) - float(f2["celkem_s_dph"])) < 0.01,
        "polozky_ok": norm1 == norm2,
        "norm1": norm1, "norm2": norm2,
        "rozdily_1_nema_2": [x for x in norm1 if x not in norm2],
        "rozdily_2_nema_1": [x for x in norm2 if x not in norm1],
    })


@app.route("/api/oznac-obsahove-duplicity", methods=["POST"])
@vyzaduj_prihlaseni
def api_oznac_obsahove_duplicity():
    """Označí duplicitní faktury. Stačí splnit jedno z kritérií:
    A) datum + VS (číslo faktury)
    B) datum + částka (tol. 1 Kč) + položky
    C) datum + položky (bez ohledu na částku)
    D) datum + VS + částka
    """
    import re as _re

    def _norm_nazev(n):
        n = str(n).strip().upper()
        n = _re.sub(r'^(ARO|MC|FL|CBA)\s+', '', n)
        n = _re.sub(r'\s+(KG|G|L|ML|KS|PC|BG|SW|CA)$', '', n)
        return n

    def _norm_polozky(rows):
        result = []
        for p in rows:
            nazev = p["nazev"] if isinstance(p, dict) else p[0]
            mnozstvi = float(p["mnozstvi"] if isinstance(p, dict) else p[1])
            if _re.search(r'KUP\s+V[IÍ]CE|PLA[TŤ]\s+M[EÉ]N[EĚ]', nazev.upper()):
                continue
            result.append((_norm_nazev(nazev), mnozstvi))
        return sorted(result)

    try:
        oznaceno = 0
        with get_db() as conn:
            faktury = conn.execute(
                "SELECT id, datum_vystaveni, celkem_s_dph, cislo_faktury FROM faktury ORDER BY id ASC"
            ).fetchall()

            _pol_cache = {}
            def _get_polozky(fid):
                if fid not in _pol_cache:
                    rows = conn.execute(
                        "SELECT nazev, mnozstvi FROM polozky WHERE faktura_id=?", (fid,)
                    ).fetchall()
                    _pol_cache[fid] = _norm_polozky(rows)
                return _pol_cache[fid]

            def _oznac(fid, kid):
                conn.execute(
                    "UPDATE faktury SET duplicita_id=? WHERE id=? AND duplicita_id IS NULL",
                    (kid, fid)
                )

            for f in faktury:
                fid    = f["id"] if isinstance(f, dict) else f[0]
                datum  = f["datum_vystaveni"] if isinstance(f, dict) else f[1]
                castka = float(f["celkem_s_dph"] if isinstance(f, dict) else f[2])
                vs     = (f["cislo_faktury"] if isinstance(f, dict) else f[3] or "").strip()

                dup_row = conn.execute("SELECT duplicita_id FROM faktury WHERE id=?", (fid,)).fetchone()
                if dup_row and (dup_row["duplicita_id"] if isinstance(dup_row, dict) else dup_row[0]):
                    continue

                kandidati = conn.execute("""
                    SELECT id, celkem_s_dph, cislo_faktury FROM faktury
                    WHERE datum_vystaveni = ? AND id != ? AND duplicita_id IS NULL
                    ORDER BY id ASC
                """, (datum, fid)).fetchall()

                for k in kandidati:
                    kid     = k["id"] if isinstance(k, dict) else k[0]
                    kcastka = float(k["celkem_s_dph"] if isinstance(k, dict) else k[1])
                    kvs     = (k["cislo_faktury"] if isinstance(k, dict) else k[2] or "").strip()

                    castka_ok  = abs(castka - kcastka) < 1.0
                    vs_ok      = bool(vs and kvs and vs == kvs)
                    pol_f      = _get_polozky(fid)
                    pol_k      = _get_polozky(kid)
                    polozky_ok = bool(pol_f and pol_k and pol_f == pol_k)

                    if vs_ok or (castka_ok and polozky_ok) or polozky_ok or (vs_ok and castka_ok):
                        _oznac(fid, kid)
                        oznaceno += 1
                        break

        return jsonify({"ok": True, "oznaceno": oznaceno})
    except Exception as e:
        return jsonify({"ok": False, "chyba": str(e)}), 500


# ═══════════════════════════════════════════════════════════════
#  KALKULACE
# ═══════════════════════════════════════════════════════════════

@app.route("/api/kalkulace")
@vyzaduj_prihlaseni
def api_kalkulace_list():
    with get_db() as conn:
        rows = conn.execute("SELECT * FROM kalkulace ORDER BY nazev").fetchall()
        result = []
        for r in rows:
            d = dict(r)
            polozky = conn.execute(
                "SELECT * FROM kalkulace_polozky WHERE kalkulace_id=? ORDER BY id", (d["id"],)
            ).fetchall()
            d["polozky"] = [dict(p) for p in polozky]
            pausalni = conn.execute(
                "SELECT * FROM kalkulace_pausalni WHERE kalkulace_id=? ORDER BY id", (d["id"],)
            ).fetchall()
            d["pausalni"] = [dict(p) for p in pausalni]
            result.append(d)
    return jsonify(result)

@app.route("/api/kalkulace/<int:kid>")
@vyzaduj_prihlaseni
def api_kalkulace_get(kid):
    with get_db() as conn:
        r = conn.execute("SELECT * FROM kalkulace WHERE id=?", (kid,)).fetchone()
        if not r:
            return jsonify({"error": "Nenalezeno"}), 404
        d = dict(r)
        polozky = conn.execute(
            "SELECT * FROM kalkulace_polozky WHERE kalkulace_id=? ORDER BY id", (kid,)
        ).fetchall()
        d["polozky"] = [dict(p) for p in polozky]
        pausalni = conn.execute(
            "SELECT * FROM kalkulace_pausalni WHERE kalkulace_id=? ORDER BY id", (kid,)
        ).fetchall()
        d["pausalni"] = [dict(p) for p in pausalni]
    return jsonify(d)

@app.route("/api/kalkulace", methods=["POST"])
@vyzaduj_prihlaseni
def api_kalkulace_ulozit():
    data = request.json
    if not data.get("nazev"):
        return jsonify({"error": "Chybí název"}), 400
    polozky  = data.pop("polozky", [])
    pausalni = data.pop("pausalni", [])
    with get_db() as conn:
        cur = conn.execute("""
            INSERT INTO kalkulace (nazev, popis, prodejni_cena, cil_marze_pct, updated_at)
            VALUES (?,?,?,?,NOW())
        """, (data["nazev"], data.get("popis",""),
              float(data.get("prodejni_cena",0) or 0),
              float(data.get("cil_marze_pct",200) or 200)))
        kid = cur.lastrowid
        for p in polozky:
            conn.execute("""
                INSERT INTO kalkulace_polozky
                (kalkulace_id, nazev, mnozstvi, jednotka, cena_za_jednotku, je_baleni, baleni_ks, zdroj_ceny)
                VALUES (?,?,?,?,?,?,?,?)
            """, (kid, p.get("nazev",""), float(p.get("mnozstvi",1) or 1),
                  p.get("jednotka","ks"), float(p.get("cena_za_jednotku",0) or 0),
                  1 if p.get("je_baleni") else 0,
                  float(p.get("baleni_ks",1) or 1), p.get("zdroj_ceny","rucni")))
        for p in pausalni:
            conn.execute("INSERT INTO kalkulace_pausalni (kalkulace_id, nazev, castka) VALUES (?,?,?)",
                         (kid, p.get("nazev",""), float(p.get("castka",0) or 0)))
    return jsonify({"ok": True, "id": kid})

@app.route("/api/kalkulace/<int:kid>", methods=["PUT"])
@vyzaduj_prihlaseni
def api_kalkulace_edit(kid):
    data = request.json
    polozky  = data.pop("polozky", [])
    pausalni = data.pop("pausalni", [])
    with get_db() as conn:
        conn.execute("""
            UPDATE kalkulace SET nazev=?, popis=?, prodejni_cena=?, cil_marze_pct=?, updated_at=NOW()
            WHERE id=?
        """, (data.get("nazev",""), data.get("popis",""),
              float(data.get("prodejni_cena",0) or 0),
              float(data.get("cil_marze_pct",200) or 200), kid))
        conn.execute("DELETE FROM kalkulace_polozky WHERE kalkulace_id=?", (kid,))
        conn.execute("DELETE FROM kalkulace_pausalni WHERE kalkulace_id=?", (kid,))
        for p in polozky:
            conn.execute("""
                INSERT INTO kalkulace_polozky
                (kalkulace_id, nazev, mnozstvi, jednotka, cena_za_jednotku, je_baleni, baleni_ks, zdroj_ceny)
                VALUES (?,?,?,?,?,?,?,?)
            """, (kid, p.get("nazev",""), float(p.get("mnozstvi",1) or 1),
                  p.get("jednotka","ks"), float(p.get("cena_za_jednotku",0) or 0),
                  1 if p.get("je_baleni") else 0,
                  float(p.get("baleni_ks",1) or 1), p.get("zdroj_ceny","rucni")))
        for p in pausalni:
            conn.execute("INSERT INTO kalkulace_pausalni (kalkulace_id, nazev, castka) VALUES (?,?,?)",
                         (kid, p.get("nazev",""), float(p.get("castka",0) or 0)))
    return jsonify({"ok": True})

@app.route("/api/kalkulace/<int:kid>", methods=["DELETE"])
@vyzaduj_prihlaseni
def api_kalkulace_smazat(kid):
    with get_db() as conn:
        conn.execute("DELETE FROM kalkulace_polozky WHERE kalkulace_id=?", (kid,))
        conn.execute("DELETE FROM kalkulace_pausalni WHERE kalkulace_id=?", (kid,))
        conn.execute("DELETE FROM kalkulace WHERE id=?", (kid,))
    return jsonify({"ok": True})

@app.route("/api/kalkulace/cena-polozky")
@vyzaduj_prihlaseni
def api_kalkulace_cena_polozky():
    """Najde poslední cenu položky z faktur podle názvu (fuzzy match)."""
    nazev = request.args.get("nazev", "").strip()
    if not nazev:
        return jsonify({"cena": None, "zdroj": None})
    with get_db() as conn:
        # Hledáme v položkách faktur – poslední faktura kde se položka vyskytuje
        row = conn.execute("""
            SELECT p.cena_za_jednotku_s_dph, p.jednotka, f.datum_vystaveni, f.dodavatel
            FROM polozky p
            JOIN faktury f ON f.id = p.faktura_id
            WHERE LOWER(p.nazev) LIKE LOWER(?)
            ORDER BY f.datum_vystaveni DESC
            LIMIT 1
        """, (f"%{nazev}%",)).fetchone()
        if row:
            d = dict(row)
            return jsonify({
                "cena": float(d["cena_za_jednotku_s_dph"]),
                "jednotka": d["jednotka"],
                "datum": d["datum_vystaveni"],
                "dodavatel": d["dodavatel"],
                "zdroj": "faktura"
            })
    return jsonify({"cena": None, "zdroj": None})

@app.route("/api/oprav-sekvence")
@vyzaduj_prihlaseni
def api_oprav_sekvence():
    if session.get("role") != "admin":
        return jsonify({"error": "Pouze admin"}), 403
    if not _USE_PG:
        return jsonify({"error": "Pouze PostgreSQL"}), 400
    vysledky = {}
    for tbl in ["vystavene_faktury", "faktury", "reporty", "vyplaty", "vydaje", "bankovni_pohyby", "zbozi", "polozky"]:
        try:
            with get_db() as conn:
                conn.execute(f"CREATE SEQUENCE IF NOT EXISTS {tbl}_id_seq")
                conn.execute(f"ALTER TABLE {tbl} ALTER COLUMN id SET DEFAULT nextval('{tbl}_id_seq')")
                conn.execute(f"SELECT setval('{tbl}_id_seq', COALESCE((SELECT MAX(id) FROM {tbl}), 0) + 1, false)")
            vysledky[tbl] = "OK"
        except Exception as e:
            vysledky[tbl] = str(e)
    return jsonify(vysledky)

@app.route("/fix-prava-seq")
def fix_prava_seq():
    if not _USE_PG:
        return "Pouze pro PostgreSQL"
    try:
        with get_db() as conn:
            conn.execute("CREATE SEQUENCE IF NOT EXISTS prava_id_seq")
            conn.execute("ALTER TABLE prava ALTER COLUMN id SET DEFAULT nextval('prava_id_seq')")
            conn.execute("SELECT setval('prava_id_seq', COALESCE((SELECT MAX(id) FROM prava), 0) + 1)")
        return "OK - sekvence opravena"
    except Exception as e:
        return f"Chyba: {e}"
@app.route("/ping")
def ping():
    return "pong", 200

@app.route("/debug-ocr")
def debug_ocr():
    """Dočasný debug endpoint - testuje OCR na poslední nahraté FA"""
    if not OCR_SUPPORT or not PDF_SUPPORT:
        return f"OCR: {OCR_SUPPORT}, PDF: {PDF_SUPPORT}", 200
    try:
        import glob, io
        # Najít poslední PDF v uploads
        pdfs = sorted(glob.glob(os.path.join(UPLOAD_DIR, "*.pdf")))
        if not pdfs:
            return "Žádné PDF nenalezeno", 200
        filepath = pdfs[-1]
        result_lines = [f"Soubor: {os.path.basename(filepath)}"]
        with pdfplumber.open(filepath) as pdf:
            result_lines.append(f"Stránek: {len(pdf.pages)}")
            page = pdf.pages[-1]
            pil = page.to_image(resolution=150).original
            text = pytesseract.image_to_string(pil, lang="ces+eng")
            result_lines.append("--- Řádky s klíčovými slovy ---")
            for line in text.splitlines():
                if any(kw in line.lower() for kw in ["celkov", "castka", "strana", "celkem"]):
                    result_lines.append(repr(line))
        return "<br>".join(result_lines), 200
    except Exception as e:
        return f"Chyba: {e}", 500


def _ensure_pg_columns():
    """Přidá chybějící sloupce do PostgreSQL tabulek."""
    try:
        import psycopg2 as _pg2
        db_url = os.environ.get("DATABASE_URL", "")
        if not db_url:
            return
        conn = _pg2.connect(db_url)
        cur = conn.cursor()
        migrace = [
            "ALTER TABLE vydaje ADD COLUMN IF NOT EXISTS var_sym TEXT DEFAULT ''",
            "ALTER TABLE vydaje ADD COLUMN IF NOT EXISTS datum_zaplaceno TEXT DEFAULT ''",
            "ALTER TABLE faktury ADD COLUMN IF NOT EXISTS var_sym TEXT DEFAULT ''",
            "ALTER TABLE faktury ADD COLUMN IF NOT EXISTS datum_zaplaceno TEXT DEFAULT ''",
            "ALTER TABLE vystavene_faktury ADD COLUMN IF NOT EXISTS var_sym TEXT DEFAULT ''",
            "ALTER TABLE vystavene_faktury ADD COLUMN IF NOT EXISTS datum_zaplaceno TEXT DEFAULT ''",
            "ALTER TABLE bankovni_pohyby ADD COLUMN IF NOT EXISTS var_sym TEXT DEFAULT ''",
            "ALTER TABLE bankovni_pohyby ADD COLUMN IF NOT EXISTS sparovano INTEGER DEFAULT 0",
            "ALTER TABLE bankovni_pohyby ADD COLUMN IF NOT EXISTS sparovano_typ TEXT DEFAULT ''",
            "ALTER TABLE bankovni_pohyby ADD COLUMN IF NOT EXISTS sparovano_id INTEGER DEFAULT NULL",
        ]
        for sql in migrace:
            try:
                cur.execute(sql)
            except Exception:
                conn.rollback()
        conn.commit()
        conn.close()
    except Exception as e:
        print(f"⚠ _ensure_pg_columns: {e}")

# Spustit migrace při startu
try:
    _ensure_pg_columns()
except Exception:
    pass



# ═══════════════════════════════════════════════════════════════
#  PENĚŽENKA — hotovostní kasa
# ═══════════════════════════════════════════════════════════════

PENEZENKA_START = "2026-03-24"

@app.route("/api/eur-kurz")
@vyzaduj_prihlaseni
def api_eur_kurz():
    """Vrátí aktuální kurz EUR/CZK z CNB."""
    import urllib.request as _ur
    try:
        req = _ur.Request("https://api.cnb.cz/cnbapi/exrates/daily?lang=EN",
                          headers={"User-Agent": "faktury-makro/1.0"})
        with _ur.urlopen(req, timeout=5) as resp:
            data = json.loads(resp.read())
        eur = next((r for r in data.get("rates", []) if r["currencyCode"] == "EUR"), None)
        if eur:
            return jsonify({"kurz": round(eur["rate"] / eur["amount"], 4)})
    except Exception as e:
        app.logger.warning(f"CNB kurz chyba: {e}")
    return jsonify({"kurz": 25.0})


@app.route("/api/usd-kurz")
@vyzaduj_prihlaseni
def api_usd_kurz():
    """Vrátí aktuální kurz USD/CZK z CNB."""
    import urllib.request as _ur
    try:
        req = _ur.Request("https://api.cnb.cz/cnbapi/exrates/daily?lang=EN",
                          headers={"User-Agent": "faktury-makro/1.0"})
        with _ur.urlopen(req, timeout=5) as resp:
            data = json.loads(resp.read())
        usd = next((r for r in data.get("rates", []) if r["currencyCode"] == "USD"), None)
        if usd:
            return jsonify({"kurz": round(usd["rate"] / usd["amount"], 4)})
    except Exception as e:
        app.logger.warning(f"CNB USD kurz chyba: {e}")
    return jsonify({"kurz": 23.0})

BANKY_SLOUPCE = ["rb_fp","rb_mr","rb_cff","rb_radek","air_fp","air_mr","air_cff","air_radek","kb_radek"]

@app.route("/api/penezenka")
@vyzaduj_prihlaseni
def api_penezenka_list():
    import datetime as _dt
    dnes = _dt.date.today().isoformat()
    with get_db() as conn:
        zaznamy = conn.execute("SELECT * FROM penezenka ORDER BY datum DESC").fetchall()
        r_hot = conn.execute("""
            SELECT COALESCE(SUM(hotovost), 0) as hot, COALESCE(SUM(vydaje), 0) as vyd
            FROM reporty WHERE datum >= ? AND datum <= ?
        """, (PENEZENKA_START, dnes)).fetchone()
        hot = float(r_hot["hot"] if isinstance(r_hot, dict) else r_hot[0])
        vyd = float(r_hot["vyd"] if isinstance(r_hot, dict) else r_hot[1])
    return jsonify({
        "zaznamy": [dict(r) for r in zaznamy],
        "teoreticky_stav": round(hot - vyd, 0),
        "hotovost_celkem": round(hot, 0),
        "vydaje_celkem": round(vyd, 0),
        "od_data": PENEZENKA_START,
    })

@app.route("/api/penezenka", methods=["POST"])
@vyzaduj_prihlaseni
def api_penezenka_ulozit():
    data = request.json or {}
    datum = data.get("datum", "")
    if not datum:
        return jsonify({"error": "Chybí datum"}), 400
    import json as _json
    import traceback as _tb
    try:
        # Zjistit aktuální sloupce v tabulce
        with get_db() as conn:
            if _USE_PG:
                cur = conn.execute("SELECT column_name FROM information_schema.columns WHERE table_name='penezenka'")
                pen_cols = [r["column_name"] for r in cur.fetchall()]
            else:
                pen_cols = [row[1] for row in conn.execute("PRAGMA table_info(penezenka)").fetchall()]

            # Přidat chybějící sloupce za běhu
            for col in ["hotovost","rb_fp","rb_mr","rb_cff","rb_radek","air_fp","air_mr","air_cff","air_radek","kb_radek","xtb_czk","xtb_eur","t212","etoro","sporeni"]:
                if col not in pen_cols:
                    try:
                        conn.execute(f"ALTER TABLE penezenka ADD COLUMN {col} REAL DEFAULT 0")
                        pen_cols.append(col)
                    except Exception: pass
            if "extras" not in pen_cols:
                try:
                    conn.execute("ALTER TABLE penezenka ADD COLUMN extras TEXT DEFAULT '[]'")
                    pen_cols.append("extras")
                except Exception: pass

            # INSERT pouze se sloupci které existují
            cols = ["datum"]
            vals = [datum]
            for col in ["hotovost","rb_fp","rb_mr","rb_cff","rb_radek","air_fp","air_mr","air_cff","air_radek","kb_radek","xtb_czk","xtb_eur","t212","etoro","sporeni"]:
                if col in pen_cols:
                    cols.append(col)
                    vals.append(float(data.get(col, 0) or 0))
            if "extras" in pen_cols:
                cols.append("extras")
                vals.append(_json.dumps(data.get("extras", []), ensure_ascii=False))
            if "poznamka" in pen_cols:
                cols.append("poznamka")
                vals.append(data.get("poznamka", ""))

            placeholders = ",".join(["?"] * len(cols))
            cur = conn.execute(
                f"INSERT INTO penezenka ({','.join(cols)}) VALUES ({placeholders})",
                vals
            )
        return jsonify({"ok": True, "id": cur.lastrowid})
    except Exception as e:
        app.logger.error(f"Penezenka save error: {_tb.format_exc()}")
        return jsonify({"error": str(e)}), 500

@app.route("/api/penezenka/<int:pid>", methods=["PUT"])
@vyzaduj_prihlaseni
def api_penezenka_edit(pid):
    import json as _json, traceback as _tb
    data = request.json or {}
    try:
        with get_db() as conn:
            if _USE_PG:
                cur = conn.execute("SELECT column_name FROM information_schema.columns WHERE table_name='penezenka'")
                pen_cols = [r["column_name"] for r in cur.fetchall()]
            else:
                pen_cols = [row[1] for row in conn.execute("PRAGMA table_info(penezenka)").fetchall()]
            cols = ["datum"]
            vals = [data.get("datum","")]
            for col in ["hotovost","rb_fp","rb_mr","rb_cff","rb_radek","air_fp","air_mr","air_cff","air_radek","kb_radek","xtb_czk","xtb_eur","t212","etoro","sporeni"]:
                if col in pen_cols:
                    cols.append(col)
                    vals.append(float(data.get(col,0) or 0))
            if "extras" in pen_cols:
                cols.append("extras")
                vals.append(_json.dumps(data.get("extras",[]), ensure_ascii=False))
            if "poznamka" in pen_cols:
                cols.append("poznamka")
                vals.append(data.get("poznamka",""))
            set_clause = ", ".join(f"{c}=?" for c in cols)
            vals.append(pid)
            conn.execute(f"UPDATE penezenka SET {set_clause} WHERE id=?", vals)
        return jsonify({"ok": True})
    except Exception as e:
        app.logger.error(f"Penezenka edit error: {_tb.format_exc()}")
        return jsonify({"error": str(e)}), 500

@app.route("/api/penezenka/<int:pid>", methods=["DELETE"])
@vyzaduj_prihlaseni
def api_penezenka_delete(pid):
    with get_db() as conn:
        conn.execute("DELETE FROM penezenka WHERE id=?", (pid,))
    return jsonify({"ok": True})



# ═══════════════════════════════════════════════════════════════
#  DLUHY — půjčky kamarádům
# ═══════════════════════════════════════════════════════════════

@app.route("/api/dluhy")
@vyzaduj_prihlaseni
def api_dluhy_list():
    with get_db() as conn:
        osoby = conn.execute("SELECT * FROM dluhy_osoby ORDER BY jmeno").fetchall()
        result = []
        for o in osoby:
            oid = o["id"] if isinstance(o, dict) else o[0]
            jmeno = o["jmeno"] if isinstance(o, dict) else o[1]
            poznamka = o["poznamka"] if isinstance(o, dict) else o[2]
            transakce = conn.execute(
                "SELECT * FROM dluhy_transakce WHERE osoba_id=? ORDER BY datum ASC, id ASC", (oid,)
            ).fetchall()
            trans_list = [dict(t) for t in transakce]
            celkem = sum(float(t["castka"] if isinstance(t, dict) else t[2]) for t in transakce)
            prvni = trans_list[0]["datum"] if trans_list else None
            result.append({
                "id": oid, "jmeno": jmeno, "poznamka": poznamka,
                "celkem": round(celkem, 0),
                "prvni_pujcka": prvni,
                "transakce": trans_list,
            })
    return jsonify(result)

@app.route("/api/dluhy/osoby", methods=["POST"])
@vyzaduj_prihlaseni
def api_dluhy_nova_osoba():
    data = request.json or {}
    jmeno = (data.get("jmeno") or "").strip()
    if not jmeno:
        return jsonify({"error": "Chybí jméno"}), 400
    with get_db() as conn:
        cur = conn.execute(
            "INSERT INTO dluhy_osoby (jmeno, poznamka) VALUES (?,?)",
            (jmeno, data.get("poznamka",""))
        )
    return jsonify({"ok": True, "id": cur.lastrowid})

@app.route("/api/dluhy/transakce", methods=["POST"])
@vyzaduj_prihlaseni
def api_dluhy_nova_transakce():
    data = request.json or {}
    osoba_id = data.get("osoba_id")
    datum    = data.get("datum","")
    castka   = float(data.get("castka", 0) or 0)
    if not osoba_id or not datum:
        return jsonify({"error": "Chybí osoba nebo datum"}), 400
    with get_db() as conn:
        cur = conn.execute(
            "INSERT INTO dluhy_transakce (osoba_id, datum, castka, poznamka) VALUES (?,?,?,?)",
            (osoba_id, datum, castka, data.get("poznamka",""))
        )
    return jsonify({"ok": True, "id": cur.lastrowid})

@app.route("/api/dluhy/transakce/<int:tid>", methods=["DELETE"])
@vyzaduj_prihlaseni
def api_dluhy_smazat_transakci(tid):
    with get_db() as conn:
        conn.execute("DELETE FROM dluhy_transakce WHERE id=?", (tid,))
    return jsonify({"ok": True})

@app.route("/api/dluhy/osoby/<int:oid>", methods=["DELETE"])
@vyzaduj_prihlaseni
def api_dluhy_smazat_osobu(oid):
    with get_db() as conn:
        conn.execute("DELETE FROM dluhy_transakce WHERE osoba_id=?", (oid,))
        conn.execute("DELETE FROM dluhy_osoby WHERE id=?", (oid,))
    return jsonify({"ok": True})


# ════════════════════════════════════════════════════════════════
#  TRVALÉ PŘÍKAZY
# ════════════════════════════════════════════════════════════════

@app.route("/api/trvale-prikazy")
@vyzaduj_prihlaseni
def api_trvale_prikazy_seznam():
    try:
        conn = psycopg2.connect(os.environ["DATABASE_URL"])
        cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
        cur.execute("SELECT * FROM trvale_prikazy ORDER BY den_v_mesici, dodavatel")
        rows = cur.fetchall()
        cur.close()
        conn.close()
        return jsonify({"prikazy": [dict(r) for r in rows]})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/api/trvale-prikazy", methods=["POST"])
@vyzaduj_prihlaseni
def api_trvale_prikazy_ulozit():
    data = request.json or {}
    try:
        conn = psycopg2.connect(os.environ["DATABASE_URL"])
        cur = conn.cursor()
        cur.execute("""
            INSERT INTO trvale_prikazy
                (lokace, dodavatel, popis, zpusob_uhrady, castka, den_v_mesici, aktivni, poznamka)
            VALUES (%s,%s,%s,%s,%s,%s,%s,%s)
            RETURNING id
        """, (
            data.get("lokace", ""),
            data.get("dodavatel", ""),
            data.get("popis", ""),
            data.get("zpusob_uhrady", "převodem"),
            float(data.get("castka", 0)),
            int(data.get("den_v_mesici", 1)),
            1,
            data.get("poznamka", ""),
        ))
        nid = cur.fetchone()[0]
        conn.commit()
        cur.close()
        conn.close()
        return jsonify({"ok": True, "id": nid})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/api/trvale-prikazy/<int:pid>", methods=["PUT"])
@vyzaduj_prihlaseni
def api_trvale_prikazy_edit(pid):
    data = request.json or {}
    try:
        conn = psycopg2.connect(os.environ["DATABASE_URL"])
        cur = conn.cursor()
        cur.execute("""
            UPDATE trvale_prikazy SET
                lokace=%s, dodavatel=%s, popis=%s, zpusob_uhrady=%s,
                castka=%s, den_v_mesici=%s, aktivni=%s, poznamka=%s
            WHERE id=%s
        """, (
            data.get("lokace", ""),
            data.get("dodavatel", ""),
            data.get("popis", ""),
            data.get("zpusob_uhrady", "převodem"),
            float(data.get("castka", 0)),
            int(data.get("den_v_mesici", 1)),
            int(data.get("aktivni", 1)),
            data.get("poznamka", ""),
            pid,
        ))
        conn.commit()
        cur.close()
        conn.close()
        return jsonify({"ok": True})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/api/trvale-prikazy/<int:pid>", methods=["DELETE"])
@vyzaduj_prihlaseni
def api_trvale_prikazy_smazat(pid):
    try:
        conn = psycopg2.connect(os.environ["DATABASE_URL"])
        cur = conn.cursor()
        cur.execute("DELETE FROM trvale_prikazy WHERE id=%s", (pid,))
        conn.commit()
        cur.close()
        conn.close()
        return jsonify({"ok": True})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/api/trvale-prikazy/generovat", methods=["POST"])
@vyzaduj_prihlaseni
def api_trvale_prikazy_generovat():
    """Vygeneruje trvalé příkazy jako výdaje pro zadaný rok/měsíc (pokud ještě neexistují)."""
    data = request.json or {}
    rok  = int(data.get("rok",  date.today().year))
    mes  = int(data.get("mesic", date.today().month))
    try:
        conn = psycopg2.connect(os.environ["DATABASE_URL"])
        cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
        cur.execute("SELECT * FROM trvale_prikazy WHERE aktivni=1")
        prikazy = cur.fetchall()
        vygenerovano = 0
        preskoceno   = 0
        for p in prikazy:
            import calendar as _cal
            posledni_den = _cal.monthrange(rok, mes)[1]
            den = min(p["den_v_mesici"], posledni_den)
            datum_str = f"{rok:04d}-{mes:02d}-{den:02d}"
            # Zkontrolovat zda výdaj pro tento trvalý příkaz v tomto měsíci již existuje
            cur2 = conn.cursor()
            cur2.execute("""
                SELECT id FROM vydaje
                WHERE zdroj='trvaly_prikaz'
                  AND poznamka=%s
                  AND datum>=%s AND datum<=%s
            """, (
                f"TP:{p['id']}",
                f"{rok:04d}-{mes:02d}-01",
                f"{rok:04d}-{mes:02d}-{posledni_den:02d}",
            ))
            existing = cur2.fetchone()
            cur2.close()
            if existing:
                preskoceno += 1
                continue
            cur3 = conn.cursor()
            cur3.execute("""
                INSERT INTO vydaje
                    (firma_zkratka, dodavatel, datum, castka, zpusob_uhrady,
                     stav, popis, poznamka, zdroj, typ, stitky)
                VALUES (%s,%s,%s,%s,%s,'nezaplaceno',%s,%s,'trvaly_prikaz','soukrome','🔁 trvalý příkaz')
            """, (
                p["lokace"] or "_soukrome",
                p["dodavatel"],
                datum_str,
                p["castka"],
                p["zpusob_uhrady"],
                p["popis"],
                f"TP:{p['id']}",
            ))
            cur3.close()
            vygenerovano += 1
        conn.commit()
        conn.close()
        return jsonify({"ok": True, "vygenerovano": vygenerovano, "preskoceno": preskoceno})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


if __name__ == "__main__":
    print("=" * 55)
    print("  Správa faktur – spouštím server")
    print("  Otevři prohlížeč na: http://localhost:5000")
    print("=" * 55)
    app.run(host="0.0.0.0", port=5000, debug=False)

# ════════════════════════════════════════════════════════════════
#  DOKUMENTY
# ════════════════════════════════════════════════════════════════

def init_dokumenty():
    with get_db() as conn:
        conn.execute("""CREATE TABLE IF NOT EXISTS dokumenty (
            id          SERIAL PRIMARY KEY,
            datum       TEXT NOT NULL,
            nazev       TEXT NOT NULL,
            misto       TEXT DEFAULT 'Praha',
            kategorie   TEXT DEFAULT '',
            soubor_cesta TEXT DEFAULT '',
            soubor_url  TEXT DEFAULT '',
            created_at  TEXT DEFAULT NOW()
        )""")
        # Přidat sloupec kategorie pokud ještě neexistuje (migrace)
        try:
            conn.execute("ALTER TABLE dokumenty ADD COLUMN kategorie TEXT DEFAULT ''")
        except Exception:
            pass

init_dokumenty()

def upload_dokument_to_gcs(local_path, filename):
    bucket = get_gcs_client()
    if not bucket:
        return None
    try:
        blob = bucket.blob(f"dokumenty/{filename}")
        blob.upload_from_filename(local_path)
        url = blob.generate_signed_url(expiration=timedelta(days=7), method="GET", version="v4")
        return url
    except Exception as e:
        print(f"⚠  GCS dokumenty upload error: {e}")
        return None

def get_dokument_gcs_url(filename):
    bucket = get_gcs_client()
    if not bucket:
        return None
    try:
        blob = bucket.blob(f"dokumenty/{filename}")
        if not blob.exists():
            return None
        return blob.generate_signed_url(expiration=timedelta(days=7), method="GET", version="v4")
    except Exception as e:
        print(f"⚠  GCS dokumenty url error: {e}")
        return None

@app.route("/api/dokumenty")
@vyzaduj_prihlaseni
def api_dokumenty_list():
    with get_db() as conn:
        rows = conn.execute("SELECT * FROM dokumenty ORDER BY datum DESC, id DESC").fetchall()
    return jsonify([dict(r) for r in rows])

@app.route("/api/dokumenty", methods=["POST"])
@vyzaduj_prihlaseni
def api_dokumenty_create():
    datum     = request.form.get("datum", "")
    nazev     = request.form.get("nazev", "").strip()
    misto     = request.form.get("misto", "Praha")
    kategorie = request.form.get("kategorie", "").strip()
    if not nazev:
        return jsonify({"chyba": "Chybí název"}), 400
    soubor_cesta = ""
    soubor_url   = ""
    if "soubor" in request.files:
        f = request.files["soubor"]
        if f and f.filename:
            fname = secure_filename(f.filename)
            ts    = datetime.now().strftime("%Y%m%d_%H%M%S")
            fname = f"{ts}_{fname}"
            fpath = os.path.join(UPLOAD_DIR, fname)
            f.save(fpath)
            url = upload_dokument_to_gcs(fpath, fname)
            soubor_cesta = fname
            soubor_url   = url or ""
    with get_db() as conn:
        cur = conn.execute(
            "INSERT INTO dokumenty (datum, nazev, misto, kategorie, soubor_cesta, soubor_url) VALUES (?,?,?,?,?,?)",
            (datum, nazev, misto, kategorie, soubor_cesta, soubor_url)
        )
    return jsonify({"ok": True, "id": cur.lastrowid})

@app.route("/api/dokumenty/<int:did>", methods=["PUT"])
@vyzaduj_prihlaseni
def api_dokumenty_update(did):
    d = request.json or {}
    with get_db() as conn:
        conn.execute(
            "UPDATE dokumenty SET datum=?, nazev=?, misto=?, kategorie=? WHERE id=?",
            (d.get("datum",""), d.get("nazev",""), d.get("misto","Praha"), d.get("kategorie",""), did)
        )
    return jsonify({"ok": True})

@app.route("/api/dokumenty/<int:did>", methods=["DELETE"])
@vyzaduj_prihlaseni
def api_dokumenty_delete(did):
    with get_db() as conn:
        conn.execute("DELETE FROM dokumenty WHERE id=?", (did,))
    return jsonify({"ok": True})

@app.route("/api/dokumenty/<int:did>/url")
@vyzaduj_prihlaseni
def api_dokumenty_url(did):
    with get_db() as conn:
        row = conn.execute("SELECT soubor_cesta, soubor_url FROM dokumenty WHERE id=?", (did,)).fetchone()
    if not row:
        return jsonify({"chyba": "Nenalezeno"}), 404
    url = get_dokument_gcs_url(row["soubor_cesta"]) or row["soubor_url"] or ""
    return jsonify({"url": url})
    


